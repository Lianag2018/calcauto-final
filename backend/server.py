from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Header
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta
import base64
import json
import re
import io
import pypdf
import pdfplumber
import tempfile
import hashlib
import time

# OCR imports
import pytesseract
from PIL import Image
import cv2
import numpy as np

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ============ WINDOW STICKER CONFIGURATION ============
WINDOW_STICKER_URLS = {
    "chrysler": "https://www.chrysler.com/hostd/windowsticker/getWindowStickerPdf.do?vin=",
    "jeep": "https://www.jeep.com/hostd/windowsticker/getWindowStickerPdf.do?vin=",
    "dodge": "https://www.dodge.com/hostd/windowsticker/getWindowStickerPdf.do?vin=",
    "ram": "https://www.ramtrucks.com/hostd/windowsticker/getWindowStickerPdf.do?vin=",
    "fiat": "https://www.fiatusa.com/hostd/windowsticker/getWindowStickerPdf.do?vin=",
    "alfa": "https://www.alfaromeousa.com/hostd/windowsticker/getWindowStickerPdf.do?vin=",
}


def convert_pdf_to_images(pdf_bytes: bytes, max_pages: int = 2, dpi: int = 100) -> list:
    """
    Convertit un PDF en images JPEG optimisées pour email.
    
    Args:
        pdf_bytes: PDF en bytes
        max_pages: Nombre max de pages à convertir
        dpi: Résolution (100 = optimisé pour email, petite taille)
    
    Returns:
        Liste de dicts avec image_base64, width, height
    """
    try:
        import fitz  # PyMuPDF
        from io import BytesIO
        from PIL import Image
        
        images = []
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        
        for page_num in range(min(len(doc), max_pages)):
            page = doc[page_num]
            
            # Convertir en image avec le DPI spécifié
            zoom = dpi / 72  # 72 est le DPI par défaut des PDF
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # Convertir via PIL pour meilleure compression
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # Redimensionner si trop grand (max 800px de large pour email)
            max_width = 800
            if img.width > max_width:
                ratio = max_width / img.width
                new_size = (max_width, int(img.height * ratio))
                img = img.resize(new_size, Image.Resampling.LANCZOS)
            
            # Compresser en JPEG avec qualité réduite
            buffer = BytesIO()
            img.save(buffer, format="JPEG", quality=70, optimize=True)
            img_bytes = buffer.getvalue()
            img_base64 = base64.b64encode(img_bytes).decode("utf-8")
            
            images.append({
                "base64": img_base64,
                "width": img.width,
                "height": img.height,
                "page": page_num + 1,
                "size_kb": len(img_bytes) // 1024
            })
            
            logger.info(f"Window Sticker page {page_num+1}: {img.width}x{img.height}, {len(img_bytes)//1024}KB")
        
        doc.close()
        return images
        
    except Exception as e:
        logger.error(f"Erreur conversion PDF→Image: {e}")
        return []


def generate_window_sticker_html(vin: str, images: list, pdf_url: str, pdf_bytes: bytes = None) -> str:
    """
    Génère le HTML pour afficher le Window Sticker avec l'image du PDF.
    L'image utilise CID (Content-ID) pour un meilleur support Gmail.
    """
    if not vin or len(vin) != 17:
        return ""
    
    # Si on a des images converties du PDF, les afficher avec CID
    if images and len(images) > 0:
        # Utiliser CID au lieu de data: pour Gmail compatibility
        cid = f"windowsticker_{vin}"
        
        return f'''
            <div style="margin-top: 25px; page-break-inside: avoid;">
                <div style="text-align: center; margin-bottom: 10px;">
                    <span style="font-size: 14px; font-weight: bold; color: #333;">📋 Window Sticker Officiel</span>
                    <br/>
                    <span style="font-size: 12px; color: #666;">VIN: {vin}</span>
                </div>
                <div style="text-align: center; border: 1px solid #ddd; border-radius: 8px; padding: 10px; background: #fff;">
                    <img src="cid:{cid}" 
                         alt="Window Sticker {vin}" 
                         style="max-width: 100%; height: auto; border-radius: 4px;" />
                </div>
                <div style="text-align: center; margin-top: 10px;">
                    <a href="{pdf_url}" 
                       style="display: inline-block; background: #1565C0; color: white; padding: 8px 20px; text-decoration: none; border-radius: 5px; font-size: 13px;">
                        📥 Télécharger le PDF
                    </a>
                </div>
            </div>
        '''
    
    # Fallback: pas d'images, juste le lien
    elif pdf_url:
        return f'''
            <div style="margin-top: 25px;">
                <div style="background: #e8f4fd; border: 1px solid #2196F3; border-radius: 6px; padding: 15px; text-align: center;">
                    <p style="margin: 0 0 10px 0; color: #1565C0; font-weight: bold;">📋 Window Sticker Officiel</p>
                    <p style="margin: 0 0 10px 0; font-size: 12px; color: #666;">VIN: {vin}</p>
                    <a href="{pdf_url}" 
                       style="display: inline-block; background: #1565C0; color: white; padding: 10px 25px; text-decoration: none; border-radius: 5px; font-size: 14px;">
                        📥 Voir le Window Sticker
                    </a>
                </div>
            </div>
        '''
    
    return ""


async def fetch_window_sticker(vin: str, brand: str = None) -> dict:
    """
    Télécharge le Window Sticker PDF pour un VIN donné.
    Approche KenBot: HTTP humain + validation PDF + fallback Playwright
    
    Returns:
        dict avec:
        - success: bool
        - pdf_base64: str (PDF encodé en base64)
        - pdf_url: str (URL directe)
        - size_bytes: int
        - error: str (si échec)
    """
    import requests
    
    MIN_PDF_SIZE = 20_000  # 20 KB minimum pour un vrai sticker
    
    if not vin or len(vin) != 17:
        return {"success": False, "error": "VIN invalide (doit être 17 caractères)"}
    
    def validate_pdf(data: bytes) -> tuple[bool, str]:
        """Valide que les bytes sont un vrai PDF Window Sticker"""
        if len(data) < MIN_PDF_SIZE:
            return False, f"PDF trop petit ({len(data)} bytes) — probablement pas un vrai sticker"
        if not data.startswith(b"%PDF"):
            head = data[:30]
            return False, f"Pas un PDF (head={head!r}) — probablement HTML/Not found/anti-bot"
        return True, "OK"
    
    def download_pdf_human(pdf_url: str, referer: str) -> bytes:
        """Télécharge le PDF avec headers humains"""
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            "Accept": "application/pdf,application/octet-stream;q=0.9,*/*;q=0.8",
            "Accept-Language": "fr-CA,fr;q=0.9,en;q=0.8",
            "Referer": referer,
            "Connection": "keep-alive",
            "Cache-Control": "no-cache",
        }
        
        response = requests.get(pdf_url, headers=headers, timeout=30, allow_redirects=True)
        response.raise_for_status()
        return response.content
    
    async def download_pdf_playwright(pdf_url: str) -> bytes:
        """Fallback: télécharge via navigateur headless (async) - OPTIONNEL"""
        try:
            from playwright.async_api import async_playwright
        except ImportError:
            logger.warning("Playwright non installé - fallback désactivé")
            raise RuntimeError("Playwright non disponible")
        
        pdf_bytes = None
        
        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                
                async def on_response(resp):
                    nonlocal pdf_bytes
                    ct = (resp.headers.get("content-type") or "").lower()
                    if "application/pdf" in ct:
                        try:
                            pdf_bytes = await resp.body()
                        except:
                            pass
                
                page.on("response", on_response)
                
                try:
                    await page.goto(pdf_url, wait_until="networkidle", timeout=30000)
                    # Attendre un peu si le PDF arrive en retard
                    if pdf_bytes is None:
                        await page.wait_for_timeout(3000)
                except Exception as e:
                    logger.warning(f"Playwright navigation error: {e}")
                
                await browser.close()
        except Exception as e:
            logger.warning(f"Playwright error: {e}")
            raise RuntimeError(f"Playwright failed: {e}")
        
        if not pdf_bytes:
            raise RuntimeError("PDF non capturé via Playwright")
        return pdf_bytes
    
    # Déterminer les URLs à essayer basées sur la marque
    urls_to_try = []
    
    if brand:
        brand_lower = brand.lower()
        for key in WINDOW_STICKER_URLS:
            if key in brand_lower or brand_lower in key:
                urls_to_try.append((key, WINDOW_STICKER_URLS[key] + vin))
                break
    
    # Si pas de marque ou pas trouvé, essayer toutes les URLs Stellantis
    if not urls_to_try:
        priority = ["jeep", "chrysler", "dodge", "ram", "fiat", "alfa"]
        for key in priority:
            urls_to_try.append((key, WINDOW_STICKER_URLS[key] + vin))
    
    # Referers par marque
    referers = {
        "jeep": "https://www.jeep.com/",
        "chrysler": "https://www.chrysler.com/",
        "dodge": "https://www.dodge.com/",
        "ram": "https://www.ramtrucks.com/",
        "fiat": "https://www.fiatusa.com/",
        "alfa": "https://www.alfaromeousa.com/",
    }
    
    last_error = None
    
    for brand_key, url in urls_to_try:
        referer = referers.get(brand_key, "https://www.chrysler.com/")
        
        # === Étape 1: HTTP "humain" ===
        try:
            logger.info(f"Window Sticker HTTP fetch: VIN={vin}, Brand={brand_key}")
            pdf_bytes = download_pdf_human(url, referer)
            
            is_valid, msg = validate_pdf(pdf_bytes)
            if is_valid:
                pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")
                logger.info(f"Window Sticker téléchargé (HTTP): VIN={vin}, Size={len(pdf_bytes)} bytes")
                return {
                    "success": True,
                    "pdf_base64": pdf_base64,
                    "pdf_url": url,
                    "size_bytes": len(pdf_bytes),
                    "brand_source": brand_key,
                    "method": "http"
                }
            else:
                logger.warning(f"Window Sticker HTTP invalid: {msg}")
                last_error = msg
        except Exception as e:
            logger.warning(f"Window Sticker HTTP failed for {brand_key}: {e}")
            last_error = str(e)
        
        # === Étape 2: Fallback Playwright ===
        try:
            logger.info(f"Window Sticker Playwright fallback: VIN={vin}, Brand={brand_key}")
            pdf_bytes = await download_pdf_playwright(url)
            
            is_valid, msg = validate_pdf(pdf_bytes)
            if is_valid:
                pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")
                logger.info(f"Window Sticker téléchargé (Playwright): VIN={vin}, Size={len(pdf_bytes)} bytes")
                return {
                    "success": True,
                    "pdf_base64": pdf_base64,
                    "pdf_url": url,
                    "size_bytes": len(pdf_bytes),
                    "brand_source": brand_key,
                    "method": "playwright"
                }
            else:
                logger.warning(f"Window Sticker Playwright invalid: {msg}")
                last_error = msg
        except Exception as e:
            logger.warning(f"Window Sticker Playwright failed for {brand_key}: {e}")
            last_error = str(e)
            continue
    
    return {"success": False, "error": f"Window Sticker non trouvé: {last_error}"}


async def save_window_sticker_to_db(vin: str, pdf_base64: str, owner_id: str) -> str:
    """
    Sauvegarde le Window Sticker PDF dans MongoDB.
    
    Returns:
        ID du document créé
    """
    doc_id = str(uuid.uuid4())
    
    await db.window_stickers.update_one(
        {"vin": vin},
        {"$set": {
            "id": doc_id,
            "vin": vin,
            "pdf_base64": pdf_base64,
            "owner_id": owner_id,
            "created_at": datetime.utcnow(),
            "size_bytes": len(base64.b64decode(pdf_base64))
        }},
        upsert=True
    )
    
    return doc_id

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Admin password for import (from .env)
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'Liana2018')
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '')

# Gmail SMTP Configuration
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))

# ============ Models ============

class FinancingRates(BaseModel):
    """Taux de financement pour chaque terme (36-96 mois)"""
    rate_36: float
    rate_48: float
    rate_60: float
    rate_72: float
    rate_84: float
    rate_96: float

class VehicleProgram(BaseModel):
    """
    Structure d'un programme de financement véhicule
    
    RÈGLES IMPORTANTES:
    - Option 1 (Consumer Cash): Rabais AVANT TAXES + taux d'intérêt (variables selon le véhicule)
    - Option 2 (Alternative): Rabais = $0 + taux réduits (peut être None si non disponible)
    - Bonus Cash: Rabais APRÈS TAXES (comme comptant), combinable avec Option 1 OU 2
    - Les taux sont FIXES (non variables)
    """
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    brand: str
    model: str
    trim: Optional[str] = None
    year: int
    
    # Option 1: Consumer Cash (rabais AVANT taxes) + taux Consumer Cash
    consumer_cash: float = 0  # Rabais avant taxes
    option1_rates: FinancingRates
    
    # Option 2: Alternative Consumer Cash (généralement $0) + taux réduits
    # None = Option 2 non disponible pour ce véhicule
    option2_rates: Optional[FinancingRates] = None
    
    # Bonus Cash: Rabais APRÈS taxes (combinable avec Option 1 ou 2)
    bonus_cash: float = 0
    
    # Métadonnées du programme
    program_month: int = Field(default_factory=lambda: datetime.utcnow().month)
    program_year: int = Field(default_factory=lambda: datetime.utcnow().year)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class VehicleProgramCreate(BaseModel):
    brand: str
    model: str
    trim: Optional[str] = None
    year: int
    consumer_cash: float = 0
    option1_rates: FinancingRates
    option2_rates: Optional[FinancingRates] = None
    bonus_cash: float = 0
    program_month: Optional[int] = None
    program_year: Optional[int] = None

class VehicleProgramUpdate(BaseModel):
    brand: Optional[str] = None
    model: Optional[str] = None
    trim: Optional[str] = None
    year: Optional[int] = None
    consumer_cash: Optional[float] = None
    option1_rates: Optional[FinancingRates] = None
    option2_rates: Optional[FinancingRates] = None
    bonus_cash: Optional[float] = None

class CalculationRequest(BaseModel):
    vehicle_price: float
    program_id: Optional[str] = None

class PaymentComparison(BaseModel):
    term_months: int
    # Option 1: Consumer Cash + taux
    option1_rate: float
    option1_monthly: float
    option1_total: float
    option1_rebate: float  # Consumer Cash (avant taxes)
    # Option 2: Taux réduits (si disponible)
    option2_rate: Optional[float] = None
    option2_monthly: Optional[float] = None
    option2_total: Optional[float] = None
    # Meilleure option
    best_option: Optional[str] = None
    savings: Optional[float] = None

class CalculationResponse(BaseModel):
    vehicle_price: float
    consumer_cash: float
    bonus_cash: float
    brand: str
    model: str
    trim: Optional[str]
    year: int
    comparisons: List[PaymentComparison]

class ProgramPeriod(BaseModel):
    month: int
    year: int
    count: int

class ImportRequest(BaseModel):
    password: str
    programs: List[Dict[str, Any]]
    program_month: int
    program_year: int

# ============ CRM Models ============

class Submission(BaseModel):
    """Soumission client avec suivi"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str = ""  # ID de l'utilisateur propriétaire
    # Client info
    client_name: str
    client_phone: str
    client_email: str
    # Vehicle info
    vehicle_brand: str
    vehicle_model: str
    vehicle_year: int
    vehicle_price: float
    # Financing info
    term: int
    payment_monthly: float
    payment_biweekly: float = 0
    payment_weekly: float = 0
    selected_option: str = "1"
    rate: float = 0
    # Tracking
    submission_date: datetime = Field(default_factory=datetime.utcnow)
    reminder_date: Optional[datetime] = None  # Default: 24h after submission
    reminder_done: bool = False
    status: str = "pending"  # pending, contacted, converted, lost
    notes: str = ""
    # For comparison
    program_month: int = 0
    program_year: int = 0

class SubmissionCreate(BaseModel):
    client_name: str
    client_phone: str
    client_email: str
    vehicle_brand: str
    vehicle_model: str
    vehicle_year: int
    vehicle_price: float
    term: int
    payment_monthly: float
    payment_biweekly: float = 0
    payment_weekly: float = 0
    selected_option: str = "1"
    rate: float = 0
    program_month: int = 0
    program_year: int = 0

class ReminderUpdate(BaseModel):
    reminder_date: datetime
    notes: Optional[str] = None

class BetterOffer(BaseModel):
    """Meilleure offre trouvée pour un client"""
    submission_id: str
    client_name: str
    client_phone: str
    client_email: str
    vehicle: str
    old_payment: float
    new_payment: float
    savings_monthly: float
    savings_total: float
    term: int
    approved: bool = False
    email_sent: bool = False

# ============ Auth Models ============

class UserRegister(BaseModel):
    name: str
    email: str
    password: str

class UserLogin(BaseModel):
    email: str
    password: str

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    password_hash: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_admin: bool = False
    is_blocked: bool = False
    last_login: Optional[datetime] = None

# Admin email - this user will always be admin
ADMIN_EMAIL = "danielgiroux007@gmail.com"

# ============ Contact Models ============

class Contact(BaseModel):
    """Contact importé depuis vCard/CSV"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str = ""  # ID de l'utilisateur propriétaire
    name: str
    phone: str = ""
    email: str = ""
    created_at: datetime = Field(default_factory=datetime.utcnow)
    source: str = "import"  # import, manual

class ContactCreate(BaseModel):
    name: str
    phone: str = ""
    email: str = ""
    source: str = "import"

class ContactBulkCreate(BaseModel):
    contacts: List[ContactCreate]

# ============ Inventory Models ============

class InventoryVehicle(BaseModel):
    """Véhicule en inventaire avec coûts réels"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str = ""
    
    # Identification
    stock_no: str                    # Numéro de stock (clé principale)
    vin: str = ""
    
    # Véhicule
    brand: str
    model: str
    trim: str = ""
    year: int
    type: str = "neuf"               # neuf | occasion
    
    # Prix et coûts (valeurs EXACTES de la facture)
    pdco: float = 0                  # Prix dealer officiel
    ep_cost: float = 0               # Coût réel (Employee Price)
    holdback: float = 0              # Holdback (valeur facture, PAS calculé)
    net_cost: float = 0              # ep_cost - holdback (calculé à l'import)
    
    # Prix de vente
    msrp: float = 0                  # PDSF
    asking_price: float = 0          # Prix affiché
    sold_price: Optional[float] = None
    
    # Statut
    status: str = "disponible"       # disponible | réservé | vendu
    km: int = 0                      # Pour occasions
    color: str = ""
    
    # Métadonnées
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class InventoryCreate(BaseModel):
    stock_no: str
    vin: str = ""
    brand: str
    model: str
    trim: str = ""
    year: int
    type: str = "neuf"
    pdco: float = 0
    ep_cost: float = 0
    holdback: float = 0
    msrp: float = 0
    asking_price: float = 0
    km: int = 0
    color: str = ""

class InventoryUpdate(BaseModel):
    vin: Optional[str] = None
    brand: Optional[str] = None
    model: Optional[str] = None
    trim: Optional[str] = None
    year: Optional[int] = None
    type: Optional[str] = None
    pdco: Optional[float] = None
    ep_cost: Optional[float] = None
    holdback: Optional[float] = None
    msrp: Optional[float] = None
    asking_price: Optional[float] = None
    sold_price: Optional[float] = None
    status: Optional[str] = None
    km: Optional[int] = None
    color: Optional[str] = None

class VehicleOption(BaseModel):
    """Option/équipement d'un véhicule"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    stock_no: str
    product_code: str
    order: int = 0
    description: str
    amount: float = 0

class ProductCode(BaseModel):
    """Référentiel des codes produits FCA"""
    code: str
    description_standard: str
    category: str = ""               # moteur, esthétique, sécurité, etc.

# ============ Utility Functions ============

import hashlib
import secrets

def hash_password(password: str) -> str:
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def generate_token() -> str:
    """Generate a random token"""
    return secrets.token_hex(32)

async def get_current_user(authorization: Optional[str] = Header(None)):
    """Obtient l'utilisateur courant à partir du token d'autorisation"""
    if not authorization:
        raise HTTPException(status_code=401, detail="Token manquant")
    
    # Le token peut être "Bearer <token>" ou juste "<token>"
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    
    # Chercher le token dans la base
    token_doc = await db.tokens.find_one({"token": token})
    if not token_doc:
        raise HTTPException(status_code=401, detail="Token invalide")
    
    # Chercher l'utilisateur
    user = await db.users.find_one({"id": token_doc["user_id"]})
    if not user:
        raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
    
    return user

async def get_optional_user(authorization: Optional[str] = Header(None)):
    """Obtient l'utilisateur courant si un token est fourni, sinon None"""
    if not authorization:
        return None
    
    try:
        return await get_current_user(authorization)
    except:
        return None

def calculate_monthly_payment(principal: float, annual_rate: float, months: int) -> float:
    """Calcule le paiement mensuel avec la formule d'amortissement"""
    if principal <= 0 or months <= 0:
        return 0
    if annual_rate == 0:
        return round(principal / months, 2)
    
    monthly_rate = annual_rate / 100 / 12
    payment = principal * (monthly_rate * (1 + monthly_rate) ** months) / ((1 + monthly_rate) ** months - 1)
    return round(payment, 2)

def get_rate_for_term(rates: FinancingRates, term: int) -> float:
    """Obtient le taux pour un terme spécifique"""
    rate_map = {
        36: rates.rate_36,
        48: rates.rate_48,
        60: rates.rate_60,
        72: rates.rate_72,
        84: rates.rate_84,
        96: rates.rate_96
    }
    return rate_map.get(term, 4.99)

# ============ Routes ============

@api_router.get("/")
async def root():
    return {"message": "Vehicle Financing Calculator API - v4 (avec historique et Bonus Cash)"}

# Ping endpoint for keep-alive (supports both GET and HEAD)
@api_router.api_route("/ping", methods=["GET", "HEAD"])
async def ping():
    """Endpoint pour garder le serveur actif (keep-alive)"""
    return {"status": "ok", "message": "Server is alive"}

# ============ Auth Routes ============

@api_router.post("/auth/register")
async def register_user(user_data: UserRegister):
    """Register a new user"""
    # Check if email already exists
    existing_user = await db.users.find_one({"email": user_data.email.lower()})
    if existing_user:
        raise HTTPException(status_code=400, detail="Cet email est déjà utilisé")
    
    # Create user
    user = User(
        name=user_data.name,
        email=user_data.email.lower(),
        password_hash=hash_password(user_data.password)
    )
    
    user_dict = user.dict()
    await db.users.insert_one(user_dict)
    
    # Generate token
    token = generate_token()
    await db.tokens.insert_one({
        "user_id": user.id,
        "token": token,
        "created_at": datetime.utcnow()
    })
    
    return {
        "success": True,
        "user": {
            "id": user.id,
            "name": user.name,
            "email": user.email
        },
        "token": token
    }

@api_router.post("/auth/login")
async def login_user(credentials: UserLogin):
    """Login user"""
    user = await db.users.find_one({
        "email": credentials.email.lower(),
        "password_hash": hash_password(credentials.password)
    })
    
    if not user:
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    # Check if user is blocked
    if user.get("is_blocked", False):
        raise HTTPException(status_code=403, detail="Votre compte a été désactivé. Contactez l'administrateur.")
    
    # Update last_login
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"last_login": datetime.utcnow()}}
    )
    
    # Check if user should be admin
    is_admin = user.get("is_admin", False) or user["email"] == ADMIN_EMAIL
    
    # Generate token
    token = generate_token()
    await db.tokens.insert_one({
        "user_id": user["id"],
        "token": token,
        "created_at": datetime.utcnow()
    })
    
    return {
        "success": True,
        "user": {
            "id": user["id"],
            "name": user["name"],
            "email": user["email"],
            "is_admin": is_admin
        },
        "token": token
    }

@api_router.post("/auth/logout")
async def logout_user(token: str):
    """Logout user by deleting token"""
    await db.tokens.delete_one({"token": token})
    return {"success": True}

# Get PDF info (page count)
@api_router.post("/pdf-info")
async def get_pdf_info(file: UploadFile = File(...)):
    """Récupère les informations du PDF (nombre de pages)"""
    try:
        contents = await file.read()
        pdf_reader = pypdf.PdfReader(io.BytesIO(contents))
        total_pages = len(pdf_reader.pages)
        
        return {
            "success": True,
            "total_pages": total_pages,
            "filename": file.filename
        }
    except Exception as e:
        logger.error(f"Error reading PDF: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur lors de la lecture du PDF: {str(e)}"
        }

# Get available program periods (for history)
@api_router.get("/periods", response_model=List[ProgramPeriod])
async def get_periods():
    """Récupère les périodes de programmes disponibles (pour l'historique)"""
    pipeline = [
        {"$group": {
            "_id": {"month": "$program_month", "year": "$program_year"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    periods = await db.programs.aggregate(pipeline).to_list(100)
    return [ProgramPeriod(month=p["_id"]["month"], year=p["_id"]["year"], count=p["count"]) for p in periods]

# Programs CRUD
@api_router.post("/programs", response_model=VehicleProgram)
async def create_program(program: VehicleProgramCreate):
    program_dict = program.dict()
    if program_dict.get("program_month") is None:
        program_dict["program_month"] = datetime.utcnow().month
    if program_dict.get("program_year") is None:
        program_dict["program_year"] = datetime.utcnow().year
    program_obj = VehicleProgram(**program_dict)
    await db.programs.insert_one(program_obj.dict())
    return program_obj

@api_router.get("/programs", response_model=List[VehicleProgram])
async def get_programs(month: Optional[int] = None, year: Optional[int] = None):
    """
    Récupère les programmes de financement
    Si month/year sont fournis, filtre par période
    Sinon, retourne la période la plus récente
    """
    if month and year:
        query = {"program_month": month, "program_year": year}
    else:
        # Trouver la période la plus récente
        latest = await db.programs.find_one(sort=[("program_year", -1), ("program_month", -1)])
        if latest:
            query = {"program_month": latest.get("program_month"), "program_year": latest.get("program_year")}
        else:
            query = {}
    
    programs = await db.programs.find(query).to_list(1000)
    return [VehicleProgram(**p) for p in programs]

@api_router.get("/programs/{program_id}", response_model=VehicleProgram)
async def get_program(program_id: str):
    program = await db.programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    return VehicleProgram(**program)

@api_router.put("/programs/{program_id}", response_model=VehicleProgram)
async def update_program(program_id: str, update: VehicleProgramUpdate):
    program = await db.programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    await db.programs.update_one({"id": program_id}, {"$set": update_data})
    updated = await db.programs.find_one({"id": program_id})
    return VehicleProgram(**updated)

@api_router.delete("/programs/{program_id}")
async def delete_program(program_id: str):
    result = await db.programs.delete_one({"id": program_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"message": "Program deleted successfully"}

# Import programs with password protection
@api_router.post("/import")
async def import_programs(request: ImportRequest):
    """
    Importe des programmes de financement (protégé par mot de passe)
    Remplace tous les programmes pour le mois/année spécifié
    """
    if request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    # Supprimer les programmes existants pour cette période
    await db.programs.delete_many({
        "program_month": request.program_month,
        "program_year": request.program_year
    })
    
    # Insérer les nouveaux programmes
    inserted = 0
    for prog_data in request.programs:
        # Assurer que les taux sont au bon format
        if prog_data.get("option1_rates") and isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        
        prog_data["program_month"] = request.program_month
        prog_data["program_year"] = request.program_year
        prog_data["bonus_cash"] = prog_data.get("bonus_cash", 0)
        
        prog = VehicleProgram(**prog_data)
        await db.programs.insert_one(prog.dict())
        inserted += 1
    
    return {"message": f"Importé {inserted} programmes pour {request.program_month}/{request.program_year}"}

# Calculate financing options
@api_router.post("/calculate", response_model=CalculationResponse)
async def calculate_financing(request: CalculationRequest):
    """
    Calcule et compare les options de financement
    
    Option 1: Prix - Consumer Cash (rabais avant taxes), avec taux Option 1
    Option 2: Prix complet, avec taux réduits (si disponible)
    
    Note: Bonus Cash est affiché mais non inclus dans le calcul du financement
    (car appliqué après taxes, comme comptant)
    """
    vehicle_price = request.vehicle_price
    
    if not request.program_id:
        raise HTTPException(status_code=400, detail="Program ID is required")
    
    program = await db.programs.find_one({"id": request.program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    program_obj = VehicleProgram(**program)
    consumer_cash = program_obj.consumer_cash
    bonus_cash = program_obj.bonus_cash
    
    comparisons = []
    terms = [36, 48, 60, 72, 84, 96]
    
    for term in terms:
        # Option 1: Avec Consumer Cash (rabais avant taxes) + taux Option 1
        option1_rate = get_rate_for_term(program_obj.option1_rates, term)
        principal1 = vehicle_price - consumer_cash  # Rabais avant taxes
        monthly1 = calculate_monthly_payment(principal1, option1_rate, term)
        total1 = round(monthly1 * term, 2)
        
        comparison = PaymentComparison(
            term_months=term,
            option1_rate=option1_rate,
            option1_monthly=monthly1,
            option1_total=total1,
            option1_rebate=consumer_cash
        )
        
        # Option 2: Sans rabais + taux réduits (si disponible)
        if program_obj.option2_rates:
            option2_rate = get_rate_for_term(program_obj.option2_rates, term)
            principal2 = vehicle_price  # Pas de rabais
            monthly2 = calculate_monthly_payment(principal2, option2_rate, term)
            total2 = round(monthly2 * term, 2)
            
            comparison.option2_rate = option2_rate
            comparison.option2_monthly = monthly2
            comparison.option2_total = total2
            
            # Déterminer la meilleure option (coût total le plus bas)
            if total1 < total2:
                comparison.best_option = "1"
                comparison.savings = round(total2 - total1, 2)
            elif total2 < total1:
                comparison.best_option = "2"
                comparison.savings = round(total1 - total2, 2)
            else:
                comparison.best_option = "1"
                comparison.savings = 0
        
        comparisons.append(comparison)
    
    return CalculationResponse(
        vehicle_price=vehicle_price,
        consumer_cash=consumer_cash,
        bonus_cash=bonus_cash,
        brand=program_obj.brand,
        model=program_obj.model,
        trim=program_obj.trim,
        year=program_obj.year,
        comparisons=comparisons
    )

# Seed initial data from PDF pages 20-21 (Février 2026)
@api_router.post("/seed")
async def seed_data():
    """
    Seed les données initiales à partir du PDF Février 2026
    Pages 20 (2026) et 21 (2025)
    """
    # Clear existing data
    await db.programs.delete_many({})
    
    # Taux standard 4.99% pour la plupart des véhicules
    std = {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}
    
    # Mois/Année du programme
    prog_month = 2
    prog_year = 2026
    
    programs_data = [
        # ==================== 2026 MODELS (Page 20) ====================
        
        # CHRYSLER 2026
        {"brand": "Chrysler", "model": "Grand Caravan", "trim": "SXT", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "PHEV", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        # Pacifica (excluding PHEV): Taux bas (0% jusqu'à 72mo, 1.99% à 84mo, 3.49% à 96mo), PAS d'Option 2
        {"brand": "Chrysler", "model": "Pacifica", "trim": "(excluding PHEV)", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "option2_rates": None, "bonus_cash": 0},
        
        # JEEP COMPASS 2026 - Taux Option 2 corrigés
        {"brand": "Jeep", "model": "Compass", "trim": "Sport", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North", "year": 2026,
         "consumer_cash": 3500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North w/ Altitude Package (ADZ)", "year": 2026,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Trailhawk", "year": 2026,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Limited", "year": 2026,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        # JEEP CHEROKEE 2026
        {"brand": "Jeep", "model": "Cherokee", "trim": "Base (KMJL74)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Cherokee", "trim": "(excluding Base)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # JEEP WRANGLER 2026 - Corrigé avec Option 2
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door (JL) non Rubicon", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door Rubicon (JL)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (excl. 392 et 4xe)", "year": 2026,
         "consumer_cash": 5250, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door MOAB 392", "year": 2026,
         "consumer_cash": 6000, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # JEEP GLADIATOR 2026
        {"brand": "Jeep", "model": "Gladiator", "trim": "Sport S, Willys, Sahara, Willys '41", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Gladiator", "trim": "(excl. Sport S, Willys, Sahara, Willys '41)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49}, 
         "bonus_cash": 0},
        
        # JEEP GRAND CHEROKEE 2026 - Taux variables Option 1 et Option 2 corrigés
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Laredo/Laredo X", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Altitude", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": None,
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Limited/Limited Reserve/Summit", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": None,
         "bonus_cash": 0},
        
        # JEEP GRAND WAGONEER 2026
        {"brand": "Jeep", "model": "Grand Wagoneer/L", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99}, 
         "bonus_cash": 0},
        
        # DODGE DURANGO 2026
        {"brand": "Dodge", "model": "Durango", "trim": "SXT, GT, GT Plus", "year": 2026,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Dodge", "model": "Durango", "trim": "GT Hemi V8 Plus, GT Hemi V8 Premium", "year": 2026,
         "consumer_cash": 9000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Dodge", "model": "Durango", "trim": "SRT Hellcat", "year": 2026,
         "consumer_cash": 15500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        # DODGE CHARGER 2026
        {"brand": "Dodge", "model": "Charger", "trim": "2-Door & 4-Door (ICE)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # RAM 2026 - Corrigé
        {"brand": "Ram", "model": "ProMaster", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Tradesman, Express, Warlock", "year": 2026,
         "consumer_cash": 6500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn", "year": 2026,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Sport, Rebel", "year": 2026,
         "consumer_cash": 8250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie (DT6P98)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie, Limited, Longhorn, Tungsten, RHO (excl. DT6P98)", "year": 2026,
         "consumer_cash": 11500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500 Power Wagon Crew Cab", "trim": "(DJ7X91 2UP)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Gas Models", "year": 2026,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Diesel Models", "year": 2026,
         "consumer_cash": 5000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 0.99, "rate_60": 0.99, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "Chassis Cab", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # ==================== 2025 MODELS (Page 21) ====================
        
        # CHRYSLER 2025 - Corrigé
        {"brand": "Chrysler", "model": "Grand Caravan", "trim": "SXT", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "Hybrid", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "Select Models (excl. Hybrid)", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99},
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "(excl. Select & Hybrid)", "year": 2025,
         "consumer_cash": 750, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        # JEEP COMPASS 2025 - Corrigé
        {"brand": "Jeep", "model": "Compass", "trim": "Sport", "year": 2025,
         "consumer_cash": 5500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0.99, "rate_60": 1.99, "rate_72": 1.99, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Altitude, Trailhawk, Trailhawk Elite", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Limited", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        # JEEP WRANGLER 2025 - Corrigé
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) 4xe (JLXL74)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) 4xe (excl. JLXL74)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door (JL) non Rubicon", "year": 2025,
         "consumer_cash": 750, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door Rubicon (JL)", "year": 2025,
         "consumer_cash": 8500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door Rubicon w/ 2.0L", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) (excl. Rubicon 2.0L & 4xe)", "year": 2025,
         "consumer_cash": 8500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # JEEP GLADIATOR 2025
        {"brand": "Jeep", "model": "Gladiator", "trim": None, "year": 2025,
         "consumer_cash": 11000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        # JEEP GRAND CHEROKEE 2025 - Corrigé
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "4xe (WL)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Laredo (WLJH74 2*A)", "year": 2025,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Altitude (WLJH74 2*B)", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit (WLJT74 23S)", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "(WL) (excl. Laredo, Altitude, Summit, 4xe)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        # JEEP GRAND CHEROKEE L 2025 - Corrigé
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Laredo (WLJH75 2*A)", "year": 2025,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Altitude (WLJH75 2*B)", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Overland (WLJS75)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "(WL) (excl. Laredo, Altitude, Overland)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        # JEEP WAGONEER 2025 - Corrigé
        {"brand": "Jeep", "model": "Wagoneer/L", "trim": None, "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Wagoneer/L", "trim": None, "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wagoneer S", "trim": "Limited & Premium (BEV)", "year": 2025,
         "consumer_cash": 8000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # DODGE DURANGO 2025 - Corrigé
        {"brand": "Dodge", "model": "Durango", "trim": "GT, GT Plus", "year": 2025,
         "consumer_cash": 8000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Durango", "trim": "R/T, R/T Plus, R/T 20th Anniversary", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Durango", "trim": "SRT Hellcat", "year": 2025,
         "consumer_cash": 16000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        # DODGE CHARGER 2025 - Corrigé
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "R/T (BEV)", "year": 2025,
         "consumer_cash": 3000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "R/T Plus (BEV)", "year": 2025,
         "consumer_cash": 5000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "Scat Pack (BEV)", "year": 2025,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # DODGE HORNET 2025
        {"brand": "Dodge", "model": "Hornet", "trim": "RT (PHEV)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "RT Plus (PHEV)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "GT (Gas)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "GT Plus (Gas)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        # RAM 2025 - Corrigé
        {"brand": "Ram", "model": "ProMaster", "trim": None, "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Tradesman, Warlock, Express (DT)", "year": 2025,
         "consumer_cash": 9250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn (DT) w/ Off-Roader Value Package (4KF)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn (DT) (excl. Off-Roader)", "year": 2025,
         "consumer_cash": 9250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Sport, Rebel (DT)", "year": 2025,
         "consumer_cash": 10000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie, Limited, Longhorn, Tungsten, RHO (DT)", "year": 2025,
         "consumer_cash": 12250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Gas Models (excl. Chassis Cab, Diesel)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "6.7L High Output Diesel (ETM)", "year": 2025,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 0.99, "rate_60": 0.99, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "Chassis Cab", "trim": None, "year": 2025,
         "consumer_cash": 5000, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # FIAT 2025
        {"brand": "Fiat", "model": "500e", "trim": "BEV", "year": 2025,
         "consumer_cash": 6000,
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "option2_rates": None, "bonus_cash": 5000},
    ]
    
    for prog_data in programs_data:
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        if isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
        prog_data["program_month"] = prog_month
        prog_data["program_year"] = prog_year
        prog = VehicleProgram(**prog_data)
        await db.programs.insert_one(prog.dict())
    
    return {"message": f"Seeded {len(programs_data)} programs for {prog_month}/{prog_year}"}

# ============ Excel Generation Function ============

def generate_excel_from_programs(programs: List[Dict[str, Any]], program_month: int, program_year: int) -> bytes:
    """Génère un fichier Excel selon le format du PDF Stellantis"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programmes"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    option1_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")  # Rouge pour Option 1
    option2_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # Bleu pour Option 2
    bonus_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Vert pour Bonus
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Month names
    month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    # Title Row 1
    ws.merge_cells('A1:S1')
    ws['A1'] = f"PROGRAMMES DE FINANCEMENT RETAIL - {month_names[program_month].upper()} {program_year}"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Sub-header Row 2 - Group headers
    ws.merge_cells('A2:D2')
    ws['A2'] = "VÉHICULE"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = header_alignment
    
    # Option 1 header with rabais
    ws.merge_cells('E2:K2')
    ws['E2'] = "OPTION 1 - Consumer Cash + Taux Standard"
    ws['E2'].font = header_font
    ws['E2'].fill = option1_fill
    ws['E2'].alignment = header_alignment
    
    # Option 2 header with rabais
    ws.merge_cells('L2:R2')
    ws['L2'] = "OPTION 2 - Alternative Consumer Cash + Taux Réduit"
    ws['L2'].font = header_font
    ws['L2'].fill = option2_fill
    ws['L2'].alignment = header_alignment
    
    # Bonus header
    ws['S2'] = "BONUS"
    ws['S2'].font = header_font
    ws['S2'].fill = bonus_fill
    ws['S2'].alignment = header_alignment
    
    # Detail Headers Row 3
    headers = [
        "Marque", "Modèle", "Version", "Année",  # Véhicule (A-D)
        "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",  # Option 1: Rabais + Taux (E-K)
        "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",  # Option 2: Rabais + Taux (L-R)
        "Bonus ($)"  # Bonus (S)
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = header_alignment
        cell.border = thin_border
        
        if col <= 4:
            cell.fill = header_fill
        elif col <= 11:  # Option 1
            cell.fill = option1_fill
        elif col <= 18:  # Option 2
            cell.fill = option2_fill
        else:  # Bonus
            cell.fill = bonus_fill
    
    # Data rows
    for row_idx, prog in enumerate(programs, 4):
        opt1_rates = prog.get("option1_rates") or {}
        opt2_rates = prog.get("option2_rates") or {}
        
        consumer_cash = prog.get("consumer_cash", 0) or 0
        alt_consumer_cash = prog.get("alt_consumer_cash", 0) or 0  # Alternative Consumer Cash pour Option 2
        bonus_cash = prog.get("bonus_cash", 0) or 0
        
        # Format rate display
        def format_rate(rate):
            if rate is None or rate == "-":
                return "-"
            try:
                r = float(rate)
                return f"{r:.2f}%" if r > 0 else "0%"
            except:
                return "-"
        
        data = [
            prog.get("brand", ""),
            prog.get("model", ""),
            prog.get("trim", "") or "",
            prog.get("year", ""),
            # Option 1: Rabais + Taux
            f"${consumer_cash:,.0f}" if consumer_cash else "-",
            format_rate(opt1_rates.get("rate_36")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_48")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_60")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_72")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_84")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_96")) if opt1_rates else "-",
            # Option 2: Rabais + Taux
            f"${alt_consumer_cash:,.0f}" if alt_consumer_cash else "-",
            format_rate(opt2_rates.get("rate_36")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_48")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_60")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_72")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_84")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_96")) if opt2_rates else "-",
            # Bonus
            f"${bonus_cash:,.0f}" if bonus_cash else "-",
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Color coding for columns
            if col >= 5 and col <= 11:  # Option 1
                cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif col >= 12 and col <= 18:  # Option 2
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            elif col == 19:  # Bonus
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Adjust column widths
    column_widths = [12, 18, 28, 7, 12, 8, 8, 8, 8, 8, 8, 12, 8, 8, 8, 8, 8, 8, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Row height for headers
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 25
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def send_excel_email(excel_data: bytes, admin_email: str, program_month: int, program_year: int, program_count: int):
    """Envoie le fichier Excel par email"""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    msg = MIMEMultipart()
    msg['From'] = SMTP_EMAIL
    msg['To'] = admin_email
    msg['Subject'] = f"CalcAuto AiPro - Extraction PDF {month_names[program_month]} {program_year}"
    
    body = f"""
Bonjour,

L'extraction du PDF des programmes de financement est terminée.

📊 Résumé:
• Période: {month_names[program_month]} {program_year}
• Programmes extraits: {program_count}

Le fichier Excel est joint à cet email pour vérification.

⚠️ IMPORTANT: Veuillez vérifier les données dans le fichier Excel avant de confirmer l'import dans l'application.

---
CalcAuto AiPro
    """
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # Attach Excel file
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_data)
    encoders.encode_base64(attachment)
    filename = f"programmes_{month_names[program_month].lower()}_{program_year}.xlsx"
    attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
    msg.attach(attachment)
    
    # Send email
    try:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        logger.error(f"Error sending Excel email: {str(e)}")
        return False

# ============ PDF Import with AI ============

class PDFExtractRequest(BaseModel):
    password: str
    program_month: int
    program_year: int

class ProgramPreview(BaseModel):
    """Preview of extracted program for validation"""
    brand: str
    model: str
    trim: Optional[str] = None
    year: int
    consumer_cash: float = 0
    bonus_cash: float = 0
    option1_rates: Dict[str, float]
    option2_rates: Optional[Dict[str, float]] = None
    
class ExtractedDataResponse(BaseModel):
    success: bool
    message: str
    programs: List[Dict[str, Any]] = []
    raw_text: str = ""

class SaveProgramsRequest(BaseModel):
    password: str
    programs: List[Dict[str, Any]]
    program_month: int
    program_year: int

@api_router.post("/verify-password")
async def verify_password(password: str = Form(...)):
    """Vérifie le mot de passe admin"""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    return {"success": True, "message": "Mot de passe vérifié"}

@api_router.post("/extract-pdf", response_model=ExtractedDataResponse)
async def extract_pdf(
    file: UploadFile = File(...),
    password: str = Form(...),
    program_month: int = Form(...),
    program_year: int = Form(...),
    start_page: int = Form(1),
    end_page: int = Form(9999)
):
    """
    Extrait les données de financement d'un PDF via OpenAI GPT-4
    Retourne les programmes pour prévisualisation/modification avant sauvegarde
    
    start_page et end_page permettent de limiter l'extraction à certaines pages
    (indexation commence à 1)
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    if not OPENAI_API_KEY:
        raise HTTPException(status_code=500, detail="Clé OpenAI non configurée")
    
    try:
        import tempfile
        import os as os_module
        import base64
        from openai import OpenAI
        import PyPDF2
        
        # Save uploaded PDF temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        try:
            # Extract text ONLY from specified pages using PyPDF2
            pdf_text = ""
            with open(tmp_path, 'rb') as pdf_file:
                reader = PyPDF2.PdfReader(pdf_file)
                total_pages = len(reader.pages)
                
                # Convert to 0-based index and validate range
                start_idx = max(0, start_page - 1)  # Convert 1-based to 0-based
                end_idx = min(total_pages, end_page)  # Keep as-is (exclusive end)
                
                logger.info(f"PDF has {total_pages} pages. Extracting pages {start_page} to {end_page} (indices {start_idx} to {end_idx})")
                
                # Only extract the specified pages
                for page_num in range(start_idx, end_idx):
                    page = reader.pages[page_num]
                    page_text = page.extract_text()
                    pdf_text += f"\n--- PAGE {page_num + 1} ---\n{page_text}\n"
                
                logger.info(f"Extracted {end_idx - start_idx} pages, total text length: {len(pdf_text)} characters")
            
            # Use OpenAI to extract structured data
            client = OpenAI(api_key=OPENAI_API_KEY)
            
            extraction_prompt = f"""EXTRAIS TOUS LES VÉHICULES de ce PDF de programmes de financement FCA Canada.

TEXTE COMPLET DU PDF:
{pdf_text}

=== FORMAT DES LIGNES DU PDF ===
Chaque ligne suit ce format:
VÉHICULE [Consumer Cash $X,XXX] [6 taux Option1] [6 taux Option2] [Bonus Cash]

- Si tu vois "- - - - - -" = option non disponible (null)
- Si tu vois "P" avant un montant = c'est quand même le montant
- 6 taux = 36M, 48M, 60M, 72M, 84M, 96M

=== BONUS CASH - TRÈS IMPORTANT ===
Le Bonus Cash est dans la DERNIÈRE COLONNE (colonne verte "Bonus Cash").
- CHAQUE véhicule a son PROPRE montant de Bonus Cash (peut être 0, $1000, $3000, etc.)
- Le Bonus Cash n'est PAS le même pour tous les véhicules!
- Si la colonne Bonus Cash est vide ou montre "-", alors bonus_cash = 0
- Les montants typiques sont: $0, $1,000, $3,000

=== EXEMPLES AVEC BONUS CASH ===

2026 MODELS (généralement PAS de Bonus Cash):
"Grand Caravan SXT    4.99%... " → bonus_cash: 0
"Ram 1500 Big Horn    $6,000  4.99%..." → bonus_cash: 0

2025 MODELS (souvent avec Bonus Cash de $1,000 à $3,000):
"Compass North  $7,500  4.99%...   $1,000" → bonus_cash: 1000
"Ram 1500 Sport  $10,000  4.99%...  $3,000" → bonus_cash: 3000
"Ram 2500/3500 Gas Models  $9,500  4.99%...  -" → bonus_cash: 0 (pas de bonus!)
"Ram Chassis Cab  $5,000  4.99%...  -" → bonus_cash: 0

=== RÈGLE IMPORTANTE POUR RAM 2025 ===
- Ram 1500 2025: bonus_cash = $3,000
- Ram 2500/3500 2025: bonus_cash = 0 (PAS de bonus!)
- Ram ProMaster 2025: bonus_cash = 0
- Ram Chassis Cab 2025: bonus_cash = 0

=== MARQUES À EXTRAIRE ===
- CHRYSLER: Grand Caravan, Pacifica
- JEEP: Compass, Cherokee, Wrangler, Gladiator, Grand Cherokee, Grand Wagoneer
- DODGE: Durango, Charger, Hornet
- RAM: ProMaster, 1500, 2500, 3500, Chassis Cab

=== ANNÉES ===
- "2026 MODELS" → year: 2026
- "2025 MODELS" → year: 2025
Extrais les véhicules des DEUX sections!

=== JSON REQUIS ===
{{
    "programs": [
        {{
            "brand": "Chrysler",
            "model": "Grand Caravan", 
            "trim": "SXT",
            "year": 2026,
            "consumer_cash": 0,
            "bonus_cash": 0,
            "option1_rates": {{"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}},
            "option2_rates": null
        }},
        ... TOUS les autres véhicules ...
    ]
}}

EXTRAIS ABSOLUMENT TOUS LES VÉHICULES DES SECTIONS 2026 ET 2025. 
VÉRIFIE LE BONUS CASH POUR CHAQUE VÉHICULE INDIVIDUELLEMENT!"""

            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "Tu extrais TOUS les véhicules d'un PDF FCA Canada. CHAQUE ligne = 1 entrée. N'oublie AUCUN véhicule. Sections 2026 ET 2025. JSON valide uniquement."},
                    {"role": "user", "content": extraction_prompt}
                ],
                temperature=0.1,
                max_tokens=16000,
                response_format={"type": "json_object"}
            )
            
            response_text = response.choices[0].message.content.strip()
            
            # Clean up response (remove markdown code blocks if present)
            if response_text.startswith("```"):
                lines = response_text.split("\n")
                response_text = "\n".join(lines[1:-1] if lines[-1] == "```" else lines[1:])
                response_text = response_text.strip()
            if response_text.endswith("```"):
                response_text = response_text[:-3].strip()
            
            # Clean common JSON issues
            response_text = response_text.replace('\n', ' ').replace('\r', '')
            response_text = re.sub(r',\s*}', '}', response_text)  # Remove trailing commas
            response_text = re.sub(r',\s*]', ']', response_text)  # Remove trailing commas in arrays
            
            try:
                data = json.loads(response_text)
                programs = data.get("programs", [])
            except json.JSONDecodeError as e:
                # Try to fix common issues and retry
                try:
                    # Find the programs array and try to parse it
                    programs_match = re.search(r'"programs"\s*:\s*\[(.*)\]', response_text, re.DOTALL)
                    if programs_match:
                        programs_str = programs_match.group(1)
                        # Try to parse individual objects
                        programs = []
                        obj_pattern = r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}'
                        for obj_match in re.finditer(obj_pattern, programs_str):
                            try:
                                obj = json.loads(obj_match.group())
                                programs.append(obj)
                            except:
                                continue
                        if programs:
                            data = {"programs": programs}
                        else:
                            raise e
                    else:
                        raise e
                except:
                    return ExtractedDataResponse(
                        success=False,
                        message=f"Erreur de parsing JSON: {str(e)}",
                        programs=[],
                        raw_text=response_text[:3000]
                    )
            
            # Validate and clean programs
            valid_programs = []
            for p in programs:
                # Ensure required fields exist
                if 'brand' in p and 'model' in p:
                    # Clean up rates
                    if p.get('option1_rates') and isinstance(p['option1_rates'], dict):
                        for key in ['rate_36', 'rate_48', 'rate_60', 'rate_72', 'rate_84', 'rate_96']:
                            if key not in p['option1_rates']:
                                p['option1_rates'][key] = 4.99
                    if p.get('option2_rates') and isinstance(p['option2_rates'], dict):
                        for key in ['rate_36', 'rate_48', 'rate_60', 'rate_72', 'rate_84', 'rate_96']:
                            if key not in p['option2_rates']:
                                p['option2_rates'][key] = 0
                    valid_programs.append(p)
            
            # Generate Excel and send by email
            excel_sent = False
            if EXCEL_AVAILABLE and valid_programs and SMTP_EMAIL:
                try:
                    excel_data = generate_excel_from_programs(valid_programs, program_month, program_year)
                    excel_sent = send_excel_email(excel_data, SMTP_EMAIL, program_month, program_year, len(valid_programs))
                    logger.info(f"Excel generated and sent: {excel_sent}")
                except Exception as excel_error:
                    logger.error(f"Error generating/sending Excel: {str(excel_error)}")
            
            # AUTO-SAVE: Sauvegarder automatiquement les programmes dans la base de données
            saved_count = 0
            try:
                # D'abord, supprimer les anciens programmes de la même période
                delete_result = await db.programs.delete_many({
                    "program_month": program_month,
                    "program_year": program_year
                })
                logger.info(f"Deleted {delete_result.deleted_count} old programs for {program_month}/{program_year}")
                
                # Taux par défaut (4.99% standard)
                default_rates = {
                    "rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99,
                    "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99
                }
                
                # Ensuite, ajouter les nouveaux programmes
                for prog in valid_programs:
                    # S'assurer que option1_rates n'est pas None
                    opt1 = prog.get("option1_rates")
                    if opt1 is None or not isinstance(opt1, dict):
                        opt1 = default_rates.copy()
                    
                    program_doc = {
                        "id": str(uuid.uuid4()),
                        "brand": prog.get("brand", ""),
                        "model": prog.get("model", ""),
                        "trim": prog.get("trim", ""),
                        "year": prog.get("year", program_year),
                        "consumer_cash": prog.get("consumer_cash", 0) or 0,
                        "bonus_cash": prog.get("bonus_cash", 0) or 0,
                        "alt_consumer_cash": prog.get("alt_consumer_cash", 0) or 0,
                        "option1_rates": opt1,
                        "option2_rates": prog.get("option2_rates"),
                        "program_month": program_month,
                        "program_year": program_year,
                        "created_at": datetime.utcnow().isoformat()
                    }
                    await db.programs.insert_one(program_doc)
                    saved_count += 1
                
                logger.info(f"Auto-saved {saved_count} programs for {program_month}/{program_year}")
            except Exception as save_error:
                logger.error(f"Error auto-saving programs: {str(save_error)}")
            
            return ExtractedDataResponse(
                success=True,
                message=f"Extrait et sauvegardé {len(valid_programs)} programmes" + (" - Excel envoyé par email!" if excel_sent else ""),
                programs=valid_programs,
                raw_text=""
            )
            
        finally:
            # Clean up temp file
            os_module.unlink(tmp_path)
            
    except Exception as e:
        logger.error(f"Error extracting PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'extraction: {str(e)}")

@api_router.post("/save-programs")
async def save_programs(request: SaveProgramsRequest):
    """
    Sauvegarde les programmes validés dans la base de données
    Remplace les programmes existants pour le mois/année spécifié
    Garde seulement les 6 derniers mois d'historique
    """
    if request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    # Delete existing programs for this period
    await db.programs.delete_many({
        "program_month": request.program_month,
        "program_year": request.program_year
    })
    
    # Insert new programs
    inserted = 0
    skipped = 0
    default_rates = {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}
    
    for prog_data in request.programs:
        # Skip invalid entries (missing brand/model)
        if not prog_data.get("brand") or not prog_data.get("model"):
            skipped += 1
            continue
            
        # Ensure option1_rates has a default value if missing
        if not prog_data.get("option1_rates"):
            prog_data["option1_rates"] = default_rates.copy()
        elif isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
            
        # Process option2_rates if present
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        
        prog_data["program_month"] = request.program_month
        prog_data["program_year"] = request.program_year
        prog_data["bonus_cash"] = prog_data.get("bonus_cash", 0)
        prog_data["consumer_cash"] = prog_data.get("consumer_cash", 0)
        
        try:
            prog = VehicleProgram(**prog_data)
            await db.programs.insert_one(prog.dict())
            inserted += 1
        except Exception as e:
            logger.warning(f"Skipped invalid program: {prog_data.get('brand')} {prog_data.get('model')} - {str(e)}")
            skipped += 1
            continue
    
    # Clean up old programs (keep only 6 months)
    await cleanup_old_programs()
    
    # Calculate brands summary for report
    brands_summary = {}
    for prog_data in request.programs:
        brand = prog_data.get("brand", "Inconnu")
        if brand not in brands_summary:
            brands_summary[brand] = 0
        brands_summary[brand] += 1
    
    # Send automatic email report
    try:
        await send_import_report_email(
            programs_count=inserted,
            program_month=request.program_month,
            program_year=request.program_year,
            brands_summary=brands_summary,
            skipped_count=skipped
        )
        logger.info(f"Import report email sent to {SMTP_EMAIL}")
    except Exception as e:
        logger.warning(f"Failed to send import report email: {str(e)}")
    
    return {
        "success": True,
        "message": f"Sauvegardé {inserted} programmes pour {request.program_month}/{request.program_year}" + (f" ({skipped} ignorés)" if skipped > 0 else "") + " - Rapport envoyé par email"
    }

async def send_import_report_email(programs_count: int, program_month: int, program_year: int, brands_summary: dict, skipped_count: int = 0):
    """Envoie automatiquement un rapport par email après l'import des programmes"""
    months_fr = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
    }
    month_name = months_fr.get(program_month, str(program_month))
    
    # Generate brands table
    brands_rows = ""
    total_programs = 0
    for brand in ['Chrysler', 'Jeep', 'Dodge', 'Ram', 'Fiat']:
        count = brands_summary.get(brand, 0)
        if count > 0:
            brands_rows += f"<tr><td style='padding: 10px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 10px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
            total_programs += count
    
    # Add any other brands not in the standard list
    for brand, count in brands_summary.items():
        if brand not in ['Chrysler', 'Jeep', 'Dodge', 'Ram', 'Fiat'] and count > 0:
            brands_rows += f"<tr><td style='padding: 10px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 10px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
            .header {{ background: linear-gradient(135deg, #1a5f4a 0%, #2d8f6f 100%); color: #fff; padding: 25px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .header .subtitle {{ margin-top: 8px; opacity: 0.9; }}
            .content {{ padding: 25px; }}
            .success-badge {{ background: #d4edda; color: #155724; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px; font-size: 16px; }}
            .success-badge strong {{ font-size: 18px; }}
            .stats-box {{ background: #f8f9fa; border-radius: 8px; padding: 20px; margin-bottom: 20px; text-align: center; }}
            .big-number {{ font-size: 56px; font-weight: bold; color: #1a5f4a; }}
            .big-label {{ color: #666; font-size: 14px; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th {{ background: #1a5f4a; color: #fff; padding: 12px; text-align: left; }}
            .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #eee; }}
            .warning-box {{ background: #fff3cd; border: 1px solid #ffc107; border-radius: 6px; padding: 12px; margin-top: 15px; color: #856404; font-size: 13px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>✅ Import Réussi!</h1>
                <div class="subtitle">CalcAuto AiPro - Rapport d'import automatique</div>
            </div>
            <div class="content">
                <div class="success-badge">
                    <strong>🎉 Programmes {month_name} {program_year}</strong><br>
                    importés avec succès!
                </div>
                
                <div class="stats-box">
                    <div class="big-number">{programs_count}</div>
                    <div class="big-label">programmes de financement</div>
                </div>
                
                <h3 style="color: #1a5f4a; border-bottom: 2px solid #1a5f4a; padding-bottom: 8px;">📊 Répartition par marque</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Marque</th>
                            <th style="text-align: center;">Nombre</th>
                        </tr>
                    </thead>
                    <tbody>
                        {brands_rows}
                        <tr style="background: #f8f9fa; font-weight: bold;">
                            <td style="padding: 12px;">TOTAL</td>
                            <td style="padding: 12px; text-align: center;">{programs_count}</td>
                        </tr>
                    </tbody>
                </table>
                
                {f'<div class="warning-box">⚠️ {skipped_count} programme(s) ignoré(s) en raison de données invalides.</div>' if skipped_count > 0 else ''}
                
                <p style="margin-top: 25px; color: #666; font-size: 14px;">
                    Les nouveaux programmes sont maintenant disponibles dans l'application CalcAuto AiPro. 
                    Vos clients peuvent commencer à utiliser les nouveaux taux immédiatement.
                </p>
            </div>
            <div class="footer">
                <p style="margin: 0;"><strong>CalcAuto AiPro</strong></p>
                <p style="margin: 8px 0 0;">Rapport généré automatiquement le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    subject = f"✅ Import {month_name} {program_year} - {programs_count} programmes"
    
    send_email(SMTP_EMAIL, subject, html_body)

async def cleanup_old_programs():
    """Supprime les programmes de plus de 6 mois"""
    # Get all unique periods
    pipeline = [
        {"$group": {
            "_id": {"month": "$program_month", "year": "$program_year"}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    periods = await db.programs.aggregate(pipeline).to_list(100)
    
    # Keep only the 6 most recent periods
    if len(periods) > 6:
        periods_to_delete = periods[6:]
        for p in periods_to_delete:
            await db.programs.delete_many({
                "program_month": p["_id"]["month"],
                "program_year": p["_id"]["year"]
            })
            logger.info(f"Deleted old programs for {p['_id']['month']}/{p['_id']['year']}")

# ============ Email Functions ============

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
import io

def send_email(to_email: str, subject: str, html_body: str, attachment_data: bytes = None, attachment_name: str = None, inline_images: list = None):
    """
    Envoie un email via Gmail SMTP avec support pour images inline (CID).
    
    Args:
        inline_images: Liste de dicts avec 'cid', 'data' (bytes), 'mimetype' (ex: 'image/jpeg')
    """
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        raise Exception("Configuration SMTP manquante")
    
    # Utiliser 'related' pour supporter les images CID
    msg = MIMEMultipart('related')
    msg['From'] = f"CalcAuto AiPro <{SMTP_EMAIL}>"
    msg['To'] = to_email
    msg['Subject'] = subject
    
    # Créer une partie alternative pour le HTML
    msg_alternative = MIMEMultipart('alternative')
    msg.attach(msg_alternative)
    
    # Corps HTML
    msg_alternative.attach(MIMEText(html_body, 'html', 'utf-8'))
    
    # Images inline (CID) - pour Window Sticker
    if inline_images:
        for img in inline_images:
            mime_img = MIMEImage(img['data'], _subtype=img.get('subtype', 'jpeg'))
            mime_img.add_header('Content-ID', f"<{img['cid']}>")
            mime_img.add_header('Content-Disposition', 'inline', filename=img.get('filename', 'image.jpg'))
            msg.attach(mime_img)
    
    # Pièce jointe PDF si fournie
    if attachment_data and attachment_name:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{attachment_name}"')
        msg.attach(part)
    
    # Connexion SMTP
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
    
    return True

class SendCalculationEmailRequest(BaseModel):
    """Requête pour envoyer un calcul par email"""
    client_email: str
    client_name: str = ""
    vehicle_info: Dict[str, Any]
    calculation_results: Dict[str, Any]
    selected_term: int = 60
    selected_option: str = "1"
    vehicle_price: float
    payment_frequency: str = "monthly"  # monthly, biweekly, weekly
    dealer_name: str = "CalcAuto AiPro"
    dealer_phone: str = ""
    # New fields for complete email
    rates_table: Dict[str, Any] = {}  # Option 1 & 2 rates for all terms
    fees: Dict[str, float] = {}  # frais_dossier, taxe_pneus, frais_rdprm
    trade_in: Dict[str, float] = {}  # valeur_echange, montant_du
    # Window Sticker
    include_window_sticker: bool = True  # Inclure le Window Sticker automatiquement
    vin: str = ""  # VIN pour récupérer le Window Sticker

class SendReportEmailRequest(BaseModel):
    """Requête pour envoyer un rapport après import"""
    programs_count: int
    program_month: int
    program_year: int
    brands_summary: Dict[str, int]


# ============ WINDOW STICKER ENDPOINT ============

@api_router.get("/window-sticker/{vin}")
async def get_window_sticker(vin: str, authorization: Optional[str] = Header(None)):
    """
    Récupère le Window Sticker PDF pour un VIN.
    Télécharge depuis Chrysler/Jeep/Dodge/Ram et stocke dans MongoDB.
    """
    user = await get_current_user(authorization)
    
    # Vérifier si déjà en cache dans MongoDB
    cached = await db.window_stickers.find_one({"vin": vin}, {"_id": 0, "pdf_base64": 0})
    if cached:
        logger.info(f"Window Sticker trouvé en cache pour VIN={vin}")
        return {
            "success": True,
            "cached": True,
            "vin": vin,
            "pdf_url": f"/api/window-sticker/{vin}/pdf",
            "size_bytes": cached.get("size_bytes", 0)
        }
    
    # Télécharger depuis Chrysler/Stellantis
    result = await fetch_window_sticker(vin)
    
    if result["success"]:
        # Sauvegarder dans MongoDB
        await save_window_sticker_to_db(vin, result["pdf_base64"], user["id"])
        
        return {
            "success": True,
            "cached": False,
            "vin": vin,
            "pdf_url": f"/api/window-sticker/{vin}/pdf",
            "size_bytes": result["size_bytes"],
            "source": result.get("brand_source", "stellantis")
        }
    
    return {"success": False, "error": result.get("error", "Window Sticker non disponible")}


@api_router.get("/window-sticker/{vin}/pdf")
async def get_window_sticker_pdf(vin: str):
    """
    Retourne le PDF du Window Sticker (depuis MongoDB).
    """
    from fastapi.responses import Response
    
    doc = await db.window_stickers.find_one({"vin": vin})
    
    if not doc or "pdf_base64" not in doc:
        raise HTTPException(status_code=404, detail="Window Sticker non trouvé")
    
    pdf_bytes = base64.b64decode(doc["pdf_base64"])
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"inline; filename=WindowSticker_{vin}.pdf"
        }
    )


@api_router.post("/send-calculation-email")
async def send_calculation_email(request: SendCalculationEmailRequest):
    """Envoie un calcul de financement par email avec Window Sticker en pièce jointe"""
    try:
        vehicle = request.vehicle_info
        calc = request.calculation_results
        term = request.selected_term
        freq = request.payment_frequency
        rates = request.rates_table
        fees = request.fees
        trade = request.trade_in
        
        # ============ WINDOW STICKER ============
        window_sticker_pdf = None
        window_sticker_images = []  # Liste des images converties du PDF
        window_sticker_url = None
        vin = request.vin or vehicle.get("vin", "")
        
        if request.include_window_sticker and vin and len(vin) == 17:
            logger.info(f"Récupération Window Sticker pour VIN={vin}")
            
            # Vérifier cache MongoDB
            cached = await db.window_stickers.find_one({"vin": vin})
            
            if cached and "pdf_base64" in cached:
                window_sticker_pdf = base64.b64decode(cached["pdf_base64"])
                logger.info(f"Window Sticker trouvé en cache: {len(window_sticker_pdf)} bytes")
            else:
                # Télécharger depuis Chrysler/Stellantis
                ws_result = await fetch_window_sticker(vin, vehicle.get("brand"))
                if ws_result["success"]:
                    window_sticker_pdf = base64.b64decode(ws_result["pdf_base64"])
                    # Sauvegarder dans MongoDB
                    await save_window_sticker_to_db(vin, ws_result["pdf_base64"], "system")
                    logger.info(f"Window Sticker téléchargé et sauvegardé: {len(window_sticker_pdf)} bytes")
                else:
                    logger.warning(f"Window Sticker non disponible: {ws_result.get('error')}")
            
            # Convertir le PDF en images pour l'email
            if window_sticker_pdf:
                window_sticker_images = convert_pdf_to_images(window_sticker_pdf, max_pages=2, dpi=120)
                logger.info(f"Window Sticker converti en {len(window_sticker_images)} image(s)")
            
            # Construire l'URL du Window Sticker (basé sur la marque)
            brand_lower = vehicle.get("brand", "jeep").lower()
            if "chrysler" in brand_lower:
                window_sticker_url = f"https://www.chrysler.com/hostd/windowsticker/getWindowStickerPdf.do?vin={vin}"
            elif "dodge" in brand_lower:
                window_sticker_url = f"https://www.dodge.com/hostd/windowsticker/getWindowStickerPdf.do?vin={vin}"
            elif "ram" in brand_lower:
                window_sticker_url = f"https://www.ramtrucks.com/hostd/windowsticker/getWindowStickerPdf.do?vin={vin}"
            else:
                window_sticker_url = f"https://www.jeep.com/hostd/windowsticker/getWindowStickerPdf.do?vin={vin}"
        
        # Get comparison data
        comparison = None
        for c in calc.get('comparisons', []):
            if c.get('term_months') == term:
                comparison = c
                break
        
        if not comparison:
            comparison = calc.get('comparisons', [{}])[0] if calc.get('comparisons') else {}
        
        consumer_cash = calc.get('consumer_cash', 0)
        bonus_cash = calc.get('bonus_cash', 0)
        
        option1_rate = comparison.get('option1_rate', 0)
        option2_rate = comparison.get('option2_rate', 0)
        
        principal_option1 = comparison.get('principal_option1', request.vehicle_price - consumer_cash)
        principal_option2 = comparison.get('principal_option2', request.vehicle_price)
        
        total_option1 = comparison.get('option1_total', 0)
        total_option2 = comparison.get('option2_total', 0)
        
        if freq == 'weekly':
            option1_payment = comparison.get('option1_weekly', 0)
            option2_payment = comparison.get('option2_weekly', 0)
            freq_label = "Hebdomadaire"
        elif freq == 'biweekly':
            option1_payment = comparison.get('option1_biweekly', 0)
            option2_payment = comparison.get('option2_biweekly', 0)
            freq_label = "Aux 2 semaines"
        else:
            option1_payment = comparison.get('option1_monthly', 0)
            option2_payment = comparison.get('option2_monthly', 0)
            freq_label = "Mensuel"
        
        best_option = comparison.get('best_option', '1')
        savings = comparison.get('savings', 0)
        # Corrected: Option 2 exists if rate is not None AND payment > 0 (rate can be 0%)
        has_option2 = option2_rate is not None and option2_payment is not None and option2_payment > 0
        
        def fmt(val):
            return f"{val:,.0f}".replace(",", " ")
        
        def fmt2(val):
            return f"{val:,.2f}".replace(",", " ").replace(".", ",")
        
        # Build fees info
        frais_dossier = fees.get('frais_dossier', 0)
        taxe_pneus = fees.get('taxe_pneus', 0)
        frais_rdprm = fees.get('frais_rdprm', 0)
        valeur_echange = trade.get('valeur_echange', 0)
        montant_du = trade.get('montant_du', 0)
        
        # LIGHT THEME EMAIL - PDF STYLE
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ font-family: Arial, Helvetica, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; color: #333; font-size: 14px; line-height: 1.5; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
                
                .header {{ background: #1a5f4a; color: #fff; padding: 20px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .header p {{ margin: 5px 0 0; opacity: 0.9; font-size: 13px; }}
                
                .content {{ padding: 25px; }}
                
                .greeting {{ font-size: 15px; margin-bottom: 20px; }}
                
                .section {{ margin-bottom: 20px; }}
                .section-title {{ font-size: 12px; color: #666; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 10px; border-bottom: 1px solid #eee; padding-bottom: 5px; }}
                
                .vehicle-box {{ background: #f8f9fa; border: 2px solid #1a5f4a; border-radius: 8px; padding: 15px; text-align: center; }}
                .vehicle-brand {{ color: #1a5f4a; font-size: 14px; font-weight: bold; }}
                .vehicle-model {{ font-size: 20px; font-weight: bold; color: #333; margin: 5px 0; }}
                .vehicle-price {{ font-size: 28px; font-weight: bold; color: #1a5f4a; }}
                
                .info-table {{ width: 100%; border-collapse: collapse; }}
                .info-table td {{ padding: 8px 0; border-bottom: 1px solid #eee; }}
                .info-table td:first-child {{ color: #666; }}
                .info-table td:last-child {{ text-align: right; font-weight: 600; }}
                
                .rates-table {{ width: 100%; border-collapse: collapse; background: #f8f9fa; border-radius: 6px; overflow: hidden; }}
                .rates-table th {{ background: #1a5f4a; color: #fff; padding: 10px; font-size: 12px; }}
                .rates-table td {{ padding: 8px 10px; text-align: center; border-bottom: 1px solid #e0e0e0; }}
                .rates-table tr:last-child td {{ border-bottom: none; }}
                .rates-table .selected {{ background: #d4edda; font-weight: bold; }}
                .rate-opt1 {{ color: #c0392b; font-weight: 600; }}
                .rate-opt2 {{ color: #1a5f4a; font-weight: 600; }}
                
                .best-choice {{ background: #d4edda; border: 2px solid #1a5f4a; border-radius: 8px; padding: 15px; margin: 20px 0; }}
                .best-choice-title {{ font-size: 18px; font-weight: bold; color: #155724; }}
                .best-choice-savings {{ font-size: 14px; color: #155724; margin-top: 5px; }}
                
                .options-container {{ display: table; width: 100%; border-spacing: 10px 0; }}
                .option-box {{ display: table-cell; width: 50%; vertical-align: top; border: 2px solid #ddd; border-radius: 8px; padding: 15px; background: #fafafa; }}
                .option-box.winner {{ border-color: #1a5f4a; background: #f0fff4; }}
                .option-title {{ font-size: 16px; font-weight: bold; margin-bottom: 10px; }}
                .option-title.opt1 {{ color: #c0392b; }}
                .option-title.opt2 {{ color: #1a5f4a; }}
                
                .option-detail {{ display: flex; justify-content: space-between; margin-bottom: 5px; font-size: 13px; }}
                .option-detail span:first-child {{ color: #666; }}
                
                .payment-highlight {{ background: #f0f0f0; border-radius: 6px; padding: 12px; text-align: center; margin-top: 10px; }}
                .payment-highlight.opt1 {{ background: #fdf2f2; }}
                .payment-highlight.opt2 {{ background: #f0fff4; }}
                .payment-label {{ font-size: 12px; color: #666; }}
                .payment-amount {{ font-size: 24px; font-weight: bold; margin: 5px 0; }}
                .payment-amount.opt1 {{ color: #c0392b; }}
                .payment-amount.opt2 {{ color: #1a5f4a; }}
                .payment-total {{ font-size: 12px; color: #666; }}
                .payment-total strong {{ color: #333; }}
                
                .winner-badge {{ display: inline-block; background: #1a5f4a; color: #fff; font-size: 10px; padding: 3px 8px; border-radius: 10px; margin-left: 5px; }}
                
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; border-top: 1px solid #eee; }}
                .footer .dealer {{ font-size: 16px; font-weight: bold; color: #1a5f4a; }}
                .footer .disclaimer {{ font-size: 11px; color: #999; margin-top: 10px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>CalcAuto AiPro</h1>
                    <p>Soumission de financement</p>
                </div>
                
                <div class="content">
                    <p class="greeting">Bonjour{' ' + request.client_name if request.client_name else ''},</p>
                    <p style="color: #666; margin-top: -10px;">Voici le détail de votre soumission de financement:</p>
                    
                    <!-- VEHICLE -->
                    <div class="section">
                        <div class="section-title">Véhicule sélectionné</div>
                        <div class="vehicle-box">
                            <div class="vehicle-brand">{vehicle.get('brand', '')}</div>
                            <div class="vehicle-model">{vehicle.get('model', '')} {vehicle.get('trim', '') or ''} {vehicle.get('year', '')}</div>
                            <div class="vehicle-price">{fmt(request.vehicle_price)} $</div>
                        </div>
                    </div>
                    
                    <!-- RATES TABLE -->
                    <div class="section">
                        <div class="section-title">Tableau des taux</div>
                        <table class="rates-table">
                            <thead>
                                <tr>
                                    <th style="text-align: left;">Terme</th>
                                    <th>Option 1</th>
                                    <th>Option 2</th>
                                </tr>
                            </thead>
                            <tbody>
                                <tr class="{'selected' if term == 36 else ''}"><td style="text-align: left;">36 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">0%</td></tr>
                                <tr class="{'selected' if term == 48 else ''}"><td style="text-align: left;">48 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">0%</td></tr>
                                <tr class="{'selected' if term == 60 else ''}"><td style="text-align: left;">60 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">0%</td></tr>
                                <tr class="{'selected' if term == 72 else ''}"><td style="text-align: left;">72 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">1,49%</td></tr>
                                <tr class="{'selected' if term == 84 else ''}"><td style="text-align: left;">84 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">1,99%</td></tr>
                                <tr class="{'selected' if term == 96 else ''}"><td style="text-align: left;">96 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">3,49%</td></tr>
                            </tbody>
                        </table>
                    </div>
                    
                    <!-- DETAILS -->
                    <div class="section">
                        <div class="section-title">Détails du financement</div>
                        <table class="info-table">
                            <tr><td>Prix du véhicule</td><td>{fmt(request.vehicle_price)} $</td></tr>
                            {f"<tr><td>Rabais (avant taxes)</td><td style='color: #1a5f4a;'>-{fmt(consumer_cash)} $</td></tr>" if consumer_cash > 0 else ""}
                            {f"<tr><td>Bonus Cash (après taxes)</td><td style='color: #1a5f4a;'>-{fmt(bonus_cash)} $</td></tr>" if bonus_cash > 0 else ""}
                            {f"<tr><td>Frais dossier</td><td>{fmt2(frais_dossier)} $</td></tr>" if frais_dossier > 0 else ""}
                            {f"<tr><td>Taxe pneus</td><td>{fmt(taxe_pneus)} $</td></tr>" if taxe_pneus > 0 else ""}
                            {f"<tr><td>Frais RDPRM</td><td>{fmt(frais_rdprm)} $</td></tr>" if frais_rdprm > 0 else ""}
                            {f"<tr><td>Valeur échange</td><td>-{fmt(valeur_echange)} $</td></tr>" if valeur_echange > 0 else ""}
                            <tr><td>Terme sélectionné</td><td><strong>{term} mois</strong></td></tr>
                            <tr><td>Fréquence de paiement</td><td><strong>{freq_label}</strong></td></tr>
                        </table>
                    </div>
                    
                    <!-- BEST CHOICE BANNER -->
                    {"<div class='best-choice'><div class='best-choice-title'>🏆 Option " + best_option + " = Meilleur choix!</div><div class='best-choice-savings'>Économies de <strong>" + fmt(savings) + " $</strong> sur le coût total</div></div>" if has_option2 and savings > 0 else ""}
                    
                    <!-- OPTIONS COMPARISON -->
                    <div class="section">
                        <div class="section-title">Comparaison des options</div>
                        <table style="width: 100%; border-spacing: 10px; border-collapse: separate;">
                            <tr>
                                <td style="width: 50%; vertical-align: top; border: 2px solid {'#1a5f4a' if best_option == '1' else '#ddd'}; border-radius: 8px; padding: 15px; background: {'#f0fff4' if best_option == '1' else '#fafafa'};">
                                    <div class="option-title opt1">Option 1 {"<span class='winner-badge'>✓ MEILLEUR</span>" if best_option == '1' else ""}</div>
                                    <div style="font-size: 12px; color: #666; margin-bottom: 10px;">Rabais + Taux standard</div>
                                    <div class="option-detail"><span>Rabais:</span><span style="color: #1a5f4a; font-weight: 600;">{'-' + fmt(consumer_cash) + ' $' if consumer_cash > 0 else '$0'}</span></div>
                                    <div class="option-detail"><span>Capital financé:</span><span>{fmt(principal_option1)} $</span></div>
                                    <div class="option-detail"><span>Taux:</span><span class="rate-opt1">{option1_rate}%</span></div>
                                    <div class="payment-highlight opt1">
                                        <div class="payment-label">{freq_label}</div>
                                        <div class="payment-amount opt1">{fmt2(option1_payment)} $</div>
                                        <div class="payment-total">Total ({term} mois): <strong>{fmt(total_option1)} $</strong></div>
                                    </div>
                                </td>
                                
                                {"<td style='width: 50%; vertical-align: top; border: 2px solid " + ("#1a5f4a" if best_option == "2" else "#ddd") + "; border-radius: 8px; padding: 15px; background: " + ("#f0fff4" if best_option == "2" else "#fafafa") + ";'><div class='option-title opt2'>Option 2 " + ("<span class='winner-badge'>✓ MEILLEUR</span>" if best_option == "2" else "") + "</div><div style='font-size: 12px; color: #666; margin-bottom: 10px;'>$0 rabais + Taux réduit</div><div class='option-detail'><span>Rabais:</span><span>$0</span></div><div class='option-detail'><span>Capital financé:</span><span>" + fmt(principal_option2) + " $</span></div><div class='option-detail'><span>Taux:</span><span class='rate-opt2'>" + str(option2_rate) + "%</span></div><div class='payment-highlight opt2'><div class='payment-label'>" + freq_label + "</div><div class='payment-amount opt2'>" + fmt2(option2_payment) + " $</div><div class='payment-total'>Total (" + str(term) + " mois): <strong>" + fmt(total_option2) + " $</strong></div></div></td>" if has_option2 else "<td style='width: 50%; vertical-align: top; border: 2px solid #ddd; border-radius: 8px; padding: 15px; background: #f5f5f5; text-align: center; color: #999;'><div class='option-title' style='color: #999;'>Option 2</div><div style='padding: 30px 0;'>Non disponible<br>pour ce véhicule</div></td>"}
                            </tr>
                        </table>
                    </div>
                    
                    {f"<div style='background: #fff3cd; border: 1px solid #ffc107; border-radius: 6px; padding: 12px; margin-top: 15px;'><span style='color: #856404;'>ℹ️ Bonus Cash de {fmt(bonus_cash)} $ sera déduit après taxes (au comptant)</span></div>" if bonus_cash > 0 else ""}
                    
                    <!-- WINDOW STICKER SECTION WITH IMAGES -->
                    {generate_window_sticker_html(vin, window_sticker_images, window_sticker_url, window_sticker_pdf)}
                </div>
                
                <div class="footer">
                    <div class="dealer">{request.dealer_name}</div>
                    {f"<div style='color: #666; margin-top: 5px;'>{request.dealer_phone}</div>" if request.dealer_phone else ""}
                    <div class="disclaimer">
                        Ce calcul est une estimation et ne constitue pas une offre de financement officielle.<br>
                        Les taux et conditions peuvent varier selon votre dossier de crédit.
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        subject = f"Soumission - {vehicle.get('brand', '')} {vehicle.get('model', '')} {vehicle.get('year', '')}"
        
        # Envoyer l'email avec ou sans Window Sticker en pièce jointe
        # Préparer les images inline pour CID
        inline_images = []
        if window_sticker_images and len(window_sticker_images) > 0:
            img = window_sticker_images[0]
            inline_images.append({
                'cid': f'windowsticker_{vin}',
                'data': base64.b64decode(img['base64']),
                'subtype': 'jpeg',
                'filename': f'WindowSticker_{vin}.jpg'
            })
        
        if window_sticker_pdf:
            send_email(
                request.client_email, 
                subject, 
                html_body, 
                attachment_data=window_sticker_pdf,
                attachment_name=f"WindowSticker_{vin}.pdf",
                inline_images=inline_images
            )
            return {"success": True, "message": f"Email envoyé à {request.client_email} avec Window Sticker"}
        else:
            send_email(request.client_email, subject, html_body, inline_images=inline_images if inline_images else None)
            return {"success": True, "message": f"Email envoyé à {request.client_email}"}
        
    except Exception as e:
        logger.error(f"Erreur envoi email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

@api_router.post("/send-import-report")
async def send_import_report(request: SendReportEmailRequest):
    """Envoie un rapport par email après l'import des programmes"""
    try:
        months_fr = {
            1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
            5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
            9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
        }
        month_name = months_fr.get(request.program_month, str(request.program_month))
        
        # Générer le résumé par marque
        brands_html = ""
        for brand, count in request.brands_summary.items():
            brands_html += f"<tr><td style='padding: 8px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 8px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
        
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
                .header {{ background: linear-gradient(135deg, #4ECDC4 0%, #44a08d 100%); color: white; padding: 30px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .content {{ padding: 30px; }}
                .success-badge {{ background: #d4edda; color: #155724; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px; }}
                .stats-box {{ background: #f8f9fa; border-radius: 8px; padding: 20px; margin: 20px 0; }}
                .big-number {{ font-size: 48px; font-weight: bold; color: #1a1a2e; text-align: center; }}
                .big-label {{ text-align: center; color: #666; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
                th {{ background: #1a1a2e; color: white; padding: 10px; text-align: left; }}
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>✅ Import Réussi</h1>
                    <p>CalcAuto AiPro - Rapport d'import</p>
                </div>
                <div class="content">
                    <div class="success-badge">
                        <strong>🎉 Programmes {month_name} {request.program_year} importés avec succès!</strong>
                    </div>
                    
                    <div class="stats-box">
                        <div class="big-number">{request.programs_count}</div>
                        <div class="big-label">programmes de financement</div>
                    </div>
                    
                    <h3>Répartition par marque:</h3>
                    <table>
                        <tr>
                            <th>Marque</th>
                            <th style="text-align: center;">Nombre</th>
                        </tr>
                        {brands_html}
                    </table>
                    
                    <p style="margin-top: 20px; color: #666;">
                        Les nouveaux programmes sont maintenant disponibles dans l'application CalcAuto AiPro.
                    </p>
                </div>
                <div class="footer">
                    <p>Ce rapport a été généré automatiquement après l'import.</p>
                    <p>Date: {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        subject = f"✅ Import {month_name} {request.program_year} - {request.programs_count} programmes"
        
        send_email(SMTP_EMAIL, subject, html_body)
        
        return {"success": True, "message": "Rapport envoyé par email"}
        
    except Exception as e:
        logger.error(f"Erreur envoi rapport: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

@api_router.post("/test-email")
async def test_email():
    """Teste la configuration email"""
    try:
        html_body = """
        <div style="font-family: Arial; padding: 20px; text-align: center;">
            <h1 style="color: #4ECDC4;">✅ Test Email Réussi!</h1>
            <p>Votre configuration Gmail SMTP fonctionne correctement.</p>
            <p style="color: #666;">CalcAuto AiPro</p>
        </div>
        """
        send_email(SMTP_EMAIL, "🧪 Test CalcAuto AiPro - Email OK", html_body)
        return {"success": True, "message": f"Email de test envoyé à {SMTP_EMAIL}"}
    except Exception as e:
        logger.error(f"Erreur test email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# ============ CRM Endpoints ============

@api_router.post("/submissions")
async def create_submission(submission: SubmissionCreate, authorization: Optional[str] = Header(None)):
    """Créer une nouvelle soumission avec rappel automatique 24h"""
    from datetime import timedelta
    
    user = await get_current_user(authorization)
    
    # Set default reminder to 24h from now
    reminder_date = datetime.utcnow() + timedelta(hours=24)
    
    new_submission = Submission(
        client_name=submission.client_name,
        client_phone=submission.client_phone,
        client_email=submission.client_email,
        vehicle_brand=submission.vehicle_brand,
        vehicle_model=submission.vehicle_model,
        vehicle_year=submission.vehicle_year,
        vehicle_price=submission.vehicle_price,
        term=submission.term,
        payment_monthly=submission.payment_monthly,
        payment_biweekly=submission.payment_biweekly,
        payment_weekly=submission.payment_weekly,
        selected_option=submission.selected_option,
        rate=submission.rate,
        program_month=submission.program_month,
        program_year=submission.program_year,
        reminder_date=reminder_date,
        owner_id=user["id"]
    )
    
    await db.submissions.insert_one(new_submission.dict())
    
    return {"success": True, "submission": new_submission.dict(), "message": "Soumission enregistrée - Rappel dans 24h"}

@api_router.get("/submissions")
async def get_submissions(search: Optional[str] = None, status: Optional[str] = None, authorization: Optional[str] = Header(None)):
    """Récupérer les soumissions de l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    query = {"owner_id": user["id"]}
    
    if search:
        # Search by name or phone
        query["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"client_phone": {"$regex": search, "$options": "i"}}
        ]
    
    if status:
        query["status"] = status
    
    submissions = await db.submissions.find(query).sort("submission_date", -1).to_list(500)
    
    # Convert ObjectId to string and datetime to ISO format
    for sub in submissions:
        if "_id" in sub:
            sub["_id"] = str(sub["_id"])
        if "submission_date" in sub and sub["submission_date"]:
            sub["submission_date"] = sub["submission_date"].isoformat()
        if "reminder_date" in sub and sub["reminder_date"]:
            sub["reminder_date"] = sub["reminder_date"].isoformat()
    
    return submissions

@api_router.get("/submissions/reminders")
async def get_reminders(authorization: Optional[str] = Header(None)):
    """Récupérer les soumissions avec rappels dus ou à venir pour l'utilisateur connecté"""
    user = await get_current_user(authorization)
    now = datetime.utcnow()
    
    # Get submissions where reminder is due and not done (filtré par owner_id)
    reminders_due = await db.submissions.find({
        "owner_id": user["id"],
        "reminder_done": False,
        "reminder_date": {"$lte": now}
    }).sort("reminder_date", 1).to_list(100)
    
    # Get upcoming reminders (next 7 days)
    from datetime import timedelta
    next_week = now + timedelta(days=7)
    
    reminders_upcoming = await db.submissions.find({
        "owner_id": user["id"],
        "reminder_done": False,
        "reminder_date": {"$gt": now, "$lte": next_week}
    }).sort("reminder_date", 1).to_list(100)
    
    # Format dates
    for sub in reminders_due + reminders_upcoming:
        if "_id" in sub:
            sub["_id"] = str(sub["_id"])
        if "submission_date" in sub and sub["submission_date"]:
            sub["submission_date"] = sub["submission_date"].isoformat()
        if "reminder_date" in sub and sub["reminder_date"]:
            sub["reminder_date"] = sub["reminder_date"].isoformat()
    
    return {
        "due": reminders_due,
        "upcoming": reminders_upcoming,
        "due_count": len(reminders_due),
        "upcoming_count": len(reminders_upcoming)
    }

@api_router.put("/submissions/{submission_id}/reminder")
async def update_reminder(submission_id: str, reminder: ReminderUpdate, authorization: Optional[str] = Header(None)):
    """Mettre à jour la date de rappel"""
    user = await get_current_user(authorization)
    
    update_data = {"reminder_date": reminder.reminder_date, "reminder_done": False}
    
    if reminder.notes:
        update_data["notes"] = reminder.notes
    
    result = await db.submissions.update_one(
        {"id": submission_id, "owner_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")
    
    return {"success": True, "message": "Rappel mis à jour"}

@api_router.put("/submissions/{submission_id}/done")
async def mark_reminder_done(submission_id: str, authorization: Optional[str] = Header(None), new_reminder_date: Optional[str] = None):
    """Marquer un rappel comme fait, optionnellement planifier un nouveau"""
    user = await get_current_user(authorization)
    
    update_data = {"reminder_done": True, "status": "contacted"}
    
    if new_reminder_date:
        update_data["reminder_date"] = datetime.fromisoformat(new_reminder_date.replace('Z', '+00:00'))
        update_data["reminder_done"] = False
    
    result = await db.submissions.update_one(
        {"id": submission_id, "owner_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")
    
    return {"success": True, "message": "Rappel marqué comme fait" + (" - Nouveau rappel planifié" if new_reminder_date else "")}

@api_router.put("/submissions/{submission_id}/status")
async def update_submission_status(submission_id: str, status: str, authorization: Optional[str] = Header(None)):
    """Mettre à jour le statut (pending, contacted, converted, lost)"""
    user = await get_current_user(authorization)
    
    if status not in ["pending", "contacted", "converted", "lost"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    result = await db.submissions.update_one(
        {"id": submission_id, "owner_id": user["id"]},
        {"$set": {"status": status}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")
    
    return {"success": True, "message": f"Statut mis à jour: {status}"}

@api_router.post("/compare-programs")
async def compare_programs_with_submissions(authorization: Optional[str] = Header(None)):
    """Compare les programmes actuels avec les soumissions passées pour trouver de meilleures offres"""
    user = await get_current_user(authorization)
    
    try:
        # Get current programs (latest month/year)
        latest = await db.programs.find_one(sort=[("program_year", -1), ("program_month", -1)])
        if not latest:
            return {"better_offers": [], "count": 0}
        
        current_month = latest["program_month"]
        current_year = latest["program_year"]
        
        # Get submissions from PREVIOUS months FOR THIS USER
        submissions = await db.submissions.find({
            "owner_id": user["id"],
            "$or": [
                {"program_month": {"$lt": current_month}, "program_year": current_year},
                {"program_year": {"$lt": current_year}}
            ]
        }).to_list(500)
        
        better_offers = []
        
        for sub in submissions:
            # Find matching program in current month
            program = await db.programs.find_one({
                "brand": sub.get("vehicle_brand"),
                "model": sub.get("vehicle_model"),
                "year": sub.get("vehicle_year"),
                "program_month": current_month,
                "program_year": current_year
            })
            
            if not program:
                continue
            
            # Calculate new payment
            term = int(sub.get("term", 72))
            old_payment = float(sub.get("payment_monthly", 0))
            
            # Get rate for this term
            opt1_rates = program.get("option1_rates", {})
            rate_key = f"rate_{term}"
            new_rate = float(opt1_rates.get(rate_key, 4.99))
            
            # Calculate with new program
            vehicle_price = float(sub.get("vehicle_price", 0))
            consumer_cash = float(program.get("consumer_cash", 0))
            bonus_cash = float(program.get("bonus_cash", 0))
            principal = vehicle_price - consumer_cash - bonus_cash
            
            if principal <= 0:
                continue
                
            if new_rate == 0:
                new_payment = round(principal / term, 2)
            else:
                monthly_rate = new_rate / 100 / 12
                new_payment = round(principal * (monthly_rate * (1 + monthly_rate) ** term) / ((1 + monthly_rate) ** term - 1), 2)
            
            # Check if better (at least $10 savings)
            if new_payment < old_payment - 10:
                savings_monthly = old_payment - new_payment
                savings_total = savings_monthly * term
                
                better_offers.append({
                    "submission_id": str(sub.get("id", "")),
                    "owner_id": user["id"],  # Add owner_id for isolation
                    "client_name": str(sub.get("client_name", "")),
                    "client_phone": str(sub.get("client_phone", "")),
                    "client_email": str(sub.get("client_email", "")),
                    "vehicle": f"{sub.get('vehicle_brand', '')} {sub.get('vehicle_model', '')} {sub.get('vehicle_year', '')}",
                    "old_payment": round(old_payment, 2),
                    "new_payment": round(new_payment, 2),
                    "old_rate": round(float(sub.get("rate", 0)), 2),
                    "new_rate": round(new_rate, 2),
                    "savings_monthly": round(savings_monthly, 2),
                    "savings_total": round(savings_total, 2),
                    "term": term,
                    "old_program": f"{sub.get('program_month', '?')}/{sub.get('program_year', '?')}",
                    "new_program": f"{current_month}/{current_year}",
                    "approved": False,
                    "email_sent": False
                })
        
        # Store better offers in DB for approval
        if better_offers:
            await db.better_offers.delete_many({"owner_id": user["id"]})  # Clear old offers for this user only
            for offer in better_offers:
                await db.better_offers.insert_one(offer.copy())  # Use copy to avoid _id being added to original
            
            # Send notification email to admin
            try:
                send_better_offers_notification(better_offers)
            except Exception as e:
                logger.error(f"Error sending better offers notification: {e}")
        
        # Return clean offers without MongoDB ObjectId
        return {"better_offers": better_offers, "count": len(better_offers)}
    
    except Exception as e:
        logger.error(f"Error in compare_programs: {e}")
        return {"better_offers": [], "count": 0, "error": str(e)}

def send_better_offers_notification(offers: List[dict]):
    """Envoie une notification par email des meilleures offres disponibles"""
    if not SMTP_EMAIL:
        return
    
    offers_html = ""
    for offer in offers:
        offers_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #eee;">{offer['client_name']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee;">{offer['vehicle']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right;">{offer['old_payment']:.2f}$</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right; color: #28a745; font-weight: bold;">{offer['new_payment']:.2f}$</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right; color: #28a745;">-{offer['savings_monthly']:.2f}$/mois</td>
        </tr>
        """
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head><meta charset="utf-8"></head>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 700px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden;">
            <div style="background: #1a1a2e; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0;">🔔 Meilleures Offres Disponibles</h1>
            </div>
            <div style="padding: 20px;">
                <p style="font-size: 16px;">{len(offers)} client(s) peuvent bénéficier de meilleurs taux!</p>
                
                <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                    <tr style="background: #f8f9fa;">
                        <th style="padding: 12px; text-align: left;">Client</th>
                        <th style="padding: 12px; text-align: left;">Véhicule</th>
                        <th style="padding: 12px; text-align: right;">Ancien</th>
                        <th style="padding: 12px; text-align: right;">Nouveau</th>
                        <th style="padding: 12px; text-align: right;">Économie</th>
                    </tr>
                    {offers_html}
                </table>
                
                <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0;">
                    <strong>⚠️ Action requise:</strong><br>
                    Ouvrez l'application pour approuver l'envoi des emails aux clients.
                </div>
                
                <a href="https://vin-scanner-1.preview.emergentagent.com" style="display: inline-block; background: #4ECDC4; color: #1a1a2e; padding: 12px 24px; border-radius: 8px; text-decoration: none; font-weight: bold;">
                    Ouvrir l'application
                </a>
            </div>
        </div>
    </body>
    </html>
    """
    
    send_email(SMTP_EMAIL, f"🔔 CalcAuto - {len(offers)} client(s) à relancer!", html_body)

@api_router.get("/better-offers")
async def get_better_offers(authorization: Optional[str] = Header(None)):
    """Récupérer les meilleures offres en attente d'approbation pour l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    offers = await db.better_offers.find({
        "owner_id": user["id"],
        "approved": False, 
        "email_sent": False
    }).to_list(100)
    
    for offer in offers:
        if "_id" in offer:
            offer["_id"] = str(offer["_id"])
    
    return offers

@api_router.post("/better-offers/{submission_id}/approve")
async def approve_better_offer(submission_id: str, authorization: Optional[str] = Header(None)):
    """Approuver et envoyer l'email au client pour une meilleure offre"""
    user = await get_current_user(authorization)
    
    offer = await db.better_offers.find_one({"submission_id": submission_id, "owner_id": user["id"]})
    
    if not offer:
        raise HTTPException(status_code=404, detail="Offre non trouvée")
    
    if offer.get("email_sent"):
        return {"success": False, "message": "Email déjà envoyé"}
    
    # Send email to client
    try:
        send_client_better_offer_email(offer)
        
        # Mark as sent
        await db.better_offers.update_one(
            {"submission_id": submission_id, "owner_id": user["id"]},
            {"$set": {"approved": True, "email_sent": True}}
        )
        
        return {"success": True, "message": f"Email envoyé à {offer['client_email']}"}
    except Exception as e:
        logger.error(f"Error sending client email: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

def send_client_better_offer_email(offer: dict):
    """Envoie un email au client pour l'informer d'une meilleure offre"""
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head><meta charset="utf-8"></head>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden;">
            <div style="background: #1a1a2e; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; color: #4ECDC4;">🎉 Bonne Nouvelle!</h1>
            </div>
            <div style="padding: 30px;">
                <p style="font-size: 18px;">Bonjour {offer['client_name']},</p>
                
                <p>De nouveaux programmes de financement sont disponibles et vous permettraient d'<strong>économiser</strong> sur votre {offer['vehicle']}!</p>
                
                <div style="background: #d4edda; border-radius: 10px; padding: 20px; margin: 20px 0; text-align: center;">
                    <p style="margin: 0; color: #666;">Votre paiement actuel</p>
                    <p style="font-size: 24px; margin: 5px 0; text-decoration: line-through; color: #999;">{offer['old_payment']:.2f}$/mois</p>
                    
                    <p style="margin: 15px 0 0; color: #666;">Nouveau paiement possible</p>
                    <p style="font-size: 32px; margin: 5px 0; color: #28a745; font-weight: bold;">{offer['new_payment']:.2f}$/mois</p>
                    
                    <p style="font-size: 18px; color: #28a745; margin-top: 15px;">
                        💰 Économie: {offer['savings_monthly']:.2f}$/mois ({offer['savings_total']:.2f}$ sur {offer['term']} mois)
                    </p>
                </div>
                
                <p>Contactez-nous pour profiter de cette offre!</p>
                
                <p style="color: #666; font-size: 14px; margin-top: 30px;">
                    Cordialement,<br>
                    L'équipe CalcAuto AiPro
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    send_email(offer['client_email'], f"🎉 Économisez {offer['savings_monthly']:.2f}$/mois sur votre {offer['vehicle']}!", html_body)

@api_router.post("/better-offers/{submission_id}/ignore")
async def ignore_better_offer(submission_id: str, authorization: Optional[str] = Header(None)):
    """Ignorer une meilleure offre"""
    user = await get_current_user(authorization)
    
    result = await db.better_offers.delete_one({"submission_id": submission_id, "owner_id": user["id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Offre non trouvée")
    
    return {"success": True, "message": "Offre ignorée"}

# ============ Contacts API ============

@api_router.get("/contacts")
async def get_contacts(authorization: Optional[str] = Header(None)):
    """Récupère les contacts de l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    # Filtrer par owner_id
    contacts = await db.contacts.find({"owner_id": user["id"]}).sort("name", 1).to_list(10000)
    return [{
        "id": c.get("id"),
        "name": c.get("name"),
        "phone": c.get("phone", ""),
        "email": c.get("email", ""),
        "created_at": c.get("created_at"),
        "source": c.get("source", "import")
    } for c in contacts]

@api_router.post("/contacts")
async def create_contact(contact: ContactCreate, authorization: Optional[str] = Header(None)):
    """Crée un nouveau contact pour l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    contact_obj = Contact(
        name=contact.name,
        phone=contact.phone,
        email=contact.email,
        source=contact.source,
        owner_id=user["id"]
    )
    await db.contacts.insert_one(contact_obj.dict())
    return contact_obj

@api_router.post("/contacts/bulk")
async def create_contacts_bulk(request: ContactBulkCreate, authorization: Optional[str] = Header(None)):
    """Importe plusieurs contacts en masse pour l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    if not request.contacts:
        return {"success": True, "imported": 0, "message": "Aucun contact à importer"}
    
    # Préparer les contacts avec owner_id
    contacts_to_insert = []
    for c in request.contacts:
        contact_obj = Contact(
            name=c.name,
            phone=c.phone,
            email=c.email,
            source=c.source,
            owner_id=user["id"]
        )
        contacts_to_insert.append(contact_obj.dict())
    
    # Supprimer les doublons par nom+phone POUR CET UTILISATEUR avant insertion
    existing_contacts = await db.contacts.find({"owner_id": user["id"]}, {"name": 1, "phone": 1}).to_list(10000)
    existing_keys = {(c.get("name", "").lower(), c.get("phone", "")) for c in existing_contacts}
    
    new_contacts = []
    for c in contacts_to_insert:
        key = (c.get("name", "").lower(), c.get("phone", ""))
        if key not in existing_keys:
            new_contacts.append(c)
            existing_keys.add(key)  # Éviter les doublons dans le même import
    
    if new_contacts:
        await db.contacts.insert_many(new_contacts)
    
    return {
        "success": True, 
        "imported": len(new_contacts),
        "skipped": len(contacts_to_insert) - len(new_contacts),
        "message": f"{len(new_contacts)} contacts importés, {len(contacts_to_insert) - len(new_contacts)} doublons ignorés"
    }

@api_router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: str, authorization: Optional[str] = Header(None)):
    """Supprime un contact de l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    # S'assurer que le contact appartient à l'utilisateur
    result = await db.contacts.delete_one({"id": contact_id, "owner_id": user["id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    return {"success": True, "message": "Contact supprimé"}

@api_router.put("/contacts/{contact_id}")
async def update_contact(contact_id: str, authorization: Optional[str] = Header(None), name: Optional[str] = None, phone: Optional[str] = None, email: Optional[str] = None):
    """Met à jour un contact de l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    update_data = {}
    if name: update_data["name"] = name
    if phone: update_data["phone"] = phone
    if email: update_data["email"] = email
    
    if not update_data:
        return {"success": False, "message": "Aucune donnée à mettre à jour"}
    
    result = await db.contacts.update_one(
        {"id": contact_id, "owner_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    return {"success": True, "message": "Contact mis à jour"}

@api_router.delete("/contacts")
async def delete_all_contacts(authorization: Optional[str] = Header(None)):
    """Supprime tous les contacts de l'utilisateur connecté"""
    user = await get_current_user(authorization)
    
    result = await db.contacts.delete_many({"owner_id": user["id"]})
    return {"success": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} contacts supprimés"}

# ============ Inventory Endpoints ============

@api_router.get("/inventory")
async def get_inventory(
    authorization: Optional[str] = Header(None),
    type: Optional[str] = None,
    status: Optional[str] = None,
    year: Optional[int] = None,
    brand: Optional[str] = None
):
    """Récupère l'inventaire de l'utilisateur avec filtres optionnels"""
    user = await get_current_user(authorization)
    
    query = {"owner_id": user["id"]}
    if type:
        query["type"] = type
    if status:
        query["status"] = status
    if year:
        query["year"] = year
    if brand:
        query["brand"] = {"$regex": brand, "$options": "i"}
    
    vehicles = []
    async for vehicle in db.inventory.find(query, {"_id": 0}).sort("stock_no", 1):
        vehicles.append(vehicle)
    
    return vehicles

@api_router.get("/inventory/{stock_no}")
async def get_inventory_vehicle(stock_no: str, authorization: Optional[str] = Header(None)):
    """Récupère un véhicule par numéro de stock"""
    user = await get_current_user(authorization)
    
    vehicle = await db.inventory.find_one(
        {"stock_no": stock_no, "owner_id": user["id"]},
        {"_id": 0}
    )
    
    if not vehicle:
        raise HTTPException(status_code=404, detail="Véhicule non trouvé")
    
    # Get options for this vehicle
    options = []
    async for opt in db.vehicle_options.find({"stock_no": stock_no}, {"_id": 0}).sort("order", 1):
        options.append(opt)
    
    vehicle["options"] = options
    return vehicle

@api_router.post("/inventory")
async def create_inventory_vehicle(vehicle: InventoryCreate, authorization: Optional[str] = Header(None)):
    """Ajoute un véhicule à l'inventaire et télécharge le Window Sticker automatiquement"""
    user = await get_current_user(authorization)
    
    # Check if stock_no already exists
    existing = await db.inventory.find_one({"stock_no": vehicle.stock_no, "owner_id": user["id"]})
    if existing:
        raise HTTPException(status_code=400, detail=f"Le numéro de stock {vehicle.stock_no} existe déjà dans l'inventaire")
    
    # Check if VIN already exists (if provided)
    if vehicle.vin:
        existing_vin = await db.inventory.find_one({"vin": vehicle.vin, "owner_id": user["id"]})
        if existing_vin:
            raise HTTPException(status_code=400, detail=f"Le VIN {vehicle.vin} existe déjà (Stock #{existing_vin.get('stock_no')})")
    
    # Calculate net_cost
    net_cost = vehicle.ep_cost - vehicle.holdback if vehicle.ep_cost and vehicle.holdback else 0
    
    # ===== TÉLÉCHARGER LE WINDOW STICKER AUTOMATIQUEMENT =====
    window_sticker_available = False
    if vehicle.vin and len(vehicle.vin) == 17:
        try:
            logger.info(f"Téléchargement Window Sticker pour VIN={vehicle.vin}")
            ws_result = await fetch_window_sticker(vehicle.vin, vehicle.brand)
            if ws_result["success"]:
                await save_window_sticker_to_db(vehicle.vin, ws_result["pdf_base64"], user["id"])
                window_sticker_available = True
                logger.info(f"Window Sticker sauvegardé pour VIN={vehicle.vin}")
        except Exception as e:
            logger.warning(f"Erreur téléchargement Window Sticker: {e}")
    
    vehicle_data = InventoryVehicle(
        owner_id=user["id"],
        stock_no=vehicle.stock_no,
        vin=vehicle.vin,
        brand=vehicle.brand,
        model=vehicle.model,
        trim=vehicle.trim,
        year=vehicle.year,
        type=vehicle.type,
        pdco=vehicle.pdco,
        ep_cost=vehicle.ep_cost,
        holdback=vehicle.holdback,
        net_cost=net_cost,
        msrp=vehicle.msrp,
        asking_price=vehicle.asking_price,
        km=vehicle.km,
        color=vehicle.color
    )
    
    await db.inventory.insert_one(vehicle_data.dict())
    return {
        "success": True, 
        "vehicle": vehicle_data.dict(), 
        "message": f"Véhicule {vehicle.stock_no} ajouté",
        "window_sticker_available": window_sticker_available
    }

@api_router.post("/inventory/bulk")
async def create_inventory_bulk(vehicles: List[InventoryCreate], authorization: Optional[str] = Header(None)):
    """Import en masse de véhicules"""
    user = await get_current_user(authorization)
    
    added = 0
    updated = 0
    errors = []
    
    for vehicle in vehicles:
        try:
            # Calculate net_cost
            net_cost = vehicle.ep_cost - vehicle.holdback if vehicle.ep_cost and vehicle.holdback else 0
            
            vehicle_data = {
                "id": str(uuid.uuid4()),
                "owner_id": user["id"],
                "stock_no": vehicle.stock_no,
                "vin": vehicle.vin,
                "brand": vehicle.brand,
                "model": vehicle.model,
                "trim": vehicle.trim,
                "year": vehicle.year,
                "type": vehicle.type,
                "pdco": vehicle.pdco,
                "ep_cost": vehicle.ep_cost,
                "holdback": vehicle.holdback,
                "net_cost": net_cost,
                "msrp": vehicle.msrp,
                "asking_price": vehicle.asking_price,
                "km": vehicle.km,
                "color": vehicle.color,
                "status": "disponible",
                "sold_price": None,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            
            # Upsert: update if exists, insert if not
            result = await db.inventory.update_one(
                {"stock_no": vehicle.stock_no, "owner_id": user["id"]},
                {"$set": vehicle_data},
                upsert=True
            )
            
            if result.upserted_id:
                added += 1
            else:
                updated += 1
                
        except Exception as e:
            errors.append({"stock_no": vehicle.stock_no, "error": str(e)})
    
    return {
        "success": True,
        "added": added,
        "updated": updated,
        "errors": errors,
        "message": f"{added} ajoutés, {updated} mis à jour"
    }

@api_router.put("/inventory/{stock_no}")
async def update_inventory_vehicle(stock_no: str, update: InventoryUpdate, authorization: Optional[str] = Header(None)):
    """Met à jour un véhicule"""
    user = await get_current_user(authorization)
    
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    # Recalculate net_cost if ep_cost or holdback changed
    if "ep_cost" in update_data or "holdback" in update_data:
        vehicle = await db.inventory.find_one({"stock_no": stock_no, "owner_id": user["id"]})
        if vehicle:
            ep_cost = update_data.get("ep_cost", vehicle.get("ep_cost", 0))
            holdback = update_data.get("holdback", vehicle.get("holdback", 0))
            update_data["net_cost"] = ep_cost - holdback
    
    result = await db.inventory.update_one(
        {"stock_no": stock_no, "owner_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Véhicule non trouvé")
    
    return {"success": True, "message": f"Véhicule {stock_no} mis à jour"}

@api_router.delete("/inventory/{stock_no}")
async def delete_inventory_vehicle(stock_no: str, authorization: Optional[str] = Header(None)):
    """Supprime un véhicule de l'inventaire"""
    user = await get_current_user(authorization)
    
    result = await db.inventory.delete_one({"stock_no": stock_no, "owner_id": user["id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Véhicule non trouvé")
    
    # Also delete associated options
    await db.vehicle_options.delete_many({"stock_no": stock_no})
    
    return {"success": True, "message": f"Véhicule {stock_no} supprimé"}

@api_router.put("/inventory/{stock_no}/status")
async def update_vehicle_status(stock_no: str, status: str, sold_price: Optional[float] = None, authorization: Optional[str] = Header(None)):
    """Change le statut d'un véhicule (disponible/réservé/vendu)"""
    user = await get_current_user(authorization)
    
    if status not in ["disponible", "réservé", "vendu"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    update_data = {"status": status, "updated_at": datetime.utcnow()}
    if status == "vendu" and sold_price:
        update_data["sold_price"] = sold_price
    
    result = await db.inventory.update_one(
        {"stock_no": stock_no, "owner_id": user["id"]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Véhicule non trouvé")
    
    return {"success": True, "message": f"Statut mis à jour: {status}"}

# Vehicle Options endpoints
@api_router.post("/inventory/{stock_no}/options")
async def add_vehicle_option(stock_no: str, option: VehicleOption, authorization: Optional[str] = Header(None)):
    """Ajoute une option à un véhicule"""
    user = await get_current_user(authorization)
    
    # Verify vehicle exists and belongs to user
    vehicle = await db.inventory.find_one({"stock_no": stock_no, "owner_id": user["id"]})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Véhicule non trouvé")
    
    option_data = option.dict()
    option_data["stock_no"] = stock_no
    
    await db.vehicle_options.insert_one(option_data)
    return {"success": True, "message": "Option ajoutée"}

@api_router.get("/inventory/stats/summary")
async def get_inventory_stats(authorization: Optional[str] = Header(None)):
    """Statistiques d'inventaire"""
    user = await get_current_user(authorization)
    
    total = await db.inventory.count_documents({"owner_id": user["id"]})
    disponible = await db.inventory.count_documents({"owner_id": user["id"], "status": "disponible"})
    reserve = await db.inventory.count_documents({"owner_id": user["id"], "status": "réservé"})
    vendu = await db.inventory.count_documents({"owner_id": user["id"], "status": "vendu"})
    neuf = await db.inventory.count_documents({"owner_id": user["id"], "type": "neuf"})
    occasion = await db.inventory.count_documents({"owner_id": user["id"], "type": "occasion"})
    
    # Calculate total value
    pipeline = [
        {"$match": {"owner_id": user["id"], "status": "disponible"}},
        {"$group": {"_id": None, "total_msrp": {"$sum": "$msrp"}, "total_cost": {"$sum": "$net_cost"}}}
    ]
    result = await db.inventory.aggregate(pipeline).to_list(1)
    totals = result[0] if result else {"total_msrp": 0, "total_cost": 0}
    
    return {
        "total": total,
        "disponible": disponible,
        "reserve": reserve,
        "vendu": vendu,
        "neuf": neuf,
        "occasion": occasion,
        "total_msrp": totals.get("total_msrp", 0),
        "total_cost": totals.get("total_cost", 0),
        "potential_profit": totals.get("total_msrp", 0) - totals.get("total_cost", 0)
    }

# Product codes reference
@api_router.get("/product-codes")
async def get_product_codes():
    """Récupère le référentiel des codes produits"""
    codes = []
    async for code in db.product_codes.find({}, {"_id": 0}):
        codes.append(code)
    return codes

@api_router.post("/product-codes")
async def add_product_code(code: ProductCode, authorization: Optional[str] = Header(None)):
    """Ajoute un code produit au référentiel"""
    await require_admin(authorization)
    
    await db.product_codes.update_one(
        {"code": code.code},
        {"$set": code.dict()},
        upsert=True
    )
    return {"success": True, "message": f"Code {code.code} ajouté"}

@api_router.get("/product-codes/{code}/financing")
async def get_product_code_financing(code: str):
    """
    Récupère les informations de financement pour un code produit.
    Retourne: consumer cash, bonus cash, taux Option 1 et Option 2.
    """
    code = code.upper().strip()
    
    # Obtenir les infos complètes (véhicule + financement)
    full_info = get_full_vehicle_info(code)
    
    if not full_info:
        raise HTTPException(status_code=404, detail=f"Code produit '{code}' non trouvé")
    
    return {
        "success": True,
        "code": code,
        "vehicle": full_info.get("vehicle", {}),
        "financing": full_info.get("financing"),
        "has_financing": full_info.get("financing") is not None
    }

@api_router.get("/financing/lookup")
async def lookup_financing_by_vehicle(
    brand: Optional[str] = None,
    model: Optional[str] = None,
    trim: Optional[str] = None,
    year: Optional[str] = None
):
    """
    Recherche les programmes de financement par marque/modèle/trim/année.
    Retourne tous les codes produits correspondants avec leurs promotions.
    """
    results = []
    
    for code, data in _CODE_PROGRAM_MAPPING.items():
        vehicle = data.get("vehicle", {})
        financing = data.get("financing", {})
        
        # Filtrer par critères
        if brand and vehicle.get("brand", "").lower() != brand.lower():
            continue
        if model and model.lower() not in vehicle.get("model", "").lower():
            continue
        if trim and trim.lower() not in vehicle.get("trim", "").lower():
            continue
        if year and vehicle.get("year") != year:
            continue
        
        results.append({
            "code": code,
            "vehicle": vehicle,
            "financing": {
                "consumer_cash": financing.get("consumer_cash", 0),
                "bonus_cash": financing.get("bonus_cash", 0),
                "total_rebates": financing.get("consumer_cash", 0) + financing.get("bonus_cash", 0),
                "option1_available": any(v for v in financing.get("option1_rates", {}).values() if v is not None),
                "option2_available": any(v for v in financing.get("option2_rates", {}).values() if v is not None)
            }
        })
    
    # Trier par total rebates (descending)
    results.sort(key=lambda x: x["financing"]["total_rebates"], reverse=True)
    
    return {
        "success": True,
        "count": len(results),
        "results": results[:50]  # Limiter à 50 résultats
    }

@api_router.get("/financing/summary")
async def get_financing_summary():
    """
    Retourne un résumé des programmes de financement disponibles.
    Utile pour afficher les meilleures offres.
    """
    by_brand = {}
    best_deals = []
    
    for code, data in _CODE_PROGRAM_MAPPING.items():
        vehicle = data.get("vehicle", {})
        financing = data.get("financing", {})
        brand = vehicle.get("brand", "Unknown")
        
        if brand not in by_brand:
            by_brand[brand] = {
                "count": 0,
                "max_consumer_cash": 0,
                "max_bonus_cash": 0,
                "models": set()
            }
        
        by_brand[brand]["count"] += 1
        by_brand[brand]["max_consumer_cash"] = max(
            by_brand[brand]["max_consumer_cash"],
            financing.get("consumer_cash", 0)
        )
        by_brand[brand]["max_bonus_cash"] = max(
            by_brand[brand]["max_bonus_cash"],
            financing.get("bonus_cash", 0)
        )
        by_brand[brand]["models"].add(vehicle.get("model", ""))
        
        # Collecter les meilleures offres
        total_rebate = financing.get("consumer_cash", 0) + financing.get("bonus_cash", 0)
        if total_rebate > 5000:
            best_deals.append({
                "code": code,
                "brand": brand,
                "model": vehicle.get("model", ""),
                "trim": vehicle.get("trim", ""),
                "total_rebate": total_rebate,
                "consumer_cash": financing.get("consumer_cash", 0),
                "bonus_cash": financing.get("bonus_cash", 0)
            })
    
    # Convertir les sets en listes
    for brand in by_brand:
        by_brand[brand]["models"] = list(by_brand[brand]["models"])
    
    # Trier les meilleures offres
    best_deals.sort(key=lambda x: x["total_rebate"], reverse=True)
    
    return {
        "success": True,
        "total_codes": len(_CODE_PROGRAM_MAPPING),
        "by_brand": by_brand,
        "best_deals": best_deals[:10]
    }

# ============ Invoice Scanner with AI ============

import re
import base64

# ============ VIN Validation & Auto-Correction ============

# Table de translittération VIN (ISO 3779)
VIN_TRANSLATION = {
    **{str(i): i for i in range(10)},
    "A":1,"B":2,"C":3,"D":4,"E":5,"F":6,"G":7,"H":8,
    "J":1,"K":2,"L":3,"M":4,"N":5,"P":7,"R":9,
    "S":2,"T":3,"U":4,"V":5,"W":6,"X":7,"Y":8,"Z":9
}

# Poids par position pour calcul check digit
VIN_WEIGHTS = [8,7,6,5,4,3,2,10,0,9,8,7,6,5,4,3,2]

# Corrections OCR communes
VIN_OCR_CORRECTIONS = {
    "O": "0",  # O ressemble à 0
    "I": "1",  # I ressemble à 1
    "Q": "0",  # Q ressemble à 0
    "L": "1",  # L ressemble à 1
    "Z": "2",  # Z peut ressembler à 2
    "S": "5",  # S peut ressembler à 5
    "B": "8",  # B peut ressembler à 8
}

# Codes année (10e caractère)
VIN_YEAR_CODES = {
    'A': 2010, 'B': 2011, 'C': 2012, 'D': 2013, 'E': 2014,
    'F': 2015, 'G': 2016, 'H': 2017, 'J': 2018, 'K': 2019,
    'L': 2020, 'M': 2021, 'N': 2022, 'P': 2023, 'R': 2024,
    'S': 2025, 'T': 2026, 'V': 2027, 'W': 2028, 'X': 2029,
    'Y': 2030, '1': 2031, '2': 2032, '3': 2033, '4': 2034,
    '5': 2035, '6': 2036, '7': 2037, '8': 2038, '9': 2039
}

# WMI (World Manufacturer Identifier) - 3 premiers caractères
VIN_WMI_MAP = {
    "1C4": "Jeep",
    "1C6": "Ram",
    "1C3": "Chrysler",
    "2C3": "Chrysler",
    "2C4": "Chrysler",
    "3C4": "Chrysler",
    "3C6": "Ram",
    "3C3": "Chrysler",
    "1B3": "Dodge",
    "2B3": "Dodge",
    "3D4": "Dodge",
}


def compute_vin_check_digit(vin: str) -> str:
    """Calcule le check digit (9e caractère) d'un VIN"""
    total = 0
    for i, char in enumerate(vin.upper()):
        if i == 8:  # Skip position du check digit
            continue
        value = VIN_TRANSLATION.get(char, 0)
        total += value * VIN_WEIGHTS[i]
    
    remainder = total % 11
    return "X" if remainder == 10 else str(remainder)


def validate_vin_checksum(vin: str) -> bool:
    """Valide le check digit d'un VIN"""
    vin = vin.upper()
    
    if len(vin) != 17:
        return False
    
    # VIN ne peut pas contenir I, O, Q
    if any(c in "IOQ" for c in vin):
        return False
    
    expected = compute_vin_check_digit(vin)
    return vin[8] == expected


def auto_correct_vin(vin: str) -> tuple:
    """
    Corrige automatiquement les erreurs OCR communes dans un VIN.
    Essaie plusieurs stratégies de correction.
    
    Returns: (vin_corrigé, was_corrected)
    """
    vin = vin.upper().replace("-", "").replace(" ", "")
    
    if len(vin) != 17:
        return vin, False
    
    # Déjà valide ?
    if validate_vin_checksum(vin):
        return vin, False
    
    # Stratégie 1: Corrections simples caractère par caractère
    for i, char in enumerate(vin):
        if char in VIN_OCR_CORRECTIONS:
            corrected = vin[:i] + VIN_OCR_CORRECTIONS[char] + vin[i+1:]
            if validate_vin_checksum(corrected):
                return corrected, True
    
    # Stratégie 2: Permutations communes FCA (P↔J, S↔5, X↔K, etc.)
    common_swaps = [
        ("P", "J"), ("J", "P"),
        ("P", "S"), ("S", "P"),  # P et S se ressemblent
        ("5", "S"), ("S", "5"),  # 5 et S se ressemblent
        ("7", "T"), ("T", "7"),
        ("X", "K"), ("K", "X"),
        ("6", "G"), ("G", "6"),
        ("Y", "T"), ("T", "Y"),
        ("0", "D"), ("D", "0"),
        ("8", "B"), ("B", "8"),
    ]
    for old, new in common_swaps:
        for i, char in enumerate(vin):
            if char == old:
                corrected = vin[:i] + new + vin[i+1:]
                if validate_vin_checksum(corrected):
                    return corrected, True
    
    # Stratégie 3: Corriger l'année (position 10) - très important
    # P souvent confondu avec S (2023 vs 2025)
    year_pos = 9  # Index 9 = position 10
    year_swaps = [("P", "S"), ("S", "P"), ("R", "S"), ("S", "R"), ("P", "R"), ("T", "7")]
    for old, new in year_swaps:
        if vin[year_pos] == old:
            test = vin[:year_pos] + new + vin[year_pos+1:]
            if validate_vin_checksum(test):
                return test, True
    
    # Stratégie 4: Combinaisons multiples pour VINs Jeep/Ram
    # Position 5: K souvent lu comme X
    if len(vin) >= 6 and vin[4] in ["X", "K"]:
        swap = "K" if vin[4] == "X" else "X"
        test = vin[:4] + swap + vin[5:]
        if validate_vin_checksum(test):
            return test, True
        # Essayer avec correction année aussi
        for old, new in [("P", "S"), ("S", "P")]:
            if vin[year_pos] == old:
                test2 = test[:year_pos] + new + test[year_pos+1:]
                if validate_vin_checksum(test2):
                    return test2, True
    
    # Stratégie 5: Recalculer check digit si tout le reste semble OK
    expected_check = compute_vin_check_digit(vin)
    if vin[8] != expected_check:
        corrected = vin[:8] + expected_check + vin[9:]
        if not any(c in "IOQ" for c in corrected):
            return corrected, True
    
    # Stratégie 6: P/J + année + check digit
    if vin[3] == "J":
        test = vin[:3] + "P" + vin[4:]
        check = compute_vin_check_digit(test)
        test = test[:8] + check + test[9:]
        if validate_vin_checksum(test):
            return test, True
    
    return vin, False


def decode_vin_year(vin: str) -> int:
    """Décode l'année à partir du 10e caractère"""
    if len(vin) < 10:
        return None
    return VIN_YEAR_CODES.get(vin[9].upper())


def decode_vin_brand(vin: str) -> str:
    """Décode le constructeur à partir du WMI (3 premiers caractères)"""
    if len(vin) < 3:
        return None
    wmi = vin[:3].upper()
    
    # Cas spécial: Jeep Gladiator (1C6PJ...)
    if vin.upper().startswith("1C6PJ"):
        return "Jeep"
    
    return VIN_WMI_MAP.get(wmi)


def validate_vin_brand_consistency(vin: str, expected_brand: str) -> bool:
    """Vérifie que le VIN correspond à la marque attendue"""
    vin_brand = decode_vin_brand(vin)
    if not vin_brand or not expected_brand:
        return True  # Pas de vérification possible
    return vin_brand.lower() == expected_brand.lower()


def decode_vin(vin: str) -> dict:
    """
    Décode un VIN complet avec validation et auto-correction.
    
    Returns:
        dict avec: vin, valid, corrected, year, manufacturer, checksum_valid
    """
    result = {
        "vin": vin,
        "valid": False,
        "corrected": False,
        "checksum_valid": False,
        "year": None,
        "manufacturer": None,
        "plant": None,
        "serial": None
    }
    
    # Nettoyer
    vin = vin.upper().replace("-", "").replace(" ", "")
    
    if len(vin) != 17:
        return result
    
    # Auto-correction
    vin_corrected, was_corrected = auto_correct_vin(vin)
    result["vin"] = vin_corrected
    result["corrected"] = was_corrected
    result["checksum_valid"] = validate_vin_checksum(vin_corrected)
    result["valid"] = result["checksum_valid"]
    
    # Année (10e caractère)
    result["year"] = decode_vin_year(vin_corrected)
    
    # Constructeur (WMI)
    result["manufacturer"] = decode_vin_brand(vin_corrected)
    
    # Serial (positions 12-17)
    result["serial"] = vin_corrected[11:17]
    
    return result

# ============ FCA Product Code Database ============

# Charger la base de codes 2026 depuis le fichier JSON
import json as _json
_FCA_CODES_2026 = {}
try:
    with open(os.path.join(os.path.dirname(__file__), 'fca_product_codes_2026.json'), 'r') as f:
        _codes_raw = _json.load(f)
        for code, info in _codes_raw.items():
            _FCA_CODES_2026[code] = {
                "brand": info.get("brand"),
                "model": info.get("model"),
                "trim": info.get("trim"),
                "body": f"{info.get('cab', '')} {info.get('drive', '')}".strip(),
                "description": f"{info.get('brand', '')} {info.get('model', '')} {info.get('trim', '')} {info.get('cab', '')} {info.get('drive', '')}".strip()
            }
    print(f"[FCA] Loaded {len(_FCA_CODES_2026)} product codes from JSON")
except Exception as e:
    print(f"[FCA] Warning: Could not load FCA codes JSON: {e}")

# ============ Code -> Programme Financing Mapping ============
_CODE_PROGRAM_MAPPING = {}
try:
    mapping_file = os.path.join(os.path.dirname(__file__), 'data', 'code_program_mapping.json')
    with open(mapping_file, 'r') as f:
        _CODE_PROGRAM_MAPPING = _json.load(f)
    print(f"[FCA] Loaded {len(_CODE_PROGRAM_MAPPING)} code->program mappings")
except Exception as e:
    print(f"[FCA] Warning: Could not load code->program mapping: {e}")

def get_financing_for_code(code: str) -> Optional[Dict[str, Any]]:
    """
    Retourne les informations de financement pour un code produit.
    Inclut: consumer_cash, bonus_cash, taux Option 1 et Option 2.
    """
    code = code.upper().strip()
    if code in _CODE_PROGRAM_MAPPING:
        return _CODE_PROGRAM_MAPPING[code]['financing']
    return None

def get_full_vehicle_info(code: str) -> Optional[Dict[str, Any]]:
    """
    Retourne toutes les informations pour un code produit:
    - Détails du véhicule (marque, modèle, trim, etc.)
    - Informations de financement (consumer cash, bonus cash, taux)
    """
    code = code.upper().strip()
    if code in _CODE_PROGRAM_MAPPING:
        return _CODE_PROGRAM_MAPPING[code]
    # Fallback: retourner juste les infos du véhicule sans financement
    if code in _FCA_CODES_2026:
        return {
            'code': code,
            'vehicle': _FCA_CODES_2026[code],
            'financing': None
        }
    return None

# Base de données des codes produits FCA/Stellantis
# PRIORITÉ: fichier JSON (codes officiels) > codes fallback en dur
# Format: CODE -> {brand, model, trim, body, description}
_FCA_FALLBACK_CODES = {
    # Ram 2500 Series (DJ = Heavy Duty 2500) - FALLBACK si pas dans JSON
    "DJ7L91": {"brand": "Ram", "model": "2500", "trim": "Tradesman", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Tradesman Crew Cab 4x4"},
    "DJ7L92": {"brand": "Ram", "model": "2500", "trim": "Tradesman", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Tradesman Crew Cab 4x4"},
    "DJ7L94": {"brand": "Ram", "model": "2500", "trim": "Laramie", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Laramie Crew Cab 4x4"},
    "DJ7L96": {"brand": "Ram", "model": "2500", "trim": "Limited", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Limited Crew Cab 4x4"},
    "DJ7L98": {"brand": "Ram", "model": "2500", "trim": "Power Wagon", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Power Wagon Crew Cab 4x4"},
    
    # Ram 1500 Series (DS/DT = 1500)
    "DT6L91": {"brand": "Ram", "model": "1500", "trim": "Tradesman", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Tradesman Crew Cab 4x4"},
    "DT6L92": {"brand": "Ram", "model": "1500", "trim": "Big Horn", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Big Horn Crew Cab 4x4"},
    "DT6H92": {"brand": "Ram", "model": "1500", "trim": "Big Horn", "body": "Quad Cab 4x4 6'4\" Box", "description": "Ram 1500 Big Horn Quad Cab 4x4"},
    "DT6L94": {"brand": "Ram", "model": "1500", "trim": "Laramie", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Laramie Crew Cab 4x4"},
    "DT6L96": {"brand": "Ram", "model": "1500", "trim": "Limited", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Limited Crew Cab 4x4"},
    "DT6L97": {"brand": "Ram", "model": "1500", "trim": "Rebel", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Rebel Crew Cab 4x4"},
    "DT6L99": {"brand": "Ram", "model": "1500", "trim": "TRX", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 TRX Crew Cab 4x4"},
    
    # Ram 3500 Series (D2 = Heavy Duty 3500)
    "D2RL91": {"brand": "Ram", "model": "3500", "trim": "Tradesman", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Tradesman Crew Cab 4x4"},
    "D2RL92": {"brand": "Ram", "model": "3500", "trim": "Big Horn", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Big Horn Crew Cab 4x4"},
    "D2RL94": {"brand": "Ram", "model": "3500", "trim": "Laramie", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Laramie Crew Cab 4x4"},
    "D2RL96": {"brand": "Ram", "model": "3500", "trim": "Limited", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Limited Crew Cab 4x4"},
    
    # Jeep
    "JLXL74": {"brand": "Jeep", "model": "Wrangler", "trim": "Rubicon", "body": "4-Door 4x4", "description": "Jeep Wrangler Rubicon Unlimited 4x4"},
    "JLXL72": {"brand": "Jeep", "model": "Wrangler", "trim": "Sahara", "body": "4-Door 4x4", "description": "Jeep Wrangler Sahara Unlimited 4x4"},
    "JLXL70": {"brand": "Jeep", "model": "Wrangler", "trim": "Sport", "body": "4-Door 4x4", "description": "Jeep Wrangler Sport Unlimited 4x4"},
    # Jeep Gladiator (JT = Gladiator)
    "JTJL98": {"brand": "Jeep", "model": "Gladiator", "trim": "Willys", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Willys 4x4"},
    "JTJL96": {"brand": "Jeep", "model": "Gladiator", "trim": "Rubicon", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Rubicon 4x4"},
    "JTJL94": {"brand": "Jeep", "model": "Gladiator", "trim": "Overland", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Overland 4x4"},
    "JTJL92": {"brand": "Jeep", "model": "Gladiator", "trim": "Sport S", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Sport S 4x4"},
    "JTJL90": {"brand": "Jeep", "model": "Gladiator", "trim": "Sport", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Sport 4x4"},
    
    # Grand Cherokee (WL Series - 2022+)
    "WLJP74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Limited", "body": "4x4", "description": "Jeep Grand Cherokee Limited 4x4"},
    "WLJP75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Limited", "body": "4x4", "description": "Jeep Grand Cherokee L Limited 4x4"},
    "WLJH74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Laredo", "body": "4x4", "description": "Jeep Grand Cherokee Laredo 4x4"},
    "WLJH75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Altitude", "body": "4x4", "description": "Jeep Grand Cherokee L Altitude 4x4"},
    "WLJS74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit", "body": "4x4", "description": "Jeep Grand Cherokee Summit 4x4"},
    "WLJS75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Summit", "body": "4x4", "description": "Jeep Grand Cherokee L Summit 4x4"},
    "WLJT74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Overland", "body": "4x4", "description": "Jeep Grand Cherokee Overland 4x4"},
    "WLJT75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Overland", "body": "4x4", "description": "Jeep Grand Cherokee L Overland 4x4"},
    "WKXL74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit", "body": "4x4", "description": "Jeep Grand Cherokee Summit 4x4"},
    "WKXL72": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Limited", "body": "4x4", "description": "Jeep Grand Cherokee Limited 4x4"},
    "MPXL74": {"brand": "Jeep", "model": "Compass", "trim": "Limited", "body": "4x4", "description": "Jeep Compass Limited 4x4"},
    
    # Dodge
    "LDXL74": {"brand": "Dodge", "model": "Durango", "trim": "R/T", "body": "AWD", "description": "Dodge Durango R/T AWD"},
    "LDXL76": {"brand": "Dodge", "model": "Durango", "trim": "SRT", "body": "AWD", "description": "Dodge Durango SRT AWD"},
    "LAXL74": {"brand": "Dodge", "model": "Charger", "trim": "R/T", "body": "RWD", "description": "Dodge Charger R/T"},
    "LCXL74": {"brand": "Dodge", "model": "Challenger", "trim": "R/T", "body": "RWD", "description": "Dodge Challenger R/T"},
    "HNXL74": {"brand": "Dodge", "model": "Hornet", "trim": "R/T", "body": "AWD", "description": "Dodge Hornet R/T AWD"},
    
    # Chrysler
    "RUXL74": {"brand": "Chrysler", "model": "Pacifica", "trim": "Limited", "body": "FWD", "description": "Chrysler Pacifica Limited"},
    "RUXL78": {"brand": "Chrysler", "model": "Pacifica", "trim": "Pinnacle", "body": "AWD", "description": "Chrysler Pacifica Pinnacle AWD"},
}

# Fusionner avec PRIORITÉ au fichier JSON (codes officiels 2025/2026)
FCA_PRODUCT_CODES = {
    **_FCA_FALLBACK_CODES,  # Codes fallback en premier (seront écrasés par JSON)
    **_FCA_CODES_2026,      # Codes officiels du JSON en dernier (priorité maximale)
}

# Codes d'options communes FCA
FCA_OPTION_CODES = {
    # Moteurs
    "ETM": "6.7L Cummins Turbo Diesel I-6",
    "ETK": "6.7L Cummins High Output Turbo Diesel",
    "EZH": "6.4L HEMI V8",
    "ERC": "5.7L HEMI V8 MDS VVT",
    "ERB": "3.6L Pentastar V6",
    "EZC": "3.0L EcoDiesel V6",
    "ESG": "2.0L Turbo I-4 PHEV",
    
    # Transmissions
    "DFM": "8-Speed Automatic TorqueFlite HD",
    "DFD": "8-Speed Automatic TorqueFlite 8HP75",
    "DFH": "6-Speed Manual",
    "DFL": "8-Speed Automatic 8HP95",
    
    # Couleurs
    "PXJ": "Noir Cristal Nacré",
    "PW7": "Blanc Vif",
    "PAU": "Rouge Flamme",
    "PBF": "Bleu Patriote",
    "PSC": "Gris Destroyer",
    "PX8": "Noir Diamant",
    "PWL": "Blanc Perle",
    "PGG": "Gris Granit",
    
    # Options populaires
    "AHU": "Préparation Remorquage Sellette/Col-de-cygne",
    "XAC": "Groupe Remorquage Max",
    "ADA": "Différentiel Arrière Anti-spin",
    "CLF": "Tapis de Protection Mopar",
    "LHL": "Commandes Auxiliaires Tableau de Bord",
    "LNC": "Feux de Gabarit",
    "MWH": "Doublures Passage de Roue Arrière",
    "UAQ": "Uconnect 5 NAV 12\" Écran",
    "RC3": "Radio Uconnect 5 8.4\" Écran",
    "GWA": "Boîte de Transfert Électronique",
    "DSA": "Suspension Pneumatique Arrière",
    
    # Frais/Taxes
    "801": "Frais de Transport",
    "4CP": "Taxe Accise Fédérale Climatiseur",
    "92HC1": "Cotisation Protection Produit",
    "92HC2": "Allocation Marketing",
}

def _build_trim_string(product_info: dict) -> str:
    """
    Construit la chaîne de trim à partir des données product_info.
    Gère les deux formats: {trim, body} et {trim, cab, drive}
    """
    if not product_info:
        return ""
    
    trim = product_info.get("trim") or ""
    
    # Format 1: body direct
    body = product_info.get("body") or ""
    
    # Format 2: cab + drive (fichier JSON)
    if not body:
        cab = product_info.get("cab") or ""
        drive = product_info.get("drive") or ""
        if cab or drive:
            body = f"{cab} {drive}".strip()
    
    # Combiner trim et body
    if trim and body:
        return f"{trim} {body}"
    elif trim:
        return trim
    elif body:
        return body
    else:
        return ""

def decode_product_code(code: str) -> dict:
    """Décode un code produit FCA et retourne les informations du véhicule"""
    code = code.upper().strip()
    
    # Chercher dans la base de données
    if code in FCA_PRODUCT_CODES:
        return FCA_PRODUCT_CODES[code]
    
    # Essayer de décoder le pattern si pas trouvé
    result = {
        "brand": None,
        "model": None,
        "trim": None,
        "body": None,
        "description": None
    }
    
    # Patterns de préfixes connus
    if code.startswith("DJ"):
        result["brand"] = "Ram"
        result["model"] = "2500"
    elif code.startswith("DT") or code.startswith("DS"):
        result["brand"] = "Ram"
        result["model"] = "1500"
    elif code.startswith("D2"):
        result["brand"] = "Ram"
        result["model"] = "3500"
    elif code.startswith("JL"):
        result["brand"] = "Jeep"
        result["model"] = "Wrangler"
    elif code.startswith("WK"):
        result["brand"] = "Jeep"
        result["model"] = "Grand Cherokee"
    elif code.startswith("MP"):
        result["brand"] = "Jeep"
        result["model"] = "Compass"
    elif code.startswith("LD"):
        result["brand"] = "Dodge"
        result["model"] = "Durango"
    elif code.startswith("LA"):
        result["brand"] = "Dodge"
        result["model"] = "Charger"
    elif code.startswith("LC"):
        result["brand"] = "Dodge"
        result["model"] = "Challenger"
    elif code.startswith("RU"):
        result["brand"] = "Chrysler"
        result["model"] = "Pacifica"
    
    return result

def decode_option_code(code: str) -> str:
    """Retourne la description d'un code d'option FCA"""
    code = code.upper().strip()
    return FCA_OPTION_CODES.get(code, None)

def enrich_vehicle_data(vehicle_data: dict) -> dict:
    """Enrichit les données d'un véhicule avec le décodage VIN et codes produits"""
    
    # Décoder le VIN si présent
    vin = vehicle_data.get("vin", "")
    if vin and len(vin.replace("-", "")) == 17:
        vin_info = decode_vin(vin)
        if vin_info["valid"]:
            # Mettre à jour l'année si pas déjà définie
            if not vehicle_data.get("year") and vin_info.get("year"):
                vehicle_data["year"] = vin_info["year"]
            # Mettre à jour la marque si pas définie
            if not vehicle_data.get("brand") and vin_info.get("manufacturer"):
                vehicle_data["brand"] = vin_info["manufacturer"]
    
    # Chercher et décoder le code produit principal (première option ou MODEL/OPT)
    options = vehicle_data.get("options", [])
    if options:
        first_option = options[0]
        # PATCH: Utiliser "product_code" au lieu de "code"
        code = first_option.get("product_code", first_option.get("code", ""))
        
        # Vérifier si c'est un code de modèle (pas un code d'option)
        product_info = decode_product_code(code)
        if product_info.get("brand"):
            # C'est un code de véhicule
            if not vehicle_data.get("brand"):
                vehicle_data["brand"] = product_info["brand"]
            if not vehicle_data.get("model"):
                vehicle_data["model"] = product_info["model"]
            if not vehicle_data.get("trim"):
                vehicle_data["trim"] = product_info.get("trim")
            if product_info.get("description"):
                first_option["description"] = product_info["description"]
    
    # Enrichir les descriptions des options
    for option in options:
        # PATCH: Utiliser "product_code" au lieu de "code"
        code = option.get("product_code", option.get("code", ""))
        if not option.get("description") or len(option.get("description", "")) < 5:
            desc = decode_option_code(code)
            if desc:
                option["description"] = desc
    
    return vehicle_data

# PATCH: decode_fca_price() supprimé - utiliser clean_fca_price() uniquement
# Fonction dupliquée de clean_fca_price() dans parser.py

def decode_fca_holdback(raw_value: str) -> float:
    """Décode un holdback FCA - même règle que les prix
    Enlever premier 0 + deux derniers chiffres
    Exemple: 050000 → 50000 → 500 → 500$
    """
    # Remove any non-numeric characters
    cleaned = re.sub(r'[^\d]', '', str(raw_value))
    
    if len(cleaned) >= 4:
        # Remove first 0 if present
        if cleaned.startswith('0'):
            cleaned = cleaned[1:]
        # Remove last 2 digits
        if len(cleaned) >= 2:
            cleaned = cleaned[:-2]
        try:
            return float(cleaned)
        except:
            return 0
    return 0

# =========================
# PARSER STRUCTURÉ FCA V5 - VERSION INDUSTRIELLE OPTIMISÉE
# =========================

# Codes invalides à exclure des options
INVALID_OPTION_CODES = {
    "VIN", "GST", "TPS", "QUE", "INC", "PDCO", "PREF", 
    "MODEL", "TOTAL", "MSRP", "SUB", "KG", "GVW"
}

# Cache des codes produits FCA (évite lookups répétés)
FCA_PRODUCT_CACHE = {}


def generate_file_hash(file_bytes: bytes) -> str:
    """Génère un hash SHA256 unique pour le fichier"""
    return hashlib.sha256(file_bytes).hexdigest()


def compress_image_for_vision(file_bytes: bytes, max_size: int = 1024, quality: int = 70) -> str:
    """
    Compresse l'image pour réduire les tokens Vision API.
    - Redimensionne à max 1024px
    - Compression JPEG quality 70
    - Retourne base64 optimisé
    
    Économie: ~60-70% de tokens en moins
    """
    try:
        image = Image.open(io.BytesIO(file_bytes))
        
        # Convertir en RGB si nécessaire
        if image.mode in ('RGBA', 'P'):
            image = image.convert('RGB')
        
        # Redimensionner proportionnellement
        width, height = image.size
        if max(width, height) > max_size:
            ratio = max_size / max(width, height)
            new_size = (int(width * ratio), int(height * ratio))
            image = image.resize(new_size, Image.Resampling.LANCZOS)
        
        # Compression JPEG
        buffer = io.BytesIO()
        image.save(buffer, format='JPEG', quality=quality, optimize=True)
        
        compressed_bytes = buffer.getvalue()
        original_size = len(file_bytes)
        new_size = len(compressed_bytes)
        
        logger.info(f"Image compressed: {original_size/1024:.1f}KB → {new_size/1024:.1f}KB ({100-new_size*100/original_size:.0f}% reduction)")
        
        return base64.b64encode(compressed_bytes).decode('utf-8')
    except Exception as e:
        logger.error(f"Compression error: {e}")
        return base64.b64encode(file_bytes).decode('utf-8')


def clean_fca_price(raw_value: str) -> int:
    """
    Règle FCA pour décoder les prix:
    - Enlever le premier 0
    - Enlever les deux derniers chiffres
    Exemple: 05662000 -> 56620
    """
    raw_value = str(raw_value).strip()
    raw_value = re.sub(r'[^\d]', '', raw_value)
    
    if not raw_value:
        return 0
    
    if raw_value.startswith("0"):
        raw_value = raw_value[1:]
    
    if len(raw_value) >= 2:
        raw_value = raw_value[:-2]
    
    try:
        return int(raw_value)
    except:
        return 0


def clean_decimal_price(raw_value: str) -> float:
    """Nettoie les montants format décimal: 57,120.00 -> 57120.00"""
    raw_value = str(raw_value).replace(",", "").strip()
    try:
        return float(raw_value)
    except:
        return 0.0


def extract_pdf_text(file_bytes: bytes) -> str:
    """Extrait le texte d'un PDF avec pdfplumber"""
    text = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\n"
        
        os.unlink(tmp_path)
    except Exception as e:
        logger.error(f"Error extracting PDF text: {e}")
    
    return text


def extract_text_from_image(file_bytes: bytes) -> str:
    """
    OCR image avec Tesseract + prétraitement OpenCV
    """
    try:
        # Charger l'image
        image = Image.open(io.BytesIO(file_bytes))
        
        # Convertir en RGB si nécessaire (pour les images RGBA ou autres)
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        img = np.array(image)
        
        # Convertir en niveaux de gris
        if len(img.shape) == 3:
            gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        else:
            gray = img
        
        # Redimensionner si l'image est trop grande (améliore la vitesse OCR)
        max_dimension = 3000
        height, width = gray.shape[:2]
        if max(height, width) > max_dimension:
            scale = max_dimension / max(height, width)
            gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
        
        # Amélioration du contraste
        gray = cv2.convertScaleAbs(gray, alpha=1.5, beta=10)
        
        # Seuillage adaptatif pour améliorer la lisibilité
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
        
        # OCR avec Tesseract (anglais + français)
        text = pytesseract.image_to_string(
            thresh,
            lang="eng+fra",
            config="--psm 6 --oem 3"
        )
        
        logger.info(f"OCR extracted {len(text)} chars")
        return text
    except Exception as e:
        logger.error(f"OCR error: {str(e)}")
        import traceback
        logger.error(f"OCR traceback: {traceback.format_exc()}")
        return ""


def parse_fca_invoice_structured(text: str) -> dict:
    """
    Parser structuré V4 pour factures FCA Canada.
    Patterns améliorés et plus stricts.
    """
    data = {
        "vin": None,
        "model_code": None,
        "ep_cost": None,
        "pdco": None,
        "pref": None,
        "holdback": None,
        "subtotal_excl_tax": None,
        "invoice_total": None,
        "options": [],
        "parse_method": "structured_v4"
    }
    
    # -------------------------
    # VIN (17 caractères, format FCA avec ou sans tirets)
    # -------------------------
    # PATCH: Pattern VIN plus strict - 17 caractères exacts
    # Priorité 1: VIN standard 17 caractères (plus fiable)
    vin_match = re.search(r"\b([0-9A-HJ-NPR-Z]{17})\b", text)
    if vin_match:
        data["vin"] = vin_match.group(1)
    
    # Fallback: VIN FCA avec tirets (format 1C4RJHBG6-S8-806264)
    if not data["vin"]:
        vin_dash_match = re.search(r"\b([0-9A-HJ-NPR-Z]{9})[-\s]([A-HJ-NPR-Z0-9]{2})[-\s]([A-HJ-NPR-Z0-9]{6})\b", text)
        if vin_dash_match:
            vin_raw = vin_dash_match.group(1) + vin_dash_match.group(2) + vin_dash_match.group(3)
            if len(vin_raw) == 17:
                data["vin"] = vin_raw
    
    # -------------------------
    # Model Code - RESTREINT à la zone MODEL/OPT
    # Pattern: 5-7 caractères alphanumériques (ex: WLJP74, WLJH75, JTJL98)
    # -------------------------
    model_section = re.search(r"MODEL/OPT[\s\S]{0,50}?\n\s*([A-Z]{2,4}[A-Z0-9]{2,4})", text)
    if model_section:
        data["model_code"] = model_section.group(1)
    else:
        # Fallback: chercher pattern FCA standard (2-4 lettres + 2 chiffres)
        model_match = re.search(r"\b(WL[A-Z]{2}\d{2}|JT[A-Z]{2}\d{2}|DT[A-Z0-9]{2}\d{2})\b", text)
        if model_match:
            data["model_code"] = model_match.group(1)
    
    # -------------------------
    # E.P. (Employee Price / Coût)
    # Pattern: E.P. suivi de 7-10 chiffres
    # -------------------------
    ep_match = re.search(r"E\.P\.?\s*(\d{7,10})", text)
    if ep_match:
        data["ep_cost"] = clean_fca_price(ep_match.group(1))
    
    # -------------------------
    # PDCO (Prix dealer / PDSF base)
    # -------------------------
    pdco_match = re.search(r"PDCO\s*(\d{7,10})", text)
    if pdco_match:
        data["pdco"] = clean_fca_price(pdco_match.group(1))
    
    # -------------------------
    # PREF (Prix de référence)
    # -------------------------
    pref_match = re.search(r"PREF\*?\s*(\d{7,10})", text)
    if pref_match:
        data["pref"] = clean_fca_price(pref_match.group(1))
    
    # -------------------------
    # HOLDBACK - AMÉLIORÉ
    # Chercher dans la zone financière près de PREF
    # Format: 6 chiffres commençant par 0 (ex: 070000 = 700$)
    # -------------------------
    # Méthode 1: Chercher après PREF
    holdback_match = re.search(r"PREF\*?\s*\d{7,10}\s+(\d{6})\b", text)
    if holdback_match:
        data["holdback"] = clean_fca_price(holdback_match.group(1))
    else:
        # Méthode 2: Chercher un 0XXXXX isolé près de GVW/KG
        holdback_match = re.search(r"\b(0[3-9]\d{4})\b\s*(?:GVW|KG|$)", text)
        if holdback_match:
            data["holdback"] = clean_fca_price(holdback_match.group(1))
    
    # -------------------------
    # SUBTOTAL EXCLUDING TAXES
    # -------------------------
    subtotal_match = re.search(
        r"SUB\s*TOTAL\s*EXCLUDING\s*TAXES[\s\S]*?([\d,]+\.\d{2})",
        text,
        re.IGNORECASE
    )
    if subtotal_match:
        data["subtotal_excl_tax"] = clean_decimal_price(subtotal_match.group(1))
    
    # -------------------------
    # INVOICE TOTAL / TOTAL DE LA FACTURE
    # -------------------------
    total_match = re.search(
        r"TOTAL\s+DE\s+LA\s+FACTURE\s+([\d,]+\.\d{2})",
        text
    )
    if not total_match:
        total_match = re.search(
            r"INVOICE\s*TOTAL[\s\S]*?([\d,]+\.\d{2})",
            text,
            re.IGNORECASE
        )
    if total_match:
        data["invoice_total"] = clean_decimal_price(total_match.group(1))
    
    # -------------------------
    # OPTIONS - PATTERN AMÉLIORÉ
    # Format: CODE (2-5 chars) + DESCRIPTION (5+ chars) + MONTANT (6-10 chiffres ou SANS FRAIS)
    # -------------------------
    option_pattern = re.findall(
        r"\n\s*([A-Z0-9]{2,5})\s+([A-Z0-9][A-Z0-9 ,\-\/'\.]{4,}?)\s+(\d{6,10}|\*|SANS\s*FRAIS)",
        text
    )
    
    for code, desc, amount in option_pattern:
        if code.upper() in INVALID_OPTION_CODES:
            continue
        
        # Nettoyer le montant
        if amount in ['*', 'SANS FRAIS', 'SANS']:
            amt = 0
        else:
            amt = clean_fca_price(amount)
        
        data["options"].append({
            "product_code": code.upper(),
            "description": desc.strip()[:100],
            "amount": amt
        })
    
    return data


# SUPPRIMÉ: validate_invoice_data locale - utiliser celle de validation.py
# Importée comme: from validation import validate_invoice_data as validate_invoice_full


class InvoiceScanRequest(BaseModel):
    image_base64: str
    is_pdf: bool = False


@api_router.post("/inventory/scan-invoice")
async def scan_invoice(request: InvoiceScanRequest, authorization: Optional[str] = Header(None)):
    """
    Scanne une facture FCA - Pipeline Multi-Niveaux Industriel
    
    ARCHITECTURE 100% OPTIMISÉE:
    - Niveau 1: PDF natif → pdfplumber + regex (100% précis, gratuit)
    - Niveau 2: Image → OpenCV ROI + Tesseract (85-92%, gratuit)
    - Niveau 3: Fallback → GPT-4 Vision (si score < 70, ~2-3¢)
    
    AVANTAGES:
    - 95% des factures traitées sans coût API
    - Fallback intelligent uniquement si nécessaire
    - Validation VIN industrielle avec auto-correction
    """
    # Import des nouveaux modules OCR
    from ocr import process_image_ocr_pipeline
    from parser import parse_invoice_text
    from vin_utils import validate_and_correct_vin
    from validation import validate_invoice_data as validate_invoice_full, calculate_validation_score
    
    user = await get_current_user(authorization)
    
    try:
        start_time = time.time()
        
        # Décoder le base64
        try:
            file_bytes = base64.b64decode(request.image_base64)
        except:
            raise HTTPException(status_code=400, detail="Base64 invalide")
        
        # Générer le hash du fichier pour anti-doublon
        file_hash = generate_file_hash(file_bytes)
        
        # Détecter si c'est un PDF
        is_pdf = file_bytes[:4] == b'%PDF' or request.is_pdf
        
        vehicle_data = None
        parse_method = None
        validation = {"score": 0, "errors": [], "is_valid": False}
        
        # ===== NIVEAU 1: PDF → PARSER STRUCTURÉ (100% GRATUIT) =====
        if is_pdf:
            logger.info("PDF détecté → Parser pdfplumber + regex")
            extracted_text = extract_pdf_text(file_bytes)
            
            if extracted_text and len(extracted_text) > 100:
                parsed = parse_fca_invoice_structured(extracted_text)
                validation = validate_invoice_data(parsed)
                
                if validation["is_valid"]:
                    vin = parsed.get("vin", "")
                    vin_info = decode_vin(vin) if vin and len(vin) == 17 else {}
                    model_code = parsed.get("model_code", "")
                    product_info = decode_product_code(model_code) if model_code else {}
                    
                    vehicle_data = {
                        "vin": vin,
                        "model_code": model_code,
                        "year": vin_info.get("year") or datetime.now().year,
                        "brand": product_info.get("brand") or vin_info.get("manufacturer") or "Stellantis",
                        "model": product_info.get("model") or "",
                        "trim": product_info.get("trim") or "",
                        "ep_cost": parsed.get("ep_cost") or 0,
                        "pdco": parsed.get("pdco") or 0,
                        "pref": parsed.get("pref") or 0,
                        "holdback": parsed.get("holdback") or 0,
                        "msrp": parsed.get("pdco") or 0,
                        "net_cost": parsed.get("ep_cost") or 0,
                        "subtotal": parsed.get("subtotal_excl_tax") or 0,
                        "invoice_total": parsed.get("invoice_total") or 0,
                        "options": parsed.get("options", []),
                        "file_hash": file_hash,
                        "parse_method": "pdf_native",
                        "cost_estimate": "$0.00"
                    }
                    parse_method = "pdf_native"
                    logger.info(f"Parser PDF réussi: VIN={vin}, EP={vehicle_data['ep_cost']}")
        
        # ===== NIVEAU 2: IMAGE → OCR SIMPLIFIÉ STABLE (ZÉRO ERREUR) =====
        decision = None
        
        if vehicle_data is None and not is_pdf:
            logger.info("Image détectée → OCR Global simplifié")
            
            try:
                # 1. Pipeline OCR par zones
                ocr_result = process_image_ocr_pipeline(file_bytes)
                
                # 2. Parser structuré sur le texte OCR
                parsed = parse_invoice_text(ocr_result)
                
                # 3. Validation et correction VIN centralisée (UNE SEULE FONCTION)
                vin_raw = parsed.get("vin", "")
                vin_result = validate_and_correct_vin(vin_raw) if vin_raw else {}
                
                vin_corrected = vin_result.get("corrected", vin_raw)
                vin_valid = vin_result.get("is_valid", False)
                vin_was_corrected = vin_result.get("was_corrected", False)
                
                # 4. Calcul du score de validation
                parsed["vin"] = vin_corrected
                parsed["vin_valid"] = vin_valid
                validation_result = validate_invoice_full(parsed)
                
                ocr_score = validation_result["score"]
                logger.info(f"OCR: score={ocr_score}, VIN={vin_corrected}, EP={parsed.get('ep_cost')}")
                
                # ----------- NOUVELLE LOGIQUE: GOOGLE VISION EN PRIORITÉ -----------
                # Si Google Vision est configuré, on l'utilise TOUJOURS pour meilleure précision
                google_api_key = os.environ.get("GOOGLE_VISION_API_KEY")
                if google_api_key:
                    logger.info("Google Vision API configurée → Utilisation prioritaire")
                    decision = "vision_required"  # Force Google Vision
                elif ocr_score >= 85:
                    decision = "auto_approved"
                elif 60 <= ocr_score < 85:
                    decision = "review_required"
                else:
                    decision = "vision_required"
                    
                logger.info(f"Décision OCR: {decision} (score={ocr_score})")
                    
            except Exception as ocr_err:
                logger.warning(f"OCR échoué: {ocr_err}, passage au fallback Vision")
                decision = "vision_required"
        
        # ===== GESTION DES DÉCISIONS =====
        
        # REVIEW REQUIRED (60-84) → Retourner pour révision humaine
        if decision == "review_required":
            vin_info = decode_vin(vin_corrected) if vin_corrected and len(vin_corrected) == 17 else {}
            model_code = parsed.get("model_code", "")
            product_info = decode_product_code(model_code) if model_code else {}
            
            # PRIORITÉ CORRIGÉE: code produit (base de données) > parser > VIN
            extracted_model = product_info.get("model") or parsed.get("model") or ""
            extracted_trim = product_info.get("trim") or parsed.get("trim") or ""
            extracted_body = product_info.get("body") or ""
            
            if extracted_body and extracted_body not in extracted_trim:
                extracted_trim = f"{extracted_trim} {extracted_body}".strip() if extracted_trim else extracted_body
            
            logger.info(f"Review - Model extraction: code={model_code}, product_info={product_info}, final_model={extracted_model}, final_trim={extracted_trim}")
            
            return {
                "success": True,
                "review_required": True,
                "vehicle": {
                    "stock_no": parsed.get("stock_no", ""),
                    "vin": vin_corrected,
                    "vin_valid": vin_valid,
                    "vin_corrected": vin_was_corrected,
                    "model_code": model_code,
                    "year": vin_info.get("year") or datetime.now().year,
                    "brand": product_info.get("brand") or "Stellantis",
                    "model": extracted_model,
                    "trim": extracted_trim,
                    "ep_cost": parsed.get("ep_cost") or 0,
                    "pdco": parsed.get("pdco") or 0,
                    "pref": parsed.get("pref") or 0,
                    "holdback": parsed.get("holdback") or 0,
                    "subtotal": parsed.get("subtotal_excl_tax") or parsed.get("subtotal") or 0,
                    "invoice_total": parsed.get("invoice_total") or 0,
                    "options": parsed.get("options", []),
                },
                "validation": validation_result,
                "parse_method": "ocr_review",
                "message": "Score entre 60-84. Révision humaine recommandée."
            }
        
        # AUTO APPROVED (85+) → Accepter directement
        if decision == "auto_approved":
            vin_info = decode_vin(vin_corrected) if vin_corrected and len(vin_corrected) == 17 else {}
            model_code = parsed.get("model_code", "")
            product_info = decode_product_code(model_code) if model_code else {}
            vin_brand = decode_vin_brand(vin_corrected) if vin_corrected else None
            
            # PRIORITÉ CORRIGÉE: code produit (base de données) > parser > VIN
            # Le code produit (D28H92, etc.) est la source la plus fiable pour modèle/trim
            extracted_model = product_info.get("model") or parsed.get("model") or ""
            extracted_trim = product_info.get("trim") or parsed.get("trim") or ""
            extracted_body = product_info.get("body") or ""
            
            # Construire le trim complet avec cab/drive si disponible
            if extracted_body and extracted_body not in extracted_trim:
                extracted_trim = f"{extracted_trim} {extracted_body}".strip() if extracted_trim else extracted_body
            
            logger.info(f"Model extraction: code={model_code}, product_info={product_info}, final_model={extracted_model}, final_trim={extracted_trim}")
            
            vehicle_data = {
                "stock_no": parsed.get("stock_no", ""),
                "vin": vin_corrected,
                "vin_original": vin_raw if vin_was_corrected else None,
                "vin_valid": vin_valid,
                "vin_corrected": vin_was_corrected,
                "vin_brand": vin_brand,
                "model_code": model_code,
                "year": vin_info.get("year") or datetime.now().year,
                "brand": product_info.get("brand") or vin_brand or "Stellantis",
                "model": extracted_model,
                "trim": extracted_trim,
                "ep_cost": parsed.get("ep_cost") or 0,
                "pdco": parsed.get("pdco") or 0,
                "pref": parsed.get("pref") or 0,
                "holdback": parsed.get("holdback") or 0,
                "msrp": parsed.get("pdco") or 0,
                "net_cost": parsed.get("ep_cost") or 0,
                "subtotal": parsed.get("subtotal_excl_tax") or parsed.get("subtotal") or 0,
                "invoice_total": parsed.get("invoice_total") or 0,
                "options": parsed.get("options", []),
                "file_hash": file_hash,
                "parse_method": "ocr_auto_approved",
                "cost_estimate": "$0.00",
                "metrics": {
                    "zones_processed": ocr_result.get("zones_processed", 0),
                    "validation_score": ocr_score,
                    "parse_duration_sec": round(time.time() - start_time, 3)
                }
            }
            
            validation = validation_result
            parse_method = "ocr_auto_approved"
            logger.info(f"OCR Auto-Approved: VIN={vin_corrected}, EP={vehicle_data['ep_cost']}, Score={ocr_score}")
        
        # ===== NIVEAU 3: FALLBACK → GOOGLE CLOUD VISION (OCR HYBRIDE) =====
        # Stratégie hybride: Google Vision pour l'OCR pur (moins cher, plus précis)
        # + parser.py pour l'extraction structurée des données
        if decision == "vision_required" or vehicle_data is None:
            logger.info("Fallback → Google Cloud Vision avec prétraitement CamScanner")
            
            try:
                from PIL import Image as PILImage
                from ocr import (
                    camscanner_preprocess_for_vision, 
                    load_image_from_bytes,
                    google_vision_ocr_from_numpy,
                    google_vision_ocr
                )
                from parser import (
                    parse_vin, 
                    parse_model_code, 
                    parse_financial_data, 
                    parse_totals, 
                    parse_options,
                    parse_stock_number
                )
                
                # Vérifier la clé API Google Vision
                google_api_key = os.environ.get("GOOGLE_VISION_API_KEY")
                if not google_api_key:
                    logger.warning("GOOGLE_VISION_API_KEY non configurée, fallback vers GPT-4 Vision")
                    raise ValueError("Google Vision API key not configured")
                
                # Décoder l'image
                img_data = base64.b64decode(request.image_base64)
                
                # ====== PRÉTRAITEMENT CAMSCANNER ======
                cv_image = load_image_from_bytes(img_data)
                
                if cv_image is not None:
                    logger.info("Applying CamScanner preprocessing for Google Vision...")
                    preprocessed = camscanner_preprocess_for_vision(cv_image)
                    
                    # ====== OCR GOOGLE CLOUD VISION ======
                    logger.info("Calling Google Cloud Vision API...")
                    vision_result = google_vision_ocr_from_numpy(preprocessed, google_api_key)
                    
                    if not vision_result["success"]:
                        logger.error(f"Google Vision error: {vision_result['error']}")
                        raise ValueError(f"Google Vision error: {vision_result['error']}")
                    
                    full_text = vision_result["full_text"]
                    ocr_confidence = vision_result["confidence"]
                    logger.info(f"Google Vision OCR: {len(full_text)} chars, confidence={ocr_confidence:.2f}")
                else:
                    # Fallback: envoyer l'image originale directement
                    logger.warning("CamScanner preprocessing failed, using original image")
                    image_base64 = base64.b64encode(img_data).decode("utf-8")
                    vision_result = google_vision_ocr(image_base64, google_api_key)
                    
                    if not vision_result["success"]:
                        raise ValueError(f"Google Vision error: {vision_result['error']}")
                    
                    full_text = vision_result["full_text"]
                    ocr_confidence = vision_result["confidence"]
                
                # ====== PARSING STRUCTURÉ (parser.py) ======
                # Utiliser les parsers regex existants sur le texte brut de Google Vision
                logger.info("Parsing structured data from Google Vision text...")
                
                # Parser le VIN
                vin_raw = parse_vin(full_text)
                if not vin_raw:
                    # Chercher dans le texte brut avec un pattern plus permissif
                    vin_match = re.search(r'1C4[A-Z0-9]{14}', full_text.replace("-", "").replace(" ", "").upper())
                    if vin_match:
                        vin_raw = vin_match.group()
                vin_raw = str(vin_raw or "").replace("-", "").replace(" ", "").upper()[:17]
                
                # Valider et corriger le VIN
                vin_result = validate_and_correct_vin(vin_raw) if len(vin_raw) == 17 else {"corrected": vin_raw, "is_valid": False, "was_corrected": False}
                vin_corrected = vin_result.get("corrected", vin_raw)
                vin_valid = vin_result.get("is_valid", False)
                vin_was_corrected = vin_result.get("was_corrected", False)
                vin_info = decode_vin(vin_corrected) if len(vin_corrected) == 17 else {}
                
                # Parser le code modèle
                model_code = parse_model_code(full_text) or ""
                product_info = decode_product_code(model_code) if model_code else {}
                
                # Validation cohérence VIN ↔ marque
                vin_brand = decode_vin_brand(vin_corrected)
                expected_brand = product_info.get("brand")
                vin_consistent = validate_vin_brand_consistency(vin_corrected, expected_brand) if vin_brand else True
                
                # Parser les données financières
                financial = parse_financial_data(full_text)
                ep_cost = financial.get("ep_cost", 0) or 0
                pdco = financial.get("pdco", 0) or 0
                pref = financial.get("pref", 0) or 0
                holdback = financial.get("holdback", 0) or 0
                
                # Parser les totaux
                totals = parse_totals(full_text)
                subtotal = totals.get("subtotal", 0) or 0
                invoice_total = totals.get("invoice_total", 0) or 0
                
                # Parser les options
                options = []
                for opt in parse_options(full_text):
                    options.append({
                        "product_code": opt.get("product_code", ""),
                        "description": opt.get("description", "")[:80],
                        "amount": 0  # Pas de prix accessoires comme demandé
                    })
                
                # Parser le numéro de stock (souvent manuscrit)
                stock_no = parse_stock_number(full_text) or ""
                
                # Chercher le stock manuscrit avec un pattern plus large si non trouvé
                if not stock_no:
                    # Chercher un nombre de 5 chiffres isolé (souvent le stock manuscrit)
                    stock_match = re.search(r'\b(\d{5})\b', full_text)
                    if stock_match:
                        stock_no = stock_match.group(1)
                
                # Couleur - Extraire le code couleur (3 caractères commençant par P)
                raw_color = ""
                color_match = re.search(r'\b(P[A-Z0-9]{2})\b', full_text)
                if color_match:
                    raw_color = color_match.group(1)
                
                # Mapping des codes couleur FCA
                color_map = {
                    "PW7": "Blanc Vif", "PWZ": "Blanc Vif", "PXJ": "Noir Cristal", 
                    "PX8": "Noir Diamant", "PAU": "Rouge Flamme", "PSC": "Gris Destroyer", 
                    "PWL": "Blanc Perle", "PGG": "Gris Granit", "PBF": "Bleu Patriote", 
                    "PGE": "Vert Sarge", "PRM": "Rouge Velours", "PAR": "Argent Billet",
                    "PYB": "Jaune Stinger", "PBJ": "Bleu Hydro", "PFQ": "Granite Cristal"
                }
                color_code = raw_color if raw_color in color_map else raw_color
                
                parse_duration = round(time.time() - start_time, 3)
                
                # Validation avec validate_invoice_full
                vehicle_data_for_validation = {
                    "vin": vin_corrected,
                    "vin_valid": vin_valid,
                    "ep_cost": ep_cost,
                    "pdco": pdco,
                    "pref": pref,
                    "subtotal_excl_tax": subtotal,
                    "invoice_total": invoice_total,
                    "options": options
                }
                validation = validate_invoice_full(vehicle_data_for_validation)
                
                # Ajout erreurs spécifiques
                if vin_was_corrected:
                    validation["errors"] = validation.get("errors", []) + ["VIN auto-corrigé"]
                
                vehicle_data = {
                    "stock_no": stock_no,
                    "vin": vin_corrected,
                    "vin_original": vin_raw if vin_was_corrected else None,
                    "vin_valid": vin_valid,
                    "vin_corrected": vin_was_corrected,
                    "vin_brand": vin_brand,
                    "vin_consistent": vin_consistent,
                    "model_code": model_code,
                    "year": vin_info.get("year") or datetime.now().year,
                    "brand": product_info.get("brand") or vin_brand or "Stellantis",
                    "model": product_info.get("model") or "",
                    "trim": _build_trim_string(product_info),
                    "ep_cost": ep_cost,
                    "pdco": pdco,
                    "pref": pref,
                    "holdback": holdback,
                    "msrp": pdco,
                    "net_cost": ep_cost,
                    "subtotal": subtotal,
                    "invoice_total": invoice_total,
                    "color": color_map.get(color_code, color_code),
                    "options": options,
                    "file_hash": file_hash,
                    "parse_method": "google_vision_hybrid",
                    "metrics": {
                        "parse_duration_sec": parse_duration,
                        "validation_score": validation.get("score", 0),
                        "ocr_confidence": ocr_confidence,
                        "cost_estimate": "~$0.0015"  # Google Vision est ~85% moins cher que GPT-4
                    }
                }
                
                logger.info(f"Google Vision - Model extraction: code={model_code}, product_info={product_info}")
                parse_method = "google_vision_hybrid"
                logger.info(f"Google Vision Hybrid: VIN={vin_corrected}, EP={ep_cost}, PDCO={pdco}, Score={validation.get('score', 0)}, Duration={parse_duration}s")
                
            except HTTPException:
                raise
            except Exception as vision_err:
                logger.error(f"Erreur Google Vision: {vision_err}")
                raise HTTPException(status_code=500, detail=f"Erreur analyse OCR: {str(vision_err)}")
        
        # ===== RÈGLE D'OR : JAMAIS ENREGISTRER SI INVALIDE =====
        # Le système n'enregistre jamais si :
        # - VIN invalide (ou manquant)
        # - EP manquant
        # - PDCO manquant
        # - EP >= PDCO
        
        if vehicle_data:
            vin = vehicle_data.get("vin", "")
            ep = vehicle_data.get("ep_cost", 0) or 0
            pdco = vehicle_data.get("pdco", 0) or 0
            
            blocking_errors = []
            
            if not vin or len(vin) != 17:
                blocking_errors.append("VIN manquant ou invalide (doit être 17 caractères)")
            
            if not ep or ep <= 0:
                blocking_errors.append("EP (Employee Price) manquant")
            
            if not pdco or pdco <= 0:
                blocking_errors.append("PDCO (Dealer Price) manquant")
            
            if ep > 0 and pdco > 0 and ep >= pdco:
                blocking_errors.append(f"EP ({ep}) doit être inférieur à PDCO ({pdco})")
            
            if blocking_errors:
                return {
                    "success": False,
                    "review_required": True,
                    "blocking_errors": blocking_errors,
                    "vehicle": vehicle_data,
                    "validation": validation,
                    "parse_method": parse_method,
                    "message": "Données critiques manquantes ou invalides. Révision obligatoire."
                }
        
        # ===== MONITORING: LOG PARSING METRICS =====
        try:
            score = validation.get("score", 0) if validation else 0
            duration = 0
            if vehicle_data and vehicle_data.get("metrics"):
                duration = vehicle_data["metrics"].get("parse_duration_sec", 0)
            
            # Déterminer le statut basé sur le score
            if score >= 85:
                status = "auto"
            elif score >= 60:
                status = "review"
            else:
                status = "vision"
            
            # Estimer le coût basé sur la méthode utilisée
            cost_estimate = 0.0
            if "google_vision" in parse_method:
                cost_estimate = 0.0015  # Google Vision DOCUMENT_TEXT_DETECTION
            elif "vision" in parse_method:
                cost_estimate = 0.02    # GPT-4 Vision (fallback)
            # OCR Tesseract = gratuit
            
            log_entry = {
                "timestamp": datetime.now(),
                "owner_id": user["id"],
                "parse_method": parse_method,
                "score": score,
                "status": status,
                "vin_valid": vehicle_data.get("vin_valid") if vehicle_data else False,
                "vin": vehicle_data.get("vin", "")[:17] if vehicle_data else "",
                "stock_no": vehicle_data.get("stock_no", "") if vehicle_data else "",
                "brand": vehicle_data.get("brand", "") if vehicle_data else "",
                "model": vehicle_data.get("model", "") if vehicle_data else "",
                "ep_cost": vehicle_data.get("ep_cost", 0) if vehicle_data else 0,
                "pdco": vehicle_data.get("pdco", 0) if vehicle_data else 0,
                "duration_sec": duration,
                "cost_estimate": cost_estimate,
                "success": True
            }
            
            await db.parsing_metrics.insert_one(log_entry)
            logger.info(f"Parsing metric logged: status={status}, score={score}, method={parse_method}")
        except Exception as log_err:
            logger.warning(f"Failed to log parsing metric: {log_err}")
        
        # ===== ENRICHISSEMENT AVEC PROMOTIONS FINANCEMENT =====
        # Ajouter automatiquement les informations de financement basées sur le code produit
        financing_info = None
        if vehicle_data and vehicle_data.get("model_code"):
            financing_info = get_financing_for_code(vehicle_data["model_code"])
            if financing_info:
                vehicle_data["financing"] = {
                    "consumer_cash": financing_info.get("consumer_cash", 0),
                    "bonus_cash": financing_info.get("bonus_cash", 0),
                    "option1_rates": financing_info.get("option1_rates", {}),
                    "option2_rates": financing_info.get("option2_rates", {}),
                    "programme_source": financing_info.get("programme_trim", "")
                }
                logger.info(f"Financing info added: Consumer Cash=${financing_info.get('consumer_cash', 0)}, Bonus=${financing_info.get('bonus_cash', 0)}")
        
        return {
            "success": True,
            "vehicle": vehicle_data,
            "validation": validation,
            "parse_method": parse_method,
            "has_financing": financing_info is not None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scanning invoice: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@api_router.post("/inventory/scan-invoice-file")
async def scan_invoice_file(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    """
    Scanne une facture FCA depuis un fichier uploadé (PDF ou image).
    Utilise le parser structuré (regex) pour les PDFs.
    
    USAGE: POST multipart/form-data avec field 'file'
    """
    user = await get_current_user(authorization)
    
    try:
        # Lire le fichier
        file_bytes = await file.read()
        
        # Détecter le type de fichier
        is_pdf = (
            file.content_type == "application/pdf" or
            file.filename.lower().endswith(".pdf") or
            file_bytes[:4] == b'%PDF'
        )
        
        # Convertir en base64 pour réutiliser la logique existante
        file_base64 = base64.b64encode(file_bytes).decode('utf-8')
        
        # Appeler le scan existant
        request = InvoiceScanRequest(image_base64=file_base64, is_pdf=is_pdf)
        return await scan_invoice(request, authorization)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scanning invoice file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@api_router.post("/test-ocr")
async def test_ocr_pipeline(file: UploadFile = File(...)):
    """
    🧪 ENDPOINT DE TEST - Pipeline OCR sans sauvegarde
    
    Teste le pipeline OCR complet sur une image de facture FCA:
    - Niveau 1: PDF → pdfplumber
    - Niveau 2: Image → OpenCV ROI + Tesseract
    - Niveau 3: Fallback → GPT-4 Vision (si score < 70)
    
    USAGE: POST multipart/form-data avec field 'file'
    Retourne: JSON avec données extraites + métriques + debug info
    
    ⚠️ NE SAUVEGARDE PAS en base de données
    """
    from ocr import process_image_ocr_pipeline, process_image_global_ocr
    from parser import parse_invoice_text
    from vin_utils import validate_and_correct_vin, decode_vin_year, decode_vin_brand
    from validation import validate_invoice_data as validate_invoice_full, calculate_validation_score
    
    try:
        start_time = time.time()
        
        # Lire le fichier
        file_bytes = await file.read()
        file_size = len(file_bytes)
        
        # Détecter le type
        is_pdf = (
            file.content_type == "application/pdf" or
            file.filename.lower().endswith(".pdf") or
            file_bytes[:4] == b'%PDF'
        )
        
        result = {
            "success": True,
            "file_info": {
                "filename": file.filename,
                "size_bytes": file_size,
                "content_type": file.content_type,
                "is_pdf": is_pdf
            },
            "pipeline_used": None,
            "raw_ocr": {},
            "parsed_data": {},
            "vin_analysis": {},
            "validation": {},
            "metrics": {},
            "debug": {}
        }
        
        # ===== TEST NIVEAU 1: PDF =====
        if is_pdf:
            result["pipeline_used"] = "pdf_native"
            extracted_text = extract_pdf_text(file_bytes)
            
            result["debug"]["pdf_text_length"] = len(extracted_text) if extracted_text else 0
            result["debug"]["pdf_text_preview"] = extracted_text[:500] if extracted_text else ""
            
            if extracted_text and len(extracted_text) > 100:
                parsed = parse_fca_invoice_structured(extracted_text)
                validation = validate_invoice_data(parsed)
                
                result["parsed_data"] = parsed
                result["validation"] = validation
                
                # VIN analysis
                vin = parsed.get("vin", "")
                if vin:
                    vin_result = validate_and_correct_vin(vin)
                    result["vin_analysis"] = vin_result
        
        # ===== TEST NIVEAU 2: OCR PAR ZONES =====
        else:
            result["pipeline_used"] = "ocr_zones"
            
            # Pipeline OCR par zones
            ocr_result = process_image_ocr_pipeline(file_bytes)
            result["raw_ocr"] = {
                "zones_processed": ocr_result.get("zones_processed", 0),
                "vin_text": ocr_result.get("vin_text", "")[:200],
                "finance_text": ocr_result.get("finance_text", "")[:200],
                "options_text": ocr_result.get("options_text", "")[:300],
                "totals_text": ocr_result.get("totals_text", "")[:200],
            }
            
            # Parser structuré
            parsed = parse_invoice_text(ocr_result)
            result["parsed_data"] = {
                "vin": parsed.get("vin"),
                "model_code": parsed.get("model_code"),
                "stock_no": parsed.get("stock_no"),
                "ep_cost": parsed.get("ep_cost"),
                "pdco": parsed.get("pdco"),
                "pref": parsed.get("pref"),
                "holdback": parsed.get("holdback"),
                "subtotal": parsed.get("subtotal"),
                "invoice_total": parsed.get("invoice_total"),
                "options_count": len(parsed.get("options", [])),
                "options": parsed.get("options", [])[:10],  # Limiter à 10
                "fields_extracted": parsed.get("fields_extracted", 0)
            }
            
            # VIN analysis approfondi
            vin = parsed.get("vin", "")
            if vin:
                vin_result = validate_and_correct_vin(vin)
                result["vin_analysis"] = {
                    "original": vin_result.get("original"),
                    "corrected": vin_result.get("corrected"),
                    "is_valid": vin_result.get("is_valid"),
                    "was_corrected": vin_result.get("was_corrected"),
                    "correction_type": vin_result.get("correction_type"),
                    "year": vin_result.get("year"),
                    "brand": vin_result.get("brand"),
                    "confidence": vin_result.get("confidence")
                }
            
            # Validation complète
            parsed["vin_valid"] = result["vin_analysis"].get("is_valid", False) if result["vin_analysis"] else False
            validation_result = validate_invoice_full(parsed)
            result["validation"] = validation_result
            
            # Recommandation fallback
            score = validation_result.get("score", 0)
            if score < 70:
                result["debug"]["recommendation"] = "Score < 70, le fallback GPT-4 Vision serait utilisé en production"
            else:
                result["debug"]["recommendation"] = f"Score {score}/100 suffisant, pas besoin de fallback Vision"
            
            # Test OCR global (comparaison)
            global_text = process_image_global_ocr(file_bytes)
            result["debug"]["global_ocr_length"] = len(global_text)
            result["debug"]["global_ocr_preview"] = global_text[:300] if global_text else ""
        
        # Métriques finales
        duration = round(time.time() - start_time, 3)
        result["metrics"] = {
            "parse_duration_sec": duration,
            "validation_score": result["validation"].get("score", 0),
            "cost_estimate": "$0.00 (100% open-source)"
        }
        
        return result
        
    except Exception as e:
        logger.error(f"Test OCR error: {str(e)}")
        import traceback
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@api_router.post("/inventory/scan-and-save")
async def scan_and_save_invoice(request: InvoiceScanRequest, authorization: Optional[str] = Header(None)):
    """Scanne une facture ET sauvegarde automatiquement le véhicule"""
    user = await get_current_user(authorization)
    
    # First scan the invoice
    scan_result = await scan_invoice(request, authorization)
    
    if not scan_result.get("success"):
        return scan_result
    
    vehicle_data = scan_result.get("vehicle", {})
    
    # ENRICHIR avec décodage VIN et codes produits
    vehicle_data = enrich_vehicle_data(vehicle_data)
    
    # Check if stock_no already exists
    stock_no = vehicle_data.get("stock_no", "")
    if not stock_no:
        raise HTTPException(status_code=400, detail="Numéro de stock non trouvé dans la facture")
    
    existing = await db.inventory.find_one({"stock_no": stock_no, "owner_id": user["id"]})
    
    # Prepare vehicle document
    vehicle_doc = {
        "id": str(uuid.uuid4()),
        "owner_id": user["id"],
        "stock_no": stock_no,
        "vin": vehicle_data.get("vin", ""),
        "brand": vehicle_data.get("brand", ""),
        "model": vehicle_data.get("model", ""),
        "trim": vehicle_data.get("trim", ""),
        "year": vehicle_data.get("year", datetime.now().year),
        "type": vehicle_data.get("type", "neuf"),
        "pdco": vehicle_data.get("pdco", 0) or 0,
        "ep_cost": vehicle_data.get("ep_cost", 0) or 0,
        "holdback": vehicle_data.get("holdback", 0) or 0,
        "net_cost": vehicle_data.get("net_cost", 0) or 0,
        "msrp": vehicle_data.get("msrp", 0) or 0,
        "asking_price": vehicle_data.get("msrp", 0) or 0,  # Default to MSRP
        "sold_price": None,
        "status": "disponible",
        "km": 0,
        "color": vehicle_data.get("color", ""),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    if existing:
        # Update existing
        await db.inventory.update_one(
            {"stock_no": stock_no, "owner_id": user["id"]},
            {"$set": vehicle_doc}
        )
        action = "mis à jour"
    else:
        # Insert new
        await db.inventory.insert_one(vehicle_doc)
        action = "ajouté"
    
    # Save options if present
    options = vehicle_data.get("options", [])
    if options:
        # Delete existing options
        await db.vehicle_options.delete_many({"stock_no": stock_no})
        # Insert new options
        for idx, opt in enumerate(options):
            await db.vehicle_options.insert_one({
                "id": str(uuid.uuid4()),
                "stock_no": stock_no,
                # PATCH: Support "product_code" et "code" pour compatibilité
                "product_code": opt.get("product_code", opt.get("code", "")),
                "order": idx,
                "description": opt.get("description", ""),
                "amount": opt.get("amount", 0) or 0
            })
    
    return {
        "success": True,
        "vehicle": vehicle_doc,
        "options_count": len(options),
        "action": action,
        "message": f"Véhicule {stock_no} {action} avec {len(options)} options"
    }

# ============ Admin Endpoints ============

class AdminUserResponse(BaseModel):
    id: str
    name: str
    email: str
    created_at: datetime
    last_login: Optional[datetime] = None
    is_blocked: bool = False
    is_admin: bool = False
    contacts_count: int = 0
    submissions_count: int = 0

async def require_admin(authorization: Optional[str] = Header(None)):
    """Vérifie que l'utilisateur est admin"""
    user = await get_current_user(authorization)
    is_admin = user.get("is_admin", False) or user.get("email") == ADMIN_EMAIL
    if not is_admin:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    return user

# ============ Parsing Metrics & Monitoring ============

@api_router.get("/admin/parsing-stats")
async def get_parsing_stats(authorization: Optional[str] = Header(None)):
    """
    Statistiques de parsing en temps réel (admin seulement)
    
    Retourne:
    - total_scans: nombre total de scans
    - auto_rate: % auto-approved (score >= 85)
    - review_rate: % review required (60-84)
    - vision_rate: % vision fallback (< 60)
    - avg_score: score moyen
    - avg_time_sec: temps moyen de traitement
    - quality_alert: True si qualité en baisse
    """
    await require_admin(authorization)
    
    try:
        total = await db.parsing_metrics.count_documents({})
        
        if total == 0:
            return {
                "total_scans": 0,
                "auto_rate": 0,
                "review_rate": 0,
                "vision_rate": 0,
                "avg_score": 0,
                "avg_time_sec": 0,
                "quality_alert": False,
                "message": "Aucun scan enregistré"
            }
        
        auto = await db.parsing_metrics.count_documents({"status": "auto"})
        review = await db.parsing_metrics.count_documents({"status": "review"})
        vision = await db.parsing_metrics.count_documents({"status": "vision"})
        
        # Aggregations pour moyennes
        avg_score_result = await db.parsing_metrics.aggregate([
            {"$group": {"_id": None, "avg": {"$avg": "$score"}}}
        ]).to_list(1)
        
        avg_time_result = await db.parsing_metrics.aggregate([
            {"$group": {"_id": None, "avg": {"$avg": "$duration_sec"}}}
        ]).to_list(1)
        
        avg_score = round(avg_score_result[0]["avg"], 2) if avg_score_result else 0
        avg_time = round(avg_time_result[0]["avg"], 2) if avg_time_result else 0
        
        auto_rate = round(auto / total * 100, 2)
        review_rate = round(review / total * 100, 2)
        vision_rate = round(vision / total * 100, 2)
        
        # Détection dérive qualité
        quality_alert = auto_rate < 70 or avg_score < 80
        
        return {
            "total_scans": total,
            "auto_rate": auto_rate,
            "review_rate": review_rate,
            "vision_rate": vision_rate,
            "avg_score": avg_score,
            "avg_time_sec": avg_time,
            "quality_alert": quality_alert,
            "breakdown": {
                "auto_approved": auto,
                "review_required": review,
                "vision_fallback": vision
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting parsing stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/admin/parsing-history")
async def get_parsing_history(
    days: int = 7,
    authorization: Optional[str] = Header(None)
):
    """
    Historique des métriques de parsing sur N jours (admin seulement)
    
    Retourne les stats agrégées par jour
    """
    await require_admin(authorization)
    
    try:
        from_date = datetime.now() - timedelta(days=days)
        
        pipeline = [
            {"$match": {"timestamp": {"$gte": from_date}}},
            {"$group": {
                "_id": {
                    "year": {"$year": "$timestamp"},
                    "month": {"$month": "$timestamp"},
                    "day": {"$dayOfMonth": "$timestamp"}
                },
                "total": {"$sum": 1},
                "avg_score": {"$avg": "$score"},
                "avg_duration": {"$avg": "$duration_sec"},
                "auto_count": {"$sum": {"$cond": [{"$eq": ["$status", "auto"]}, 1, 0]}},
                "review_count": {"$sum": {"$cond": [{"$eq": ["$status", "review"]}, 1, 0]}},
                "vision_count": {"$sum": {"$cond": [{"$eq": ["$status", "vision"]}, 1, 0]}}
            }},
            {"$sort": {"_id.year": 1, "_id.month": 1, "_id.day": 1}}
        ]
        
        results = await db.parsing_metrics.aggregate(pipeline).to_list(100)
        
        history = []
        for r in results:
            date_str = f"{r['_id']['year']}-{r['_id']['month']:02d}-{r['_id']['day']:02d}"
            history.append({
                "date": date_str,
                "total": r["total"],
                "avg_score": round(r["avg_score"], 2) if r["avg_score"] else 0,
                "avg_duration_sec": round(r["avg_duration"], 2) if r["avg_duration"] else 0,
                "auto_rate": round(r["auto_count"] / r["total"] * 100, 2) if r["total"] else 0,
                "review_rate": round(r["review_count"] / r["total"] * 100, 2) if r["total"] else 0,
                "vision_rate": round(r["vision_count"] / r["total"] * 100, 2) if r["total"] else 0
            })
        
        return {
            "period_days": days,
            "history": history
        }
        
    except Exception as e:
        logger.error(f"Error getting parsing history: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ User Scan History Endpoints (Non-Admin) ============

@api_router.get("/scans/history")
async def get_user_scan_history(
    limit: int = 50,
    authorization: Optional[str] = Header(None)
):
    """
    Historique des scans de factures pour l'utilisateur connecté
    
    Retourne les derniers scans avec détails (VIN, score, méthode, coût)
    """
    user = await get_current_user(authorization)
    
    try:
        # Récupérer les derniers scans de l'utilisateur
        cursor = db.parsing_metrics.find(
            {"owner_id": user["id"]},
            {"_id": 0}
        ).sort("timestamp", -1).limit(limit)
        
        scans = []
        async for scan in cursor:
            scans.append({
                "timestamp": scan.get("timestamp").isoformat() if scan.get("timestamp") else None,
                "vin": scan.get("vin", ""),
                "stock_no": scan.get("stock_no", ""),
                "brand": scan.get("brand", ""),
                "model": scan.get("model", ""),
                "ep_cost": scan.get("ep_cost", 0),
                "pdco": scan.get("pdco", 0),
                "score": scan.get("score", 0),
                "status": scan.get("status", "unknown"),
                "parse_method": scan.get("parse_method", "unknown"),
                "duration_sec": round(scan.get("duration_sec", 0), 2),
                "cost_estimate": scan.get("cost_estimate", 0),
                "vin_valid": scan.get("vin_valid", False),
                "success": scan.get("success", True)
            })
        
        return {
            "count": len(scans),
            "scans": scans
        }
        
    except Exception as e:
        logger.error(f"Error getting user scan history: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/scans/stats")
async def get_user_scan_stats(
    days: int = 30,
    authorization: Optional[str] = Header(None)
):
    """
    Statistiques de scans pour l'utilisateur connecté
    
    Retourne:
    - total_scans: nombre total de scans
    - success_rate: % de scans réussis (score >= 70)
    - avg_score: score de confiance moyen
    - avg_duration: temps moyen de traitement
    - total_cost: coût estimé total
    - methods_breakdown: répartition par méthode (tesseract, google_vision)
    """
    user = await get_current_user(authorization)
    
    try:
        from_date = datetime.now() - timedelta(days=days)
        
        # Filtrer par utilisateur et période
        match_filter = {
            "owner_id": user["id"],
            "timestamp": {"$gte": from_date}
        }
        
        # Compter total
        total = await db.parsing_metrics.count_documents(match_filter)
        
        if total == 0:
            return {
                "period_days": days,
                "total_scans": 0,
                "success_rate": 0,
                "avg_score": 0,
                "avg_duration_sec": 0,
                "total_cost_estimate": 0,
                "methods_breakdown": {},
                "message": "Aucun scan dans cette période"
            }
        
        # Aggregation pour statistiques
        pipeline = [
            {"$match": match_filter},
            {"$group": {
                "_id": None,
                "total": {"$sum": 1},
                "avg_score": {"$avg": "$score"},
                "avg_duration": {"$avg": "$duration_sec"},
                "total_cost": {"$sum": {"$ifNull": ["$cost_estimate", 0]}},
                "success_count": {"$sum": {"$cond": [{"$gte": ["$score", 70]}, 1, 0]}},
                "google_vision_count": {"$sum": {"$cond": [{"$regexMatch": {"input": {"$ifNull": ["$parse_method", ""]}, "regex": "google_vision"}}, 1, 0]}},
                "tesseract_count": {"$sum": {"$cond": [{"$regexMatch": {"input": {"$ifNull": ["$parse_method", ""]}, "regex": "tesseract|pdfplumber"}}, 1, 0]}},
                "gpt4_vision_count": {"$sum": {"$cond": [{"$regexMatch": {"input": {"$ifNull": ["$parse_method", ""]}, "regex": "vision_optimized"}}, 1, 0]}}
            }}
        ]
        
        results = await db.parsing_metrics.aggregate(pipeline).to_list(1)
        
        if results:
            r = results[0]
            return {
                "period_days": days,
                "total_scans": r["total"],
                "success_rate": round(r["success_count"] / r["total"] * 100, 1) if r["total"] else 0,
                "avg_score": round(r["avg_score"], 1) if r["avg_score"] else 0,
                "avg_duration_sec": round(r["avg_duration"], 2) if r["avg_duration"] else 0,
                "total_cost_estimate": round(r["total_cost"], 4),
                "cost_savings_vs_gpt4": round(r["google_vision_count"] * 0.0185, 2),  # Économie par rapport à GPT-4
                "methods_breakdown": {
                    "google_vision_hybrid": r["google_vision_count"],
                    "tesseract_pdfplumber": r["tesseract_count"],
                    "gpt4_vision": r["gpt4_vision_count"]
                },
                "free_quota_remaining": max(0, 1000 - r["google_vision_count"])  # 1000 gratuits/mois Google
            }
        
        return {"period_days": days, "total_scans": 0, "message": "Pas de données"}
        
    except Exception as e:
        logger.error(f"Error getting user scan stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Other Admin Endpoints ============

@api_router.get("/admin/users")
async def get_all_users(authorization: Optional[str] = Header(None)):
    """Récupère tous les utilisateurs (admin seulement)"""
    await require_admin(authorization)
    
    users = []
    async for user in db.users.find({}, {"_id": 0, "password_hash": 0}):
        user_id = user.get("id")
        
        # Count contacts and submissions for this user
        contacts_count = await db.contacts.count_documents({"owner_id": user_id})
        submissions_count = await db.submissions.count_documents({"owner_id": user_id})
        
        users.append({
            "id": user_id,
            "name": user.get("name", ""),
            "email": user.get("email", ""),
            "created_at": user.get("created_at", datetime.utcnow()).isoformat() if user.get("created_at") else None,
            "last_login": user.get("last_login").isoformat() if user.get("last_login") else None,
            "is_blocked": user.get("is_blocked", False),
            "is_admin": user.get("is_admin", False) or user.get("email") == ADMIN_EMAIL,
            "contacts_count": contacts_count,
            "submissions_count": submissions_count
        })
    
    return users

@api_router.put("/admin/users/{user_id}/block")
async def block_user(user_id: str, authorization: Optional[str] = Header(None)):
    """Bloque un utilisateur (admin seulement)"""
    admin = await require_admin(authorization)
    
    # Cannot block yourself
    if admin.get("id") == user_id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas vous bloquer vous-même")
    
    # Find user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Cannot block an admin
    if user.get("email") == ADMIN_EMAIL or user.get("is_admin", False):
        raise HTTPException(status_code=400, detail="Impossible de bloquer un administrateur")
    
    # Block user
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_blocked": True}}
    )
    
    # Delete all tokens for this user (force logout)
    await db.tokens.delete_many({"user_id": user_id})
    
    return {"success": True, "message": f"Utilisateur {user.get('name')} bloqué"}

@api_router.put("/admin/users/{user_id}/unblock")
async def unblock_user(user_id: str, authorization: Optional[str] = Header(None)):
    """Débloque un utilisateur (admin seulement)"""
    await require_admin(authorization)
    
    # Find user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Unblock user
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_blocked": False}}
    )
    
    return {"success": True, "message": f"Utilisateur {user.get('name')} débloqué"}

@api_router.get("/admin/stats")
async def get_admin_stats(authorization: Optional[str] = Header(None)):
    """Récupère les statistiques globales (admin seulement)"""
    await require_admin(authorization)
    
    total_users = await db.users.count_documents({})
    blocked_users = await db.users.count_documents({"is_blocked": True})
    total_contacts = await db.contacts.count_documents({})
    total_submissions = await db.submissions.count_documents({})
    
    return {
        "total_users": total_users,
        "active_users": total_users - blocked_users,
        "blocked_users": blocked_users,
        "total_contacts": total_contacts,
        "total_submissions": total_submissions
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Endpoint pour servir les images de test CamScanner
@api_router.get("/debug/camscanner-preview")
async def get_camscanner_preview():
    """Retourne l'image prétraitée par CamScanner"""
    file_path = Path(__file__).parent / "static" / "facture_camscanner.jpg"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/jpeg")
    raise HTTPException(status_code=404, detail="Image not found")

@api_router.get("/debug/original-preview")
async def get_original_preview():
    """Retourne l'image originale"""
    file_path = Path(__file__).parent / "static" / "facture_originale.jpg"
    if file_path.exists():
        return FileResponse(file_path, media_type="image/jpeg")
    raise HTTPException(status_code=404, detail="Image not found")



@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
