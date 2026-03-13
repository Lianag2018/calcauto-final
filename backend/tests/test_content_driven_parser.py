"""
Test suite for content-driven table classification parser rewrite.
Tests: PDF extraction, SCI Lease parsing, Excel generation, validation endpoint.
Focus: Verifying content-based table identification instead of hardcoded indices.
"""
import pytest
import requests
import os
import json
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://lease-extraction.preview.emergentagent.com').rstrip('/')
PDF_PATH = "/app/backend/data/march2026_source.pdf"
PDF_PASSWORD = "Liana2018"
PROGRAM_MONTH = 3
PROGRAM_YEAR = 2026


class TestPDFExtraction:
    """Test POST /api/extract-pdf returns correct program counts and data."""

    def test_extract_pdf_returns_93_programs(self):
        """Verify extraction returns exactly 93 programs for March 2026."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        result = response.json()
        assert result.get('success') is True, f"Extraction failed: {result.get('message')}"
        
        programs = result.get('programs', [])
        assert len(programs) == 93, f"Expected 93 programs, got {len(programs)}"
        print(f"✓ Extracted {len(programs)} programs")

    def test_extract_pdf_returns_73_sci_lease(self):
        """Verify SCI Lease extraction returns 73 vehicles total."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        sci_count = result.get('sci_lease_count', 0)
        assert sci_count == 73, f"Expected 73 SCI Lease vehicles, got {sci_count}"
        print(f"✓ SCI Lease count: {sci_count}")

    def test_programs_by_year_breakdown(self):
        """Verify year distribution: 36 for 2026, 43 for 2025, 14 for 2024."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        programs = result.get('programs', [])
        
        by_year = {}
        for p in programs:
            y = p.get('year', 0)
            by_year[y] = by_year.get(y, 0) + 1
        
        assert by_year.get(2026, 0) == 36, f"Expected 36 for 2026, got {by_year.get(2026, 0)}"
        assert by_year.get(2025, 0) == 43, f"Expected 43 for 2025, got {by_year.get(2025, 0)}"
        assert by_year.get(2024, 0) == 14, f"Expected 14 for 2024, got {by_year.get(2024, 0)}"
        print(f"✓ Year breakdown: 2026={by_year.get(2026)}, 2025={by_year.get(2025)}, 2024={by_year.get(2024)}")

    def test_brand_distribution(self):
        """Verify brand counts: Chrysler=8, Jeep=39, Dodge=24, Ram=20, Fiat=2."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        programs = result.get('programs', [])
        
        by_brand = {}
        for p in programs:
            b = p.get('brand', 'Unknown')
            by_brand[b] = by_brand.get(b, 0) + 1
        
        assert by_brand.get('Chrysler', 0) == 8, f"Chrysler: expected 8, got {by_brand.get('Chrysler', 0)}"
        assert by_brand.get('Jeep', 0) == 39, f"Jeep: expected 39, got {by_brand.get('Jeep', 0)}"
        assert by_brand.get('Dodge', 0) == 24, f"Dodge: expected 24, got {by_brand.get('Dodge', 0)}"
        assert by_brand.get('Ram', 0) == 20, f"Ram: expected 20, got {by_brand.get('Ram', 0)}"
        assert by_brand.get('Fiat', 0) == 2, f"Fiat: expected 2, got {by_brand.get('Fiat', 0)}"
        print(f"✓ Brand distribution: {by_brand}")


class TestMSRPDiscountBugFix:
    """Test that MSRP discount lines are correctly ignored (not treated as consumer_cash)."""

    def test_wagoneer_s_2024_consumer_cash_zero(self):
        """Wagoneer S 2024 consumer_cash must be $0 (MSRP discount fix)."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        programs = result.get('programs', [])
        
        wagoneer_s = [p for p in programs if 'wagoneer s' in p.get('model', '').lower() and p.get('year') == 2024]
        assert len(wagoneer_s) > 0, "Wagoneer S 2024 not found in programs"
        
        for ws in wagoneer_s:
            cc = ws.get('consumer_cash', -1)
            assert cc == 0, f"Wagoneer S 2024 consumer_cash should be $0 (MSRP discount), got ${cc}"
        print(f"✓ Wagoneer S 2024 consumer_cash = $0 (MSRP discount correctly ignored)")

    def test_charger_daytona_2024_consumer_cash_zero(self):
        """Charger Daytona 2024 models consumer_cash must be $0."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        programs = result.get('programs', [])
        
        daytona = [p for p in programs if 'daytona' in p.get('model', '').lower() or 'daytona' in p.get('trim', '').lower()]
        daytona_2024 = [p for p in daytona if p.get('year') == 2024]
        
        assert len(daytona_2024) > 0, "Charger Daytona 2024 not found"
        
        for d in daytona_2024:
            cc = d.get('consumer_cash', -1)
            assert cc == 0, f"Charger Daytona 2024 ({d.get('trim')}) consumer_cash should be $0, got ${cc}"
        print(f"✓ Charger Daytona 2024 models consumer_cash = $0 (verified {len(daytona_2024)} models)")


class TestSCILeaseFixes:
    """Test SCI Lease parsing with content-driven column detection."""

    def test_power_wagon_2026_lease_cash_6000(self):
        """SCI Lease Power Wagon 2026 lease_cash must be $6000 (not $8)."""
        # Load from saved SCI data file
        sci_path = f"/app/backend/data/sci_lease_rates_mar{PROGRAM_YEAR}.json"
        assert os.path.exists(sci_path), f"SCI file not found: {sci_path}"
        
        with open(sci_path, 'r') as f:
            sci_data = json.load(f)
        
        v2026 = sci_data.get('vehicles_2026', [])
        # Find the actual Power Wagon entry (not entries that just mention it in exclusion text)
        power_wagon = [v for v in v2026 if v.get('model', '').lower().startswith('ram 2500 power wagon')]
        
        assert len(power_wagon) > 0, "Power Wagon 2026 not found in SCI Lease data"
        
        for pw in power_wagon:
            lc = pw.get('lease_cash', -1)
            assert lc == 6000, f"Power Wagon lease_cash should be $6,000, got ${lc}"
        print(f"✓ Power Wagon 2026 lease_cash = $6,000")

    def test_sci_lease_vehicle_counts_by_year(self):
        """SCI Lease: 33 vehicles for 2026, 40 for 2025."""
        sci_path = f"/app/backend/data/sci_lease_rates_mar{PROGRAM_YEAR}.json"
        assert os.path.exists(sci_path), f"SCI file not found: {sci_path}"
        
        with open(sci_path, 'r') as f:
            sci_data = json.load(f)
        
        v2026 = sci_data.get('vehicles_2026', [])
        v2025 = sci_data.get('vehicles_2025', [])
        
        assert len(v2026) == 33, f"Expected 33 vehicles for 2026, got {len(v2026)}"
        assert len(v2025) == 40, f"Expected 40 vehicles for 2025, got {len(v2025)}"
        print(f"✓ SCI Lease counts: 2026={len(v2026)}, 2025={len(v2025)}")

    def test_sci_lease_model_names_uppercase(self):
        """SCI Lease data: all model names should start with uppercase letter."""
        sci_path = f"/app/backend/data/sci_lease_rates_mar{PROGRAM_YEAR}.json"
        assert os.path.exists(sci_path), f"SCI file not found: {sci_path}"
        
        with open(sci_path, 'r') as f:
            sci_data = json.load(f)
        
        all_vehicles = sci_data.get('vehicles_2026', []) + sci_data.get('vehicles_2025', [])
        
        lowercase_models = []
        for v in all_vehicles:
            model = v.get('model', '')
            if model and model[0].islower():
                lowercase_models.append(model)
        
        assert len(lowercase_models) == 0, f"Models starting with lowercase: {lowercase_models}"
        print(f"✓ All {len(all_vehicles)} SCI Lease models start with uppercase")


class TestBonusCash:
    """Test bonus cash extraction."""

    def test_fiat_500e_2025_bonus_cash_5000(self):
        """Fiat 500e 2025 bonus_cash must be $5000."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        programs = result.get('programs', [])
        
        fiat_500e = [p for p in programs if '500e' in p.get('model', '').lower() and p.get('year') == 2025]
        assert len(fiat_500e) > 0, "Fiat 500e 2025 not found"
        
        for f500 in fiat_500e:
            bc = f500.get('bonus_cash', 0)
            assert bc == 5000, f"Fiat 500e 2025 bonus_cash should be $5,000, got ${bc}"
        print(f"✓ Fiat 500e 2025 bonus_cash = $5,000")


class TestLoyaltyMarkers:
    """Test loyalty P marker extraction."""

    def test_loyalty_markers_count(self):
        """17 vehicles should have loyalty markers."""
        with open(PDF_PATH, 'rb') as f:
            files = {'file': ('march2026_source.pdf', f, 'application/pdf')}
            data = {
                'password': PDF_PASSWORD,
                'program_month': PROGRAM_MONTH,
                'program_year': PROGRAM_YEAR
            }
            response = requests.post(f"{BASE_URL}/api/extract-pdf", files=files, data=data, timeout=120)
        
        assert response.status_code == 200
        result = response.json()
        programs = result.get('programs', [])
        
        loyalty_count = sum(1 for p in programs if p.get('loyalty_cash') or p.get('loyalty_opt1') or p.get('loyalty_opt2'))
        assert loyalty_count == 17, f"Expected 17 loyalty markers, got {loyalty_count}"
        print(f"✓ Loyalty marker count = {loyalty_count}")


class TestValidationEndpoint:
    """Test GET /api/validate-data endpoint."""

    def test_validate_data_returns_success(self):
        """GET /api/validate-data?month=3&year=2026 should return success=true with 0 errors."""
        response = requests.get(f"{BASE_URL}/api/validate-data?month={PROGRAM_MONTH}&year={PROGRAM_YEAR}", timeout=30)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        result = response.json()
        
        assert result.get('success') is True, f"Validation failed: {result}"
        errors = result.get('errors', [])
        assert len(errors) == 0, f"Expected 0 errors, got {len(errors)}: {errors}"
        print(f"✓ Validation passed: success={result.get('success')}, errors={len(errors)}, warnings={len(result.get('warnings', []))}")


class TestExcelDownload:
    """Test Excel generation and download."""

    def test_download_excel_returns_valid_xlsx(self):
        """GET /api/download-excel?month=3&year=2026 should return valid .xlsx."""
        response = requests.get(f"{BASE_URL}/api/download-excel?month={PROGRAM_MONTH}&year={PROGRAM_YEAR}", timeout=60)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert 'spreadsheet' in response.headers.get('Content-Type', '').lower() or 'octet-stream' in response.headers.get('Content-Type', '').lower()
        
        content = response.content
        assert len(content) > 1000, f"Excel file too small: {len(content)} bytes"
        
        # Verify it's a valid xlsx (starts with PK for zip)
        assert content[:2] == b'PK', "Not a valid xlsx file (should start with PK)"
        print(f"✓ Excel downloaded: {len(content)} bytes")

    def test_excel_has_3_sheets(self):
        """Excel should have 3 sheets: Financement, SCI Lease, Rapport."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/download-excel?month={PROGRAM_MONTH}&year={PROGRAM_YEAR}", timeout=60)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'Financement' in sheet_names, f"Missing 'Financement' sheet: {sheet_names}"
        assert 'SCI Lease' in sheet_names, f"Missing 'SCI Lease' sheet: {sheet_names}"
        assert 'Rapport' in sheet_names, f"Missing 'Rapport' sheet: {sheet_names}"
        print(f"✓ Excel has 3 sheets: {sheet_names}")

    def test_excel_sci_lease_column_width(self):
        """Excel SCI Lease model column should be 60 chars wide and left-aligned."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/download-excel?month={PROGRAM_MONTH}&year={PROGRAM_YEAR}", timeout=60)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['SCI Lease']
        
        # Column B is Modèle
        col_b_width = ws.column_dimensions['B'].width
        assert col_b_width >= 55, f"SCI Lease model column (B) width should be >= 55, got {col_b_width}"
        
        # Check first data row alignment (row 5 should be first data after headers)
        cell_b5 = ws['B5']
        if cell_b5.value:
            alignment = cell_b5.alignment
            assert alignment is None or alignment.horizontal in ('left', None, 'general'), \
                f"Model column should be left-aligned, got {alignment.horizontal if alignment else 'None'}"
        print(f"✓ SCI Lease model column width = {col_b_width}, alignment = left")

    def test_excel_financement_trim_column_width(self):
        """Excel Financement trim column should be 55 chars wide and left-aligned."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/download-excel?month={PROGRAM_MONTH}&year={PROGRAM_YEAR}", timeout=60)
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb['Financement']
        
        # Column D is Version/Trim
        col_d_width = ws.column_dimensions['D'].width
        assert col_d_width >= 50, f"Financement trim column (D) width should be >= 50, got {col_d_width}"
        
        # Check alignment on a data row
        cell_d4 = ws['D4']
        if cell_d4.value:
            alignment = cell_d4.alignment
            assert alignment is None or alignment.horizontal in ('left', None, 'general'), \
                f"Trim column should be left-aligned, got {alignment.horizontal if alignment else 'None'}"
        print(f"✓ Financement trim column width = {col_d_width}, alignment = left")


class TestContentDrivenClassification:
    """Test the new content-driven table classification logic."""

    def test_parser_classify_table_function(self):
        """Test _classify_table returns correct types based on content."""
        import sys
        sys.path.insert(0, '/app/backend/services')
        from pdfplumber_parser import _classify_table
        
        # Test rates table (many % values, 15+ cols)
        rates_table = [
            ['Vehicle', 'CC', 'P', '36M', '48M', '60M', '72M', '84M', '96M', 'ACC', 'P', '36M', '48M', '60M', '72M', '84M', '96M', 'Bonus'],
            ['Test Car', '$1000', 'P', '4.99%', '4.99%', '4.99%', '4.99%', '4.99%', '4.99%', '$0', '', '0%', '0%', '1.99%', '2.99%', '3.99%', '4.99%', '$500'],
        ]
        assert _classify_table(rates_table) == 'rates', f"Should classify as rates"
        
        # Test names table (brand codes)
        names_table = [
            ['RELSYRHC', 'Grand Caravan SXT'],
            ['', 'Pacifica'],
            ['PEEJ', 'Compass Sport'],
        ]
        assert _classify_table(names_table) == 'names', f"Should classify as names (brand codes)"
        
        print("✓ _classify_table correctly identifies rates and names tables")

    def test_parser_detect_rate_columns(self):
        """Test _detect_rate_columns finds correct column positions."""
        import sys
        sys.path.insert(0, '/app/backend/services')
        from pdfplumber_parser import _detect_rate_columns
        
        # Simulated table with rate headers
        table = [
            ['', '', '', '', '', '36M', '48M', '60M', '72M', '84M', '96M', '', '', '', '', '', '36M', '48M', '60M', '72M', '84M', '96M'],
            ['Grand Caravan', '$1000', 'P', '', '', '4.99%', '4.99%', '4.99%', '4.99%', '4.99%', '4.99%', '', '$0', '', '', '', '0%', '0%', '1.99%', '2.99%', '3.99%', '4.99%'],
        ]
        
        opt1, opt2 = _detect_rate_columns(table)
        assert opt1 == 5, f"opt1_start should be 5, got {opt1}"
        assert opt2 == 16, f"opt2_start should be 16, got {opt2}"
        
        print(f"✓ _detect_rate_columns: opt1={opt1}, opt2={opt2}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
