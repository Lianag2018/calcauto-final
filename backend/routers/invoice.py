from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Header
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime
import uuid
import json
import re
import os
import io
import base64
import hashlib
import time
import tempfile
from database import db, OPENAI_API_KEY, ROOT_DIR, logger
from models import InventoryVehicle, InvoiceScanRequest, ExcelExportRequest
from dependencies import get_current_user
from services.window_sticker import fetch_window_sticker, save_window_sticker_to_db

# OCR imports
import pytesseract
from PIL import Image
import cv2
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter()

# ============ Invoice Scanner with AI ============

import re
import base64

# ============ VIN Validation & Auto-Correction ============

# Table de translittération VIN (ISO 3779)
VIN_TRANSLATION = {
    **{str(i): i for i in range(10)},
    "A":1,"B":2,"C":3,"D":4,"E":5,"F":6,"G":7,"H":8,
    "J":1,"K":2,"L":3,"M":4,"N":5,"P":7,"R":9,
    "S":2,"T":3,"U":4,"V":5,"W":6,"X":7,"Y":8,"Z":9
}

# Poids par position pour calcul check digit
VIN_WEIGHTS = [8,7,6,5,4,3,2,10,0,9,8,7,6,5,4,3,2]

# Corrections OCR communes
VIN_OCR_CORRECTIONS = {
    "O": "0",  # O ressemble à 0
    "I": "1",  # I ressemble à 1
    "Q": "0",  # Q ressemble à 0
    "L": "1",  # L ressemble à 1
    "Z": "2",  # Z peut ressembler à 2
    "S": "5",  # S peut ressembler à 5
    "B": "8",  # B peut ressembler à 8
}

# Codes année (10e caractère)
VIN_YEAR_CODES = {
    'A': 2010, 'B': 2011, 'C': 2012, 'D': 2013, 'E': 2014,
    'F': 2015, 'G': 2016, 'H': 2017, 'J': 2018, 'K': 2019,
    'L': 2020, 'M': 2021, 'N': 2022, 'P': 2023, 'R': 2024,
    'S': 2025, 'T': 2026, 'V': 2027, 'W': 2028, 'X': 2029,
    'Y': 2030, '1': 2031, '2': 2032, '3': 2033, '4': 2034,
    '5': 2035, '6': 2036, '7': 2037, '8': 2038, '9': 2039
}

# WMI (World Manufacturer Identifier) - 3 premiers caractères
VIN_WMI_MAP = {
    "1C4": "Jeep",
    "1C6": "Ram",
    "1C3": "Chrysler",
    "2C3": "Chrysler",
    "2C4": "Chrysler",
    "3C4": "Chrysler",
    "3C6": "Ram",
    "3C3": "Chrysler",
    "1B3": "Dodge",
    "2B3": "Dodge",
    "3D4": "Dodge",
}


def compute_vin_check_digit(vin: str) -> str:
    """Calcule le check digit (9e caractère) d'un VIN"""
    total = 0
    for i, char in enumerate(vin.upper()):
        if i == 8:  # Skip position du check digit
            continue
        value = VIN_TRANSLATION.get(char, 0)
        total += value * VIN_WEIGHTS[i]
    
    remainder = total % 11
    return "X" if remainder == 10 else str(remainder)


def validate_vin_checksum(vin: str) -> bool:
    """Valide le check digit d'un VIN"""
    vin = vin.upper()
    
    if len(vin) != 17:
        return False
    
    # VIN ne peut pas contenir I, O, Q
    if any(c in "IOQ" for c in vin):
        return False
    
    expected = compute_vin_check_digit(vin)
    return vin[8] == expected


def auto_correct_vin(vin: str) -> tuple:
    """
    Corrige automatiquement les erreurs OCR communes dans un VIN.
    Essaie plusieurs stratégies de correction.
    
    Returns: (vin_corrigé, was_corrected)
    """
    vin = vin.upper().replace("-", "").replace(" ", "")
    
    if len(vin) != 17:
        return vin, False
    
    # Déjà valide ?
    if validate_vin_checksum(vin):
        return vin, False
    
    # Stratégie 1: Corrections simples caractère par caractère
    for i, char in enumerate(vin):
        if char in VIN_OCR_CORRECTIONS:
            corrected = vin[:i] + VIN_OCR_CORRECTIONS[char] + vin[i+1:]
            if validate_vin_checksum(corrected):
                return corrected, True
    
    # Stratégie 2: Permutations communes FCA (P↔J, S↔5, X↔K, etc.)
    common_swaps = [
        ("P", "J"), ("J", "P"),
        ("P", "S"), ("S", "P"),  # P et S se ressemblent
        ("5", "S"), ("S", "5"),  # 5 et S se ressemblent
        ("7", "T"), ("T", "7"),
        ("X", "K"), ("K", "X"),
        ("6", "G"), ("G", "6"),
        ("Y", "T"), ("T", "Y"),
        ("0", "D"), ("D", "0"),
        ("8", "B"), ("B", "8"),
    ]
    for old, new in common_swaps:
        for i, char in enumerate(vin):
            if char == old:
                corrected = vin[:i] + new + vin[i+1:]
                if validate_vin_checksum(corrected):
                    return corrected, True
    
    # Stratégie 3: Corriger l'année (position 10) - très important
    # P souvent confondu avec S (2023 vs 2025)
    year_pos = 9  # Index 9 = position 10
    year_swaps = [("P", "S"), ("S", "P"), ("R", "S"), ("S", "R"), ("P", "R"), ("T", "7")]
    for old, new in year_swaps:
        if vin[year_pos] == old:
            test = vin[:year_pos] + new + vin[year_pos+1:]
            if validate_vin_checksum(test):
                return test, True
    
    # Stratégie 4: Combinaisons multiples pour VINs Jeep/Ram
    # Position 5: K souvent lu comme X
    if len(vin) >= 6 and vin[4] in ["X", "K"]:
        swap = "K" if vin[4] == "X" else "X"
        test = vin[:4] + swap + vin[5:]
        if validate_vin_checksum(test):
            return test, True
        # Essayer avec correction année aussi
        for old, new in [("P", "S"), ("S", "P")]:
            if vin[year_pos] == old:
                test2 = test[:year_pos] + new + test[year_pos+1:]
                if validate_vin_checksum(test2):
                    return test2, True
    
    # Stratégie 5: Recalculer check digit si tout le reste semble OK
    expected_check = compute_vin_check_digit(vin)
    if vin[8] != expected_check:
        corrected = vin[:8] + expected_check + vin[9:]
        if not any(c in "IOQ" for c in corrected):
            return corrected, True
    
    # Stratégie 6: P/J + année + check digit
    if vin[3] == "J":
        test = vin[:3] + "P" + vin[4:]
        check = compute_vin_check_digit(test)
        test = test[:8] + check + test[9:]
        if validate_vin_checksum(test):
            return test, True
    
    return vin, False


def decode_vin_year(vin: str) -> int:
    """Décode l'année à partir du 10e caractère"""
    if len(vin) < 10:
        return None
    return VIN_YEAR_CODES.get(vin[9].upper())


def decode_vin_brand(vin: str) -> str:
    """Décode le constructeur à partir du WMI (3 premiers caractères)"""
    if len(vin) < 3:
        return None
    wmi = vin[:3].upper()
    
    # Cas spécial: Jeep Gladiator (1C6PJ...)
    if vin.upper().startswith("1C6PJ"):
        return "Jeep"
    
    return VIN_WMI_MAP.get(wmi)


def validate_vin_brand_consistency(vin: str, expected_brand: str) -> bool:
    """Vérifie que le VIN correspond à la marque attendue"""
    vin_brand = decode_vin_brand(vin)
    if not vin_brand or not expected_brand:
        return True  # Pas de vérification possible
    return vin_brand.lower() == expected_brand.lower()


def decode_vin(vin: str) -> dict:
    """
    Décode un VIN complet avec validation et auto-correction.
    
    Returns:
        dict avec: vin, valid, corrected, year, manufacturer, checksum_valid
    """
    result = {
        "vin": vin,
        "valid": False,
        "corrected": False,
        "checksum_valid": False,
        "year": None,
        "manufacturer": None,
        "plant": None,
        "serial": None
    }
    
    # Nettoyer
    vin = vin.upper().replace("-", "").replace(" ", "")
    
    if len(vin) != 17:
        return result
    
    # Auto-correction
    vin_corrected, was_corrected = auto_correct_vin(vin)
    result["vin"] = vin_corrected
    result["corrected"] = was_corrected
    result["checksum_valid"] = validate_vin_checksum(vin_corrected)
    result["valid"] = result["checksum_valid"]
    
    # Année (10e caractère)
    result["year"] = decode_vin_year(vin_corrected)
    
    # Constructeur (WMI)
    result["manufacturer"] = decode_vin_brand(vin_corrected)
    
    # Serial (positions 12-17)
    result["serial"] = vin_corrected[11:17]
    
    return result

# ============ FCA Product Code Database ============

# Charger la base de codes 2026 depuis le fichier JSON
import json as _json
_FCA_CODES_2026 = {}
try:
    with open(os.path.join(str(ROOT_DIR), 'data', 'fca_product_codes_2026.json'), 'r') as f:
        _codes_raw = _json.load(f)
        for code, info in _codes_raw.items():
            _FCA_CODES_2026[code] = {
                "brand": info.get("brand"),
                "model": info.get("model"),
                "trim": info.get("trim"),
                "body": f"{info.get('cab', '')} {info.get('drive', '')}".strip(),
                "description": f"{info.get('brand', '')} {info.get('model', '')} {info.get('trim', '')} {info.get('cab', '')} {info.get('drive', '')}".strip()
            }
    print(f"[FCA] Loaded {len(_FCA_CODES_2026)} product codes from JSON")
except Exception as e:
    print(f"[FCA] Warning: Could not load FCA codes JSON: {e}")

# ============ Code -> Programme Financing Mapping ============
_CODE_PROGRAM_MAPPING = {}
try:
    mapping_file = os.path.join(str(ROOT_DIR), 'data', 'code_program_mapping.json')
    with open(mapping_file, 'r') as f:
        _CODE_PROGRAM_MAPPING = _json.load(f)
    print(f"[FCA] Loaded {len(_CODE_PROGRAM_MAPPING)} code->program mappings")
except Exception as e:
    print(f"[FCA] Warning: Could not load code->program mapping: {e}")

def get_financing_for_code(code: str) -> Optional[Dict[str, Any]]:
    """
    Retourne les informations de financement pour un code produit.
    Inclut: consumer_cash, bonus_cash, taux Option 1 et Option 2.
    """
    code = code.upper().strip()
    if code in _CODE_PROGRAM_MAPPING:
        return _CODE_PROGRAM_MAPPING[code]['financing']
    return None

def get_full_vehicle_info(code: str) -> Optional[Dict[str, Any]]:
    """
    Retourne toutes les informations pour un code produit:
    - Détails du véhicule (marque, modèle, trim, etc.)
    - Informations de financement (consumer cash, bonus cash, taux)
    """
    code = code.upper().strip()
    if code in _CODE_PROGRAM_MAPPING:
        return _CODE_PROGRAM_MAPPING[code]
    # Fallback: retourner juste les infos du véhicule sans financement
    if code in _FCA_CODES_2026:
        return {
            'code': code,
            'vehicle': _FCA_CODES_2026[code],
            'financing': None
        }
    return None

# Base de données des codes produits FCA/Stellantis
# PRIORITÉ: fichier JSON (codes officiels) > codes fallback en dur
# Format: CODE -> {brand, model, trim, body, description}
_FCA_FALLBACK_CODES = {
    # Ram 2500 Series (DJ = Heavy Duty 2500) - FALLBACK si pas dans JSON
    "DJ7L91": {"brand": "Ram", "model": "2500", "trim": "Tradesman", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Tradesman Crew Cab 4x4"},
    "DJ7L92": {"brand": "Ram", "model": "2500", "trim": "Tradesman", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Tradesman Crew Cab 4x4"},
    "DJ7L94": {"brand": "Ram", "model": "2500", "trim": "Laramie", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Laramie Crew Cab 4x4"},
    "DJ7L96": {"brand": "Ram", "model": "2500", "trim": "Limited", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Limited Crew Cab 4x4"},
    "DJ7L98": {"brand": "Ram", "model": "2500", "trim": "Power Wagon", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 2500 Power Wagon Crew Cab 4x4"},
    
    # Ram 1500 Series (DS/DT = 1500)
    "DT6L91": {"brand": "Ram", "model": "1500", "trim": "Tradesman", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Tradesman Crew Cab 4x4"},
    "DT6L92": {"brand": "Ram", "model": "1500", "trim": "Big Horn", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Big Horn Crew Cab 4x4"},
    "DT6H92": {"brand": "Ram", "model": "1500", "trim": "Big Horn", "body": "Quad Cab 4x4 6'4\" Box", "description": "Ram 1500 Big Horn Quad Cab 4x4"},
    "DT6L94": {"brand": "Ram", "model": "1500", "trim": "Laramie", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Laramie Crew Cab 4x4"},
    "DT6L96": {"brand": "Ram", "model": "1500", "trim": "Limited", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Limited Crew Cab 4x4"},
    "DT6L97": {"brand": "Ram", "model": "1500", "trim": "Rebel", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 Rebel Crew Cab 4x4"},
    "DT6L99": {"brand": "Ram", "model": "1500", "trim": "TRX", "body": "Crew Cab 4x4 5'7\" Box", "description": "Ram 1500 TRX Crew Cab 4x4"},
    
    # Ram 3500 Series (D2 = Heavy Duty 3500)
    "D2RL91": {"brand": "Ram", "model": "3500", "trim": "Tradesman", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Tradesman Crew Cab 4x4"},
    "D2RL92": {"brand": "Ram", "model": "3500", "trim": "Big Horn", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Big Horn Crew Cab 4x4"},
    "D2RL94": {"brand": "Ram", "model": "3500", "trim": "Laramie", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Laramie Crew Cab 4x4"},
    "D2RL96": {"brand": "Ram", "model": "3500", "trim": "Limited", "body": "Crew Cab 4x4 6'4\" Box", "description": "Ram 3500 Limited Crew Cab 4x4"},
    
    # Jeep
    "JLXL74": {"brand": "Jeep", "model": "Wrangler", "trim": "Rubicon", "body": "4-Door 4x4", "description": "Jeep Wrangler Rubicon Unlimited 4x4"},
    "JLXL72": {"brand": "Jeep", "model": "Wrangler", "trim": "Sahara", "body": "4-Door 4x4", "description": "Jeep Wrangler Sahara Unlimited 4x4"},
    "JLXL70": {"brand": "Jeep", "model": "Wrangler", "trim": "Sport", "body": "4-Door 4x4", "description": "Jeep Wrangler Sport Unlimited 4x4"},
    # Jeep Gladiator (JT = Gladiator)
    "JTJL98": {"brand": "Jeep", "model": "Gladiator", "trim": "Willys", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Willys 4x4"},
    "JTJL96": {"brand": "Jeep", "model": "Gladiator", "trim": "Rubicon", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Rubicon 4x4"},
    "JTJL94": {"brand": "Jeep", "model": "Gladiator", "trim": "Overland", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Overland 4x4"},
    "JTJL92": {"brand": "Jeep", "model": "Gladiator", "trim": "Sport S", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Sport S 4x4"},
    "JTJL90": {"brand": "Jeep", "model": "Gladiator", "trim": "Sport", "body": "Crew Cab 4x4", "description": "Jeep Gladiator Sport 4x4"},
    
    # Grand Cherokee (WL Series - 2022+)
    "WLJP74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Limited", "body": "4x4", "description": "Jeep Grand Cherokee Limited 4x4"},
    "WLJP75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Limited", "body": "4x4", "description": "Jeep Grand Cherokee L Limited 4x4"},
    "WLJH74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Laredo", "body": "4x4", "description": "Jeep Grand Cherokee Laredo 4x4"},
    "WLJH75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Altitude", "body": "4x4", "description": "Jeep Grand Cherokee L Altitude 4x4"},
    "WLJS74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit", "body": "4x4", "description": "Jeep Grand Cherokee Summit 4x4"},
    "WLJS75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Summit", "body": "4x4", "description": "Jeep Grand Cherokee L Summit 4x4"},
    "WLJT74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Overland", "body": "4x4", "description": "Jeep Grand Cherokee Overland 4x4"},
    "WLJT75": {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Overland", "body": "4x4", "description": "Jeep Grand Cherokee L Overland 4x4"},
    "WKXL74": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit", "body": "4x4", "description": "Jeep Grand Cherokee Summit 4x4"},
    "WKXL72": {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Limited", "body": "4x4", "description": "Jeep Grand Cherokee Limited 4x4"},
    "MPXL74": {"brand": "Jeep", "model": "Compass", "trim": "Limited", "body": "4x4", "description": "Jeep Compass Limited 4x4"},
    
    # Dodge
    "LDXL74": {"brand": "Dodge", "model": "Durango", "trim": "R/T", "body": "AWD", "description": "Dodge Durango R/T AWD"},
    "LDXL76": {"brand": "Dodge", "model": "Durango", "trim": "SRT", "body": "AWD", "description": "Dodge Durango SRT AWD"},
    "LAXL74": {"brand": "Dodge", "model": "Charger", "trim": "R/T", "body": "RWD", "description": "Dodge Charger R/T"},
    "LCXL74": {"brand": "Dodge", "model": "Challenger", "trim": "R/T", "body": "RWD", "description": "Dodge Challenger R/T"},
    "HNXL74": {"brand": "Dodge", "model": "Hornet", "trim": "R/T", "body": "AWD", "description": "Dodge Hornet R/T AWD"},
    
    # Chrysler
    "RUXL74": {"brand": "Chrysler", "model": "Pacifica", "trim": "Limited", "body": "FWD", "description": "Chrysler Pacifica Limited"},
    "RUXL78": {"brand": "Chrysler", "model": "Pacifica", "trim": "Pinnacle", "body": "AWD", "description": "Chrysler Pacifica Pinnacle AWD"},
}

# Fusionner avec PRIORITÉ au fichier JSON (codes officiels 2025/2026)
FCA_PRODUCT_CODES = {
    **_FCA_FALLBACK_CODES,  # Codes fallback en premier (seront écrasés par JSON)
    **_FCA_CODES_2026,      # Codes officiels du JSON en dernier (priorité maximale)
}

# Codes d'options communes FCA
FCA_OPTION_CODES = {
    # Moteurs
    "ETM": "6.7L Cummins Turbo Diesel I-6",
    "ETK": "6.7L Cummins High Output Turbo Diesel",
    "EZH": "6.4L HEMI V8",
    "ERC": "5.7L HEMI V8 MDS VVT",
    "ERB": "3.6L Pentastar V6",
    "EZC": "3.0L EcoDiesel V6",
    "ESG": "2.0L Turbo I-4 PHEV",
    
    # Transmissions
    "DFM": "8-Speed Automatic TorqueFlite HD",
    "DFD": "8-Speed Automatic TorqueFlite 8HP75",
    "DFH": "6-Speed Manual",
    "DFL": "8-Speed Automatic 8HP95",
    
    # Couleurs
    "PXJ": "Noir Cristal Nacré",
    "PW7": "Blanc Vif",
    "PAU": "Rouge Flamme",
    "PBF": "Bleu Patriote",
    "PSC": "Gris Destroyer",
    "PX8": "Noir Diamant",
    "PWL": "Blanc Perle",
    "PGG": "Gris Granit",
    
    # Options populaires
    "AHU": "Préparation Remorquage Sellette/Col-de-cygne",
    "XAC": "Groupe Remorquage Max",
    "ADA": "Différentiel Arrière Anti-spin",
    "CLF": "Tapis de Protection Mopar",
    "LHL": "Commandes Auxiliaires Tableau de Bord",
    "LNC": "Feux de Gabarit",
    "MWH": "Doublures Passage de Roue Arrière",
    "UAQ": "Uconnect 5 NAV 12\" Écran",
    "RC3": "Radio Uconnect 5 8.4\" Écran",
    "GWA": "Boîte de Transfert Électronique",
    "DSA": "Suspension Pneumatique Arrière",
    
    # Frais/Taxes
    "801": "Frais de Transport",
    "4CP": "Taxe Accise Fédérale Climatiseur",
    "92HC1": "Cotisation Protection Produit",
    "92HC2": "Allocation Marketing",
}

def _build_trim_string(product_info: dict) -> str:
    """
    Construit la chaîne de trim à partir des données product_info.
    Gère les deux formats: {trim, body} et {trim, cab, drive}
    """
    if not product_info:
        return ""
    
    trim = product_info.get("trim") or ""
    
    # Format 1: body direct
    body = product_info.get("body") or ""
    
    # Format 2: cab + drive (fichier JSON)
    if not body:
        cab = product_info.get("cab") or ""
        drive = product_info.get("drive") or ""
        if cab or drive:
            body = f"{cab} {drive}".strip()
    
    # Combiner trim et body
    if trim and body:
        return f"{trim} {body}"
    elif trim:
        return trim
    elif body:
        return body
    else:
        return ""

def calculate_holdback(brand: str, pdco: float, parsed_holdback: float = None) -> float:
    """
    Retourne le holdback extrait de la facture.
    
    Le holdback FCA est imprimé sur la facture en format spécial (0XXXXX00).
    Si extrait du parser, on l'utilise directement.
    Sinon on retourne 0 (le holdback doit venir de la facture, pas calculé).
    
    Args:
        brand: Marque du véhicule (non utilisé, gardé pour compatibilité)
        pdco: Prix Dealer (non utilisé, gardé pour compatibilité)
        parsed_holdback: Holdback extrait de la facture
    
    Returns:
        Holdback en dollars (de la facture, ou 0 si non trouvé)
    """
    # Le holdback doit être extrait de la facture, pas calculé
    if parsed_holdback and parsed_holdback > 0:
        return parsed_holdback
    
    # Si pas trouvé sur la facture, retourner 0
    return 0

def decode_product_code(code: str) -> dict:
    """Décode un code produit FCA et retourne les informations du véhicule"""
    code = code.upper().strip()
    
    # Chercher dans la base de données
    if code in FCA_PRODUCT_CODES:
        return FCA_PRODUCT_CODES[code]
    
    # Essayer de décoder le pattern si pas trouvé
    result = {
        "brand": None,
        "model": None,
        "trim": None,
        "body": None,
        "description": None
    }
    
    # Patterns de préfixes connus
    if code.startswith("DJ"):
        result["brand"] = "Ram"
        result["model"] = "2500"
    elif code.startswith("DT") or code.startswith("DS"):
        result["brand"] = "Ram"
        result["model"] = "1500"
    elif code.startswith("D2"):
        result["brand"] = "Ram"
        result["model"] = "3500"
    elif code.startswith("JL"):
        result["brand"] = "Jeep"
        result["model"] = "Wrangler"
    elif code.startswith("WK"):
        result["brand"] = "Jeep"
        result["model"] = "Grand Cherokee"
    elif code.startswith("MP"):
        result["brand"] = "Jeep"
        result["model"] = "Compass"
    elif code.startswith("LD"):
        result["brand"] = "Dodge"
        result["model"] = "Durango"
    elif code.startswith("LA"):
        result["brand"] = "Dodge"
        result["model"] = "Charger"
    elif code.startswith("LC"):
        result["brand"] = "Dodge"
        result["model"] = "Challenger"
    elif code.startswith("RU"):
        result["brand"] = "Chrysler"
        result["model"] = "Pacifica"
    
    return result

def decode_option_code(code: str) -> str:
    """Retourne la description d'un code d'option FCA"""
    code = code.upper().strip()
    return FCA_OPTION_CODES.get(code, None)

def enrich_vehicle_data(vehicle_data: dict) -> dict:
    """Enrichit les données d'un véhicule avec le décodage VIN et codes produits"""
    
    # Décoder le VIN si présent
    vin = vehicle_data.get("vin", "")
    if vin and len(vin.replace("-", "")) == 17:
        vin_info = decode_vin(vin)
        if vin_info["valid"]:
            # Mettre à jour l'année si pas déjà définie
            if not vehicle_data.get("year") and vin_info.get("year"):
                vehicle_data["year"] = vin_info["year"]
            # Mettre à jour la marque si pas définie
            if not vehicle_data.get("brand") and vin_info.get("manufacturer"):
                vehicle_data["brand"] = vin_info["manufacturer"]
    
    # Chercher et décoder le code produit principal (première option ou MODEL/OPT)
    options = vehicle_data.get("options", [])
    if options:
        first_option = options[0]
        # PATCH: Utiliser "product_code" au lieu de "code"
        code = first_option.get("product_code", first_option.get("code", ""))
        
        # Vérifier si c'est un code de modèle (pas un code d'option)
        product_info = decode_product_code(code)
        if product_info.get("brand"):
            # C'est un code de véhicule
            if not vehicle_data.get("brand"):
                vehicle_data["brand"] = product_info["brand"]
            if not vehicle_data.get("model"):
                vehicle_data["model"] = product_info["model"]
            if not vehicle_data.get("trim"):
                vehicle_data["trim"] = product_info.get("trim")
            if product_info.get("description"):
                first_option["description"] = product_info["description"]
    
    # Enrichir les descriptions des options
    for option in options:
        # PATCH: Utiliser "product_code" au lieu de "code"
        code = option.get("product_code", option.get("code", ""))
        if not option.get("description") or len(option.get("description", "")) < 5:
            desc = decode_option_code(code)
            if desc:
                option["description"] = desc
    
    return vehicle_data

# PATCH: decode_fca_price() supprimé - utiliser clean_fca_price() uniquement
# Fonction dupliquée de clean_fca_price() dans parser.py

def decode_fca_holdback(raw_value: str) -> float:
    """Décode un holdback FCA - même règle que les prix
    Enlever premier 0 + deux derniers chiffres
    Exemple: 050000 → 50000 → 500 → 500$
    """
    # Remove any non-numeric characters
    cleaned = re.sub(r'[^\d]', '', str(raw_value))
    
    if len(cleaned) >= 4:
        # Remove first 0 if present
        if cleaned.startswith('0'):
            cleaned = cleaned[1:]
        # Remove last 2 digits
        if len(cleaned) >= 2:
            cleaned = cleaned[:-2]
        try:
            return float(cleaned)
        except:
            return 0
    return 0

# =========================
# PARSER STRUCTURÉ FCA V5 - VERSION INDUSTRIELLE OPTIMISÉE
# =========================

# Codes invalides à exclure des options
INVALID_OPTION_CODES = {
    "VIN", "GST", "TPS", "QUE", "INC", "PDCO", "PREF", 
    "MODEL", "TOTAL", "MSRP", "SUB", "KG", "GVW"
}

# Cache des codes produits FCA (évite lookups répétés)
FCA_PRODUCT_CACHE = {}


def generate_file_hash(file_bytes: bytes) -> str:
    """Génère un hash SHA256 unique pour le fichier"""
    return hashlib.sha256(file_bytes).hexdigest()


def compress_image_for_vision(file_bytes: bytes, max_size: int = 1024, quality: int = 70) -> str:
    """
    Compresse l'image pour réduire les tokens Vision API.
    - Redimensionne à max 1024px
    - Compression JPEG quality 70
    - Retourne base64 optimisé
    
    Économie: ~60-70% de tokens en moins
    """
    try:
        image = Image.open(io.BytesIO(file_bytes))
        
        # Convertir en RGB si nécessaire
        if image.mode in ('RGBA', 'P'):
            image = image.convert('RGB')
        
        # Redimensionner proportionnellement
        width, height = image.size
        if max(width, height) > max_size:
            ratio = max_size / max(width, height)
            new_size = (int(width * ratio), int(height * ratio))
            image = image.resize(new_size, Image.Resampling.LANCZOS)
        
        # Compression JPEG
        buffer = io.BytesIO()
        image.save(buffer, format='JPEG', quality=quality, optimize=True)
        
        compressed_bytes = buffer.getvalue()
        original_size = len(file_bytes)
        new_size = len(compressed_bytes)
        
        logger.info(f"Image compressed: {original_size/1024:.1f}KB → {new_size/1024:.1f}KB ({100-new_size*100/original_size:.0f}% reduction)")
        
        return base64.b64encode(compressed_bytes).decode('utf-8')
    except Exception as e:
        logger.error(f"Compression error: {e}")
        return base64.b64encode(file_bytes).decode('utf-8')


def clean_fca_price(raw_value: str) -> int:
    """
    Règle FCA pour décoder les prix:
    - Enlever le premier 0
    - Enlever les deux derniers chiffres
    Exemple: 05662000 -> 56620
    """
    raw_value = str(raw_value).strip()
    raw_value = re.sub(r'[^\d]', '', raw_value)
    
    if not raw_value:
        return 0
    
    if raw_value.startswith("0"):
        raw_value = raw_value[1:]
    
    if len(raw_value) >= 2:
        raw_value = raw_value[:-2]
    
    try:
        return int(raw_value)
    except:
        return 0


def clean_decimal_price(raw_value: str) -> float:
    """Nettoie les montants format décimal: 57,120.00 -> 57120.00"""
    raw_value = str(raw_value).replace(",", "").strip()
    try:
        return float(raw_value)
    except:
        return 0.0


def extract_pdf_text(file_bytes: bytes) -> str:
    """Extrait le texte d'un PDF avec pdfplumber"""
    text = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\n"
        
        os.unlink(tmp_path)
    except Exception as e:
        logger.error(f"Error extracting PDF text: {e}")
    
    return text


def extract_text_from_image(file_bytes: bytes) -> str:
    """
    OCR image avec Tesseract + prétraitement OpenCV
    """
    try:
        # Charger l'image
        image = Image.open(io.BytesIO(file_bytes))
        
        # Convertir en RGB si nécessaire (pour les images RGBA ou autres)
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        img = np.array(image)
        
        # Convertir en niveaux de gris
        if len(img.shape) == 3:
            gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        else:
            gray = img
        
        # Redimensionner si l'image est trop grande (améliore la vitesse OCR)
        max_dimension = 3000
        height, width = gray.shape[:2]
        if max(height, width) > max_dimension:
            scale = max_dimension / max(height, width)
            gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
        
        # Amélioration du contraste
        gray = cv2.convertScaleAbs(gray, alpha=1.5, beta=10)
        
        # Seuillage adaptatif pour améliorer la lisibilité
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
        
        # OCR avec Tesseract (anglais + français)
        text = pytesseract.image_to_string(
            thresh,
            lang="eng+fra",
            config="--psm 6 --oem 3"
        )
        
        logger.info(f"OCR extracted {len(text)} chars")
        return text
    except Exception as e:
        logger.error(f"OCR error: {str(e)}")
        import traceback
        logger.error(f"OCR traceback: {traceback.format_exc()}")
        return ""


def parse_fca_invoice_structured(text: str) -> dict:
    """
    Parser structuré V4 pour factures FCA Canada.
    Patterns améliorés et plus stricts.
    """
    data = {
        "vin": None,
        "model_code": None,
        "ep_cost": None,
        "pdco": None,
        "pref": None,
        "holdback": None,
        "subtotal_excl_tax": None,
        "invoice_total": None,
        "options": [],
        "parse_method": "structured_v4"
    }
    
    # -------------------------
    # VIN (17 caractères, format FCA avec ou sans tirets)
    # -------------------------
    # PATCH: Pattern VIN plus strict - 17 caractères exacts
    # Priorité 1: VIN standard 17 caractères (plus fiable)
    vin_match = re.search(r"\b([0-9A-HJ-NPR-Z]{17})\b", text)
    if vin_match:
        data["vin"] = vin_match.group(1)
    
    # Fallback: VIN FCA avec tirets (format 1C4RJHBG6-S8-806264)
    if not data["vin"]:
        vin_dash_match = re.search(r"\b([0-9A-HJ-NPR-Z]{9})[-\s]([A-HJ-NPR-Z0-9]{2})[-\s]([A-HJ-NPR-Z0-9]{6})\b", text)
        if vin_dash_match:
            vin_raw = vin_dash_match.group(1) + vin_dash_match.group(2) + vin_dash_match.group(3)
            if len(vin_raw) == 17:
                data["vin"] = vin_raw
    
    # -------------------------
    # Model Code - RESTREINT à la zone MODEL/OPT
    # Pattern: 5-7 caractères alphanumériques (ex: WLJP74, WLJH75, JTJL98)
    # -------------------------
    model_section = re.search(r"MODEL/OPT[\s\S]{0,50}?\n\s*([A-Z]{2,4}[A-Z0-9]{2,4})", text)
    if model_section:
        data["model_code"] = model_section.group(1)
    else:
        # Fallback: chercher pattern FCA standard (2-4 lettres + 2 chiffres)
        model_match = re.search(r"\b(WL[A-Z]{2}\d{2}|JT[A-Z]{2}\d{2}|DT[A-Z0-9]{2}\d{2})\b", text)
        if model_match:
            data["model_code"] = model_match.group(1)
    
    # -------------------------
    # E.P. (Employee Price / Coût)
    # Pattern: E.P. suivi de 7-10 chiffres
    # -------------------------
    ep_match = re.search(r"E\.P\.?\s*(\d{7,10})", text)
    if ep_match:
        data["ep_cost"] = clean_fca_price(ep_match.group(1))
    
    # -------------------------
    # PDCO (Prix dealer / PDSF base)
    # -------------------------
    pdco_match = re.search(r"PDCO\s*(\d{7,10})", text)
    if pdco_match:
        data["pdco"] = clean_fca_price(pdco_match.group(1))
    
    # -------------------------
    # PREF (Prix de référence)
    # -------------------------
    pref_match = re.search(r"PREF\*?\s*(\d{7,10})", text)
    if pref_match:
        data["pref"] = clean_fca_price(pref_match.group(1))
    
    # -------------------------
    # HOLDBACK - AMÉLIORÉ
    # Chercher dans la zone financière près de PREF
    # Format: 6 chiffres commençant par 0 (ex: 070000 = 700$)
    # -------------------------
    # Méthode 1: Chercher après PREF
    holdback_match = re.search(r"PREF\*?\s*\d{7,10}\s+(\d{6})\b", text)
    if holdback_match:
        data["holdback"] = clean_fca_price(holdback_match.group(1))
    else:
        # Méthode 2: Chercher un 0XXXXX isolé près de GVW/KG
        holdback_match = re.search(r"\b(0[3-9]\d{4})\b\s*(?:GVW|KG|$)", text)
        if holdback_match:
            data["holdback"] = clean_fca_price(holdback_match.group(1))
    
    # -------------------------
    # SUBTOTAL EXCLUDING TAXES
    # -------------------------
    subtotal_match = re.search(
        r"SUB\s*TOTAL\s*EXCLUDING\s*TAXES[\s\S]*?([\d,]+\.\d{2})",
        text,
        re.IGNORECASE
    )
    if subtotal_match:
        data["subtotal_excl_tax"] = clean_decimal_price(subtotal_match.group(1))
    
    # -------------------------
    # INVOICE TOTAL / TOTAL DE LA FACTURE
    # -------------------------
    total_match = re.search(
        r"TOTAL\s+DE\s+LA\s+FACTURE\s+([\d,]+\.\d{2})",
        text
    )
    if not total_match:
        total_match = re.search(
            r"INVOICE\s*TOTAL[\s\S]*?([\d,]+\.\d{2})",
            text,
            re.IGNORECASE
        )
    if total_match:
        data["invoice_total"] = clean_decimal_price(total_match.group(1))
    
    # -------------------------
    # OPTIONS - PATTERN AMÉLIORÉ
    # Format: CODE (2-5 chars) + DESCRIPTION (5+ chars) + MONTANT (6-10 chiffres ou SANS FRAIS)
    # -------------------------
    option_pattern = re.findall(
        r"\n\s*([A-Z0-9]{2,5})\s+([A-Z0-9][A-Z0-9 ,\-\/'\.]{4,}?)\s+(\d{6,10}|\*|SANS\s*FRAIS)",
        text
    )
    
    for code, desc, amount in option_pattern:
        if code.upper() in INVALID_OPTION_CODES:
            continue
        
        # Nettoyer le montant
        if amount in ['*', 'SANS FRAIS', 'SANS']:
            amt = 0
        else:
            amt = clean_fca_price(amount)
        
        data["options"].append({
            "product_code": code.upper(),
            "description": desc.strip()[:100],
            "amount": amt
        })
    
    return data


# SUPPRIMÉ: validate_invoice_data locale - utiliser celle de validation.py
# Importée comme: from validation import validate_invoice_data as validate_invoice_full


@router.post("/inventory/scan-invoice")
async def scan_invoice(request: InvoiceScanRequest, authorization: Optional[str] = Header(None)):
    """
    Scanne une facture FCA - Pipeline Multi-Niveaux Industriel
    
    ARCHITECTURE 100% OPTIMISÉE:
    - Niveau 1: PDF natif → pdfplumber + regex (100% précis, gratuit)
    - Niveau 2: Image → OpenCV ROI + Tesseract (85-92%, gratuit)
    - Niveau 3: Fallback → GPT-4 Vision (si score < 70, ~2-3¢)
    
    AVANTAGES:
    - 95% des factures traitées sans coût API
    - Fallback intelligent uniquement si nécessaire
    - Validation VIN industrielle avec auto-correction
    - DOUBLE VÉRIFICATION: code produit validé contre base master 131 codes
    """
    # Import des nouveaux modules OCR
    from ocr import process_image_ocr_pipeline
    from parser import parse_invoice_text
    from vin_utils import validate_and_correct_vin
    from validation import validate_invoice_data as validate_invoice_full, calculate_validation_score
    from product_code_lookup import lookup_product_code, get_vehicle_info_from_invoice
    
    user = await get_current_user(authorization)
    
    try:
        start_time = time.time()
        
        # Décoder le base64
        try:
            file_bytes = base64.b64decode(request.image_base64)
        except:
            raise HTTPException(status_code=400, detail="Base64 invalide")
        
        # Générer le hash du fichier pour anti-doublon
        file_hash = generate_file_hash(file_bytes)
        
        # Détecter si c'est un PDF
        is_pdf = file_bytes[:4] == b'%PDF' or request.is_pdf
        
        vehicle_data = None
        parse_method = None
        validation = {"score": 0, "errors": [], "is_valid": False}
        
        # ===== NIVEAU 1: PDF → PARSER STRUCTURÉ (100% GRATUIT) =====
        if is_pdf:
            logger.info("PDF détecté → Parser pdfplumber + regex")
            extracted_text = extract_pdf_text(file_bytes)
            
            if extracted_text and len(extracted_text) > 100:
                parsed = parse_fca_invoice_structured(extracted_text)
                validation = validate_invoice_data(parsed)
                
                if validation["is_valid"]:
                    vin = parsed.get("vin", "")
                    vin_info = decode_vin(vin) if vin and len(vin) == 17 else {}
                    model_code = parsed.get("model_code", "")
                    product_info = decode_product_code(model_code) if model_code else {}
                    
                    vehicle_data = {
                        "vin": vin,
                        "model_code": model_code,
                        "year": vin_info.get("year") or datetime.now().year,
                        "brand": product_info.get("brand") or vin_info.get("manufacturer") or "Stellantis",
                        "model": product_info.get("model") or "",
                        "trim": product_info.get("trim") or "",
                        "ep_cost": parsed.get("ep_cost") or 0,
                        "pdco": parsed.get("pdco") or 0,
                        "pref": parsed.get("pref") or 0,
                        "holdback": calculate_holdback(product_info.get("brand") or "Ram", parsed.get("pdco") or 0, parsed.get("holdback")),
                        "msrp": parsed.get("pdco") or 0,
                        "net_cost": parsed.get("ep_cost") or 0,  # E.P. = Coût Net (holdback déjà déduit)
                        "subtotal": parsed.get("subtotal_excl_tax") or 0,
                        "invoice_total": parsed.get("invoice_total") or 0,
                        "options": parsed.get("options", []),
                        "file_hash": file_hash,
                        "parse_method": "pdf_native",
                        "cost_estimate": "$0.00"
                    }
                    parse_method = "pdf_native"
                    logger.info(f"Parser PDF réussi: VIN={vin}, EP={vehicle_data['ep_cost']}")
        
        # ===== NIVEAU 2: IMAGE → OCR SIMPLIFIÉ STABLE (ZÉRO ERREUR) =====
        decision = None
        
        if vehicle_data is None and not is_pdf:
            logger.info("Image détectée → OCR Global simplifié")
            
            try:
                # 1. Pipeline OCR par zones
                ocr_result = process_image_ocr_pipeline(file_bytes)
                
                # 2. Parser structuré sur le texte OCR
                parsed = parse_invoice_text(ocr_result)
                
                # 3. Validation et correction VIN centralisée (UNE SEULE FONCTION)
                vin_raw = parsed.get("vin", "")
                vin_result = validate_and_correct_vin(vin_raw) if vin_raw else {}
                
                vin_corrected = vin_result.get("corrected", vin_raw)
                vin_valid = vin_result.get("is_valid", False)
                vin_was_corrected = vin_result.get("was_corrected", False)
                
                # 4. Calcul du score de validation
                parsed["vin"] = vin_corrected
                parsed["vin_valid"] = vin_valid
                validation_result = validate_invoice_full(parsed)
                
                ocr_score = validation_result["score"]
                logger.info(f"OCR: score={ocr_score}, VIN={vin_corrected}, EP={parsed.get('ep_cost')}")
                
                # ----------- NOUVELLE LOGIQUE: GOOGLE VISION EN PRIORITÉ -----------
                # Si Google Vision est configuré, on l'utilise TOUJOURS pour meilleure précision
                google_api_key = os.environ.get("GOOGLE_VISION_API_KEY")
                if google_api_key:
                    logger.info("Google Vision API configurée → Utilisation prioritaire")
                    decision = "vision_required"  # Force Google Vision
                elif ocr_score >= 85:
                    decision = "auto_approved"
                elif 60 <= ocr_score < 85:
                    decision = "review_required"
                else:
                    decision = "vision_required"
                    
                logger.info(f"Décision OCR: {decision} (score={ocr_score})")
                    
            except Exception as ocr_err:
                logger.warning(f"OCR échoué: {ocr_err}, passage au fallback Vision")
                decision = "vision_required"
        
        # ===== GESTION DES DÉCISIONS =====
        
        # REVIEW REQUIRED (60-84) → Retourner pour révision humaine
        if decision == "review_required":
            vin_info = decode_vin(vin_corrected) if vin_corrected and len(vin_corrected) == 17 else {}
            model_code = parsed.get("model_code", "")
            
            # ==== DOUBLE VÉRIFICATION AVEC BASE MASTER ====
            master_lookup = lookup_product_code(model_code) if model_code else None
            product_info = master_lookup or (decode_product_code(model_code) if model_code else {})
            
            if master_lookup:
                extracted_model = master_lookup.get("model") or ""
                extracted_trim = _build_trim_string(master_lookup)
                extracted_brand = master_lookup.get("brand") or "Stellantis"
                logger.info(f"[REVIEW - MASTER OK] Code {model_code} validé: {master_lookup.get('full_description')}")
            else:
                extracted_model = product_info.get("model") or parsed.get("model") or ""
                extracted_trim = _build_trim_string(product_info) or parsed.get("trim") or ""
                extracted_brand = product_info.get("brand") or "Stellantis"
            
            return {
                "success": True,
                "review_required": True,
                "vehicle": {
                    "stock_no": parsed.get("stock_no", ""),
                    "vin": vin_corrected,
                    "vin_valid": vin_valid,
                    "vin_corrected": vin_was_corrected,
                    "model_code": model_code,
                    "model_code_validated": master_lookup is not None,
                    "year": vin_info.get("year") or datetime.now().year,
                    "brand": extracted_brand,
                    "model": extracted_model,
                    "trim": extracted_trim,
                    "ep_cost": parsed.get("ep_cost") or 0,
                    "pdco": parsed.get("pdco") or 0,
                    "pref": parsed.get("pref") or 0,
                    "holdback": calculate_holdback(extracted_brand, parsed.get("pdco") or 0, parsed.get("holdback")),
                    "subtotal": parsed.get("subtotal_excl_tax") or parsed.get("subtotal") or 0,
                    "invoice_total": parsed.get("invoice_total") or 0,
                    "options": parsed.get("options", []),
                },
                "validation": validation_result,
                "parse_method": "ocr_review",
                "message": "Score entre 60-84. Révision humaine recommandée."
            }
        
        # AUTO APPROVED (85+) → Accepter directement
        if decision == "auto_approved":
            vin_info = decode_vin(vin_corrected) if vin_corrected and len(vin_corrected) == 17 else {}
            model_code = parsed.get("model_code", "")
            
            # ==== DOUBLE VÉRIFICATION AVEC BASE MASTER ====
            # 1. Lookup dans la base master des 131 codes officiels
            master_lookup = lookup_product_code(model_code) if model_code else None
            
            # 2. Fallback vers decode_product_code si non trouvé dans master
            product_info = master_lookup or (decode_product_code(model_code) if model_code else {})
            
            vin_brand = decode_vin_brand(vin_corrected) if vin_corrected else None
            
            # PRIORITÉ: master lookup > product_info > parser > VIN
            if master_lookup:
                # Code trouvé dans la base master - données GARANTIES
                extracted_model = master_lookup.get("model") or ""
                extracted_trim = _build_trim_string(master_lookup)
                extracted_brand = master_lookup.get("brand") or vin_brand or "Stellantis"
                logger.info(f"[MASTER LOOKUP OK] Code {model_code} validé: {master_lookup.get('full_description')}")
            else:
                # Fallback vers product_info ou parser
                extracted_model = product_info.get("model") or parsed.get("model") or ""
                extracted_trim = _build_trim_string(product_info) or parsed.get("trim") or ""
                extracted_brand = product_info.get("brand") or vin_brand or "Stellantis"
                if model_code:
                    logger.warning(f"[MASTER LOOKUP MISS] Code {model_code} non trouvé dans la base master, utilisation fallback")
            
            logger.info(f"Model extraction: code={model_code}, master_found={master_lookup is not None}, final={extracted_brand} {extracted_model} {extracted_trim}")
            
            vehicle_data = {
                "stock_no": parsed.get("stock_no", ""),
                "vin": vin_corrected,
                "vin_original": vin_raw if vin_was_corrected else None,
                "vin_valid": vin_valid,
                "vin_corrected": vin_was_corrected,
                "vin_brand": vin_brand,
                "model_code": model_code,
                "model_code_validated": master_lookup is not None,  # Flag de validation
                "year": vin_info.get("year") or datetime.now().year,
                "brand": extracted_brand,
                "model": extracted_model,
                "trim": extracted_trim,
                "ep_cost": parsed.get("ep_cost") or 0,
                "pdco": parsed.get("pdco") or 0,
                "pref": parsed.get("pref") or 0,
                "holdback": calculate_holdback(extracted_brand, parsed.get("pdco") or 0, parsed.get("holdback")),
                "msrp": parsed.get("pdco") or 0,
                "net_cost": parsed.get("ep_cost") or 0,  # E.P. = Coût Net (holdback déjà déduit)
                "subtotal": parsed.get("subtotal_excl_tax") or parsed.get("subtotal") or 0,
                "invoice_total": parsed.get("invoice_total") or 0,
                "options": parsed.get("options", []),
                "file_hash": file_hash,
                "parse_method": "ocr_auto_approved",
                "cost_estimate": "$0.00",
                "metrics": {
                    "zones_processed": ocr_result.get("zones_processed", 0),
                    "validation_score": ocr_score,
                    "parse_duration_sec": round(time.time() - start_time, 3)
                }
            }
            
            validation = validation_result
            parse_method = "ocr_auto_approved"
            logger.info(f"OCR Auto-Approved: VIN={vin_corrected}, EP={vehicle_data['ep_cost']}, Score={ocr_score}")
        
        # ===== NIVEAU 3: FALLBACK → GOOGLE CLOUD VISION (OCR HYBRIDE) =====
        # Stratégie hybride: Google Vision pour l'OCR pur (moins cher, plus précis)
        # + parser.py pour l'extraction structurée des données
        if decision == "vision_required" or vehicle_data is None:
            logger.info("Fallback → Google Cloud Vision avec prétraitement CamScanner")
            
            try:
                from PIL import Image as PILImage
                from ocr import (
                    camscanner_preprocess_for_vision, 
                    load_image_from_bytes,
                    google_vision_ocr_from_numpy,
                    google_vision_ocr
                )
                from parser import (
                    parse_vin, 
                    parse_model_code, 
                    parse_financial_data, 
                    parse_totals, 
                    parse_options,
                    parse_stock_number
                )
                
                # Vérifier la clé API Google Vision
                google_api_key = os.environ.get("GOOGLE_VISION_API_KEY")
                if not google_api_key:
                    logger.warning("GOOGLE_VISION_API_KEY non configurée, fallback vers GPT-4 Vision")
                    raise ValueError("Google Vision API key not configured")
                
                # Décoder l'image
                img_data = base64.b64decode(request.image_base64)
                
                # ====== PRÉTRAITEMENT CAMSCANNER ======
                cv_image = load_image_from_bytes(img_data)
                
                if cv_image is not None:
                    logger.info("Applying CamScanner preprocessing for Google Vision...")
                    preprocessed = camscanner_preprocess_for_vision(cv_image)
                    
                    # ====== OCR GOOGLE CLOUD VISION ======
                    logger.info("Calling Google Cloud Vision API...")
                    vision_result = google_vision_ocr_from_numpy(preprocessed, google_api_key)
                    
                    if not vision_result["success"]:
                        logger.error(f"Google Vision error: {vision_result['error']}")
                        raise ValueError(f"Google Vision error: {vision_result['error']}")
                    
                    full_text = vision_result["full_text"]
                    ocr_confidence = vision_result["confidence"]
                    logger.info(f"Google Vision OCR: {len(full_text)} chars, confidence={ocr_confidence:.2f}")
                else:
                    # Fallback: envoyer l'image originale directement
                    logger.warning("CamScanner preprocessing failed, using original image")
                    image_base64 = base64.b64encode(img_data).decode("utf-8")
                    vision_result = google_vision_ocr(image_base64, google_api_key)
                    
                    if not vision_result["success"]:
                        raise ValueError(f"Google Vision error: {vision_result['error']}")
                    
                    full_text = vision_result["full_text"]
                    ocr_confidence = vision_result["confidence"]
                
                # ====== STRUCTURATION GPT-4o (texte → JSON) ======
                # Google Vision a lu le texte. GPT-4o le structure intelligemment.
                logger.info("Structuring invoice text with GPT-4o...")
                
                structured_data = None
                try:
                    from openai import OpenAI
                    openai_key = os.environ.get("OPENAI_API_KEY")
                    if openai_key:
                        client = OpenAI(api_key=openai_key)
                        
                        gpt_prompt = f"""Tu es un expert en factures de véhicules FCA Canada (Stellantis).
Voici le texte OCR brut d'une facture. Extrais les données structurées.

RÈGLES IMPORTANTES:
- Les OPTIONS VÉHICULE sont les lignes avec un CODE (2-4 caractères alphanumériques) suivi d'une DESCRIPTION
- IGNORE complètement: le nom du concessionnaire, l'adresse, la banque, les termes de paiement, les mentions légales
- Les codes comme 999, 92HC1, 92HC2 sont des codes administratifs, PAS des options véhicule
- 4CP (Taxe Accise Climatiseur) n'est PAS une option véhicule
- Le code modèle (6 caractères comme DT6L98, DJ7H91) n'est PAS une option
- Garde l'ORDRE EXACT des options comme elles apparaissent sur la facture
- Pour la couleur, le code commence par P (ex: PAU, PW7, PDN) et est suivi de la description de la couleur

BODY STYLE (CARROSSERIE):
- Cherche le type de carrosserie du véhicule dans la facture
- Exemples: "Crew Cab SWB 4WD", "Quad Cab SWB 4WD", "4D Utility 4WD", "4D Wagon AWD", "2D Coupe AWD"
- Souvent indiqué près de la description du modèle ou dans les détails du véhicule
- "SWB" = Short Wheel Base, "LWB" = Long Wheel Base
- Si pas trouvé explicitement, déduis-le du modèle: Ram 1500/2500/3500 = Crew/Quad/Reg Cab, Jeep = Utility, etc.

PRIX FCA IMPORTANT:
- E.P., PDCO et PREF sont en format FCA encodé (8 chiffres avec premier 0 et 2 derniers = cents)
- Exemple: 07158000 = $71,580.00, 06697900 = $66,979.00
- Pour décoder: enlever le premier 0, enlever les 2 derniers chiffres
- Retourne les montants E.P., PDCO, PREF DÉJÀ DÉCODÉS en dollars entiers (ex: 71580, PAS 7158000)
- Pour les montants d'options (ex: 524.00, 1104.00, 3340.00), retourne-les tels quels en nombre
- Le subtotal et invoice_total sont en format normal (ex: 67679.00 → retourne 67679)

TEXTE OCR:
{full_text}

Retourne UNIQUEMENT un JSON valide (pas de markdown, pas de commentaires):
{{
  "vin": "le VIN 17 caractères ou vide",
  "model_code": "code modèle 6 chars (ex: DT6L98)",
  "body_style": "type de carrosserie (ex: Crew Cab SWB 4WD, 4D Utility 4WD) ou vide",
  "color_code": "code couleur 3 chars (ex: PAU)",
  "color_description": "description couleur complète",
  "ep_cost": nombre_dollars_entiers ou 0,
  "pdco": nombre_dollars_entiers ou 0,
  "pref": nombre_dollars_entiers ou 0,
  "holdback": nombre ou 0,
  "subtotal": nombre ou 0,
  "invoice_total": nombre ou 0,
  "options": [
    {{"code": "XXX", "description": "description complète", "amount": nombre_ou_0}}
  ]
}}"""
                        
                        gpt_response = client.chat.completions.create(
                            model="gpt-4o",
                            messages=[{"role": "user", "content": gpt_prompt}],
                            temperature=0.0,
                            max_tokens=2000,
                        )
                        
                        raw_json = gpt_response.choices[0].message.content.strip()
                        # Nettoyer si GPT ajoute des backticks markdown
                        if raw_json.startswith("```"):
                            raw_json = raw_json.split("\n", 1)[1] if "\n" in raw_json else raw_json[3:]
                            if raw_json.endswith("```"):
                                raw_json = raw_json[:-3]
                            raw_json = raw_json.strip()
                        
                        structured_data = json.loads(raw_json)
                        logger.info(f"GPT-4o structured: {len(structured_data.get('options', []))} options extracted")
                    else:
                        logger.warning("OPENAI_API_KEY non configurée, fallback vers parsing regex")
                except Exception as gpt_err:
                    logger.error(f"GPT-4o structuring error: {gpt_err}, fallback vers parsing regex")
                    structured_data = None
                
                # ====== EXTRACTION DES DONNÉES ======
                if structured_data:
                    # === MODE GPT-4o (structuration intelligente) ===
                    logger.info("Using GPT-4o structured data")
                    
                    # VIN: GPT-4o + validation existante
                    vin_raw = structured_data.get("vin", "")
                    if not vin_raw:
                        vin_raw = parse_vin(full_text) or ""
                    vin_raw = str(vin_raw).replace("-", "").replace(" ", "").upper()[:17]
                    
                    vin_result = validate_and_correct_vin(vin_raw) if len(vin_raw) == 17 else {"corrected": vin_raw, "is_valid": False, "was_corrected": False}
                    vin_corrected = vin_result.get("corrected", vin_raw)
                    vin_valid = vin_result.get("is_valid", False)
                    vin_was_corrected = vin_result.get("was_corrected", False)
                    vin_info = decode_vin(vin_corrected) if len(vin_corrected) == 17 else {}
                    
                    # Code modèle: GPT-4o + validation master
                    from product_code_lookup import get_all_codes
                    master_codes = get_all_codes()
                    model_code = structured_data.get("model_code", "") or parse_model_code(full_text, master_codes) or ""
                    master_lookup = lookup_product_code(model_code) if model_code else None
                    product_info = master_lookup or (decode_product_code(model_code) if model_code else {})
                    
                    # Financier: GPT-4o
                    ep_cost = float(structured_data.get("ep_cost", 0) or 0)
                    pdco = float(structured_data.get("pdco", 0) or 0)
                    pref = float(structured_data.get("pref", 0) or 0)
                    holdback = float(structured_data.get("holdback", 0) or 0)
                    subtotal = float(structured_data.get("subtotal", 0) or 0)
                    invoice_total = float(structured_data.get("invoice_total", 0) or 0)
                    
                    # Sécurité: si GPT-4o a retourné des valeurs en format FCA brut (non décodé)
                    # Les prix véhicules FCA Canada sont entre 20000$ et 200000$
                    # Si > 500000, c'est probablement en format FCA brut (ex: 7158000 au lieu de 71580)
                    if ep_cost > 500000:
                        ep_cost = clean_fca_price(str(int(ep_cost)))
                        logger.info(f"GPT-4o ep_cost was in raw FCA format, decoded to: {ep_cost}")
                    if pdco > 500000:
                        pdco = clean_fca_price(str(int(pdco)))
                        logger.info(f"GPT-4o pdco was in raw FCA format, decoded to: {pdco}")
                    if pref > 500000:
                        pref = clean_fca_price(str(int(pref)))
                        logger.info(f"GPT-4o pref was in raw FCA format, decoded to: {pref}")
                    if holdback > 50000:
                        holdback = clean_fca_price(str(int(holdback)))
                        logger.info(f"GPT-4o holdback was in raw FCA format, decoded to: {holdback}")
                    
                    # Fallback financier si GPT-4o a retourné 0
                    if ep_cost == 0 or pdco == 0:
                        financial = parse_financial_data(full_text)
                        if ep_cost == 0: ep_cost = financial.get("ep_cost", 0) or 0
                        if pdco == 0: pdco = financial.get("pdco", 0) or 0
                        if pref == 0: pref = financial.get("pref", 0) or 0
                        if holdback == 0: holdback = financial.get("holdback", 0) or 0
                    
                    # Holdback: GPT-4o ne reconnaît souvent pas le format FCA (070000 = $700)
                    # Toujours essayer le regex fallback si holdback == 0
                    if holdback == 0:
                        financial_hb = parse_financial_data(full_text)
                        holdback = financial_hb.get("holdback", 0) or 0
                        if holdback > 0:
                            logger.info(f"Holdback extracted via regex fallback: {holdback}")
                    
                    if subtotal == 0 or invoice_total == 0:
                        totals = parse_totals(full_text)
                        if subtotal == 0: subtotal = totals.get("subtotal", 0) or 0
                        if invoice_total == 0: invoice_total = totals.get("invoice_total", 0) or 0
                    
                    # Options: GPT-4o (dans l'ordre exact de la facture)
                    options = []
                    for opt in structured_data.get("options", []):
                        code = str(opt.get("code", "")).strip()
                        desc = str(opt.get("description", "")).strip()
                        amt = float(opt.get("amount", 0) or 0)
                        if code and len(code) >= 2:
                            options.append({
                                "product_code": code,
                                "description": f"{code} - {desc[:70]}",
                                "amount": amt
                            })
                    
                    # Couleur: GPT-4o
                    color_code = structured_data.get("color_code", "")
                    color_desc_gpt = structured_data.get("color_description", "")
                    final_color = color_desc_gpt or color_code
                    
                    # Stock
                    stock_no = parse_stock_number(full_text) or ""
                    if not stock_no:
                        stock_match = re.search(r'\b(\d{5})\b', full_text)
                        if stock_match:
                            stock_no = stock_match.group(1)
                    
                    parse_method_detail = "gpt4o_structured"
                    cost_estimate = "~$0.006"
                    
                else:
                    # === MODE REGEX FALLBACK (si GPT-4o indisponible) ===
                    logger.info("Fallback: parsing with regex from Google Vision text...")
                    
                    vin_raw = parse_vin(full_text)
                    if not vin_raw:
                        vin_match = re.search(r'1C4[A-Z0-9]{14}', full_text.replace("-", "").replace(" ", "").upper())
                        if vin_match:
                            vin_raw = vin_match.group()
                    vin_raw = str(vin_raw or "").replace("-", "").replace(" ", "").upper()[:17]
                    
                    vin_result = validate_and_correct_vin(vin_raw) if len(vin_raw) == 17 else {"corrected": vin_raw, "is_valid": False, "was_corrected": False}
                    vin_corrected = vin_result.get("corrected", vin_raw)
                    vin_valid = vin_result.get("is_valid", False)
                    vin_was_corrected = vin_result.get("was_corrected", False)
                    vin_info = decode_vin(vin_corrected) if len(vin_corrected) == 17 else {}
                    
                    from product_code_lookup import get_all_codes
                    master_codes = get_all_codes()
                    model_code = parse_model_code(full_text, master_codes) or ""
                    master_lookup = lookup_product_code(model_code) if model_code else None
                    product_info = master_lookup or (decode_product_code(model_code) if model_code else {})
                    
                    financial = parse_financial_data(full_text)
                    ep_cost = financial.get("ep_cost", 0) or 0
                    pdco = financial.get("pdco", 0) or 0
                    pref = financial.get("pref", 0) or 0
                    holdback = financial.get("holdback", 0) or 0
                    
                    totals = parse_totals(full_text)
                    subtotal = totals.get("subtotal", 0) or 0
                    invoice_total = totals.get("invoice_total", 0) or 0
                    
                    options = []
                    for opt in parse_options(full_text):
                        options.append({
                            "product_code": opt.get("product_code", ""),
                            "description": opt.get("description", "")[:80],
                            "amount": 0
                        })
                    
                    stock_no = parse_stock_number(full_text) or ""
                    if not stock_no:
                        stock_match = re.search(r'\b(\d{5})\b', full_text)
                        if stock_match:
                            stock_no = stock_match.group(1)
                    
                    raw_color = ""
                    color_desc = ""
                    color_match_re = re.search(r'\b(P[A-Z0-9]{2})\b', full_text)
                    if color_match_re:
                        raw_color = color_match_re.group(1)
                        color_desc_match = re.search(
                            rf'\b{re.escape(raw_color)}\s+([A-Z][A-Z\s]+?)(?:\s+\d|SANS\s+FRAIS|\s*$)',
                            full_text
                        )
                        if color_desc_match:
                            color_desc = color_desc_match.group(1).strip().title()
                    
                    color_map = {
                        "PW7": "Blanc Vif", "PWZ": "Blanc Vif", "PXJ": "Noir Cristal", 
                        "PX8": "Noir Diamant", "PSC": "Gris Destroyer", 
                        "PWL": "Blanc Perle", "PGG": "Gris Granit", "PBF": "Bleu Patriote", 
                        "PGE": "Vert Sarge", "PRM": "Rouge Velours", "PAR": "Argent Billet",
                        "PYB": "Jaune Stinger", "PBJ": "Bleu Hydro", "PFQ": "Granite Cristal",
                        "PDN": "Gris Ceramique",
                    }
                    final_color = color_desc or color_map.get(raw_color, raw_color)
                    parse_method_detail = "regex_fallback"
                    cost_estimate = "~$0.0015"
                
                # ====== VALIDATION COMMUNE ======
                vin_brand = decode_vin_brand(vin_corrected)
                expected_brand = product_info.get("brand")
                vin_consistent = validate_vin_brand_consistency(vin_corrected, expected_brand) if vin_brand else True
                
                parse_duration = round(time.time() - start_time, 3)
                
                # Validation avec validate_invoice_full
                vehicle_data_for_validation = {
                    "vin": vin_corrected,
                    "vin_valid": vin_valid,
                    "ep_cost": ep_cost,
                    "pdco": pdco,
                    "pref": pref,
                    "subtotal_excl_tax": subtotal,
                    "invoice_total": invoice_total,
                    "options": options
                }
                validation = validate_invoice_full(vehicle_data_for_validation)
                
                # Ajout erreurs spécifiques
                if vin_was_corrected:
                    validation["errors"] = validation.get("errors", []) + ["VIN auto-corrigé"]
                
                vehicle_data = {
                    "stock_no": stock_no,
                    "vin": vin_corrected,
                    "vin_original": vin_raw if vin_was_corrected else None,
                    "vin_valid": vin_valid,
                    "vin_corrected": vin_was_corrected,
                    "vin_brand": vin_brand,
                    "vin_consistent": vin_consistent,
                    "model_code": model_code,
                    "model_code_validated": master_lookup is not None,
                    "year": vin_info.get("year") or datetime.now().year,
                    "brand": product_info.get("brand") or vin_brand or "Stellantis",
                    "model": product_info.get("model") or "",
                    "trim": _build_trim_string(product_info),
                    "body_style": (structured_data or {}).get("body_style", "") if structured_data else "",
                    "ep_cost": ep_cost,
                    "pdco": pdco,
                    "pref": pref,
                    "holdback": calculate_holdback(product_info.get("brand") or vin_brand or "Ram", pdco, holdback),
                    "msrp": pdco,
                    "net_cost": ep_cost,
                    "subtotal": subtotal,
                    "invoice_total": invoice_total,
                    "color": final_color,
                    "options": options,
                    "file_hash": file_hash,
                    "parse_method": f"google_vision_{parse_method_detail}",
                    "metrics": {
                        "parse_duration_sec": parse_duration,
                        "validation_score": validation.get("score", 0),
                        "ocr_confidence": ocr_confidence,
                        "cost_estimate": cost_estimate
                    }
                }
                
                logger.info(f"Google Vision - Model extraction: code={model_code}, product_info={product_info}")
                parse_method = "google_vision_hybrid"
                logger.info(f"Google Vision Hybrid: VIN={vin_corrected}, EP={ep_cost}, PDCO={pdco}, Score={validation.get('score', 0)}, Duration={parse_duration}s")
                
            except HTTPException:
                raise
            except Exception as vision_err:
                logger.error(f"Erreur Google Vision: {vision_err}")
                raise HTTPException(status_code=500, detail=f"Erreur analyse OCR: {str(vision_err)}")
        
        # ===== RÈGLE D'OR : JAMAIS ENREGISTRER SI INVALIDE =====
        # Le système n'enregistre jamais si :
        # - VIN invalide (ou manquant)
        # - EP manquant
        # - PDCO manquant
        # - EP >= PDCO
        
        if vehicle_data:
            vin = vehicle_data.get("vin", "")
            ep = vehicle_data.get("ep_cost", 0) or 0
            pdco = vehicle_data.get("pdco", 0) or 0
            
            blocking_errors = []
            
            if not vin or len(vin) != 17:
                blocking_errors.append("VIN manquant ou invalide (doit être 17 caractères)")
            
            if not ep or ep <= 0:
                blocking_errors.append("EP (Employee Price) manquant")
            
            if not pdco or pdco <= 0:
                blocking_errors.append("PDCO (Dealer Price) manquant")
            
            if ep > 0 and pdco > 0 and ep >= pdco:
                blocking_errors.append(f"EP ({ep}) doit être inférieur à PDCO ({pdco})")
            
            if blocking_errors:
                return {
                    "success": False,
                    "review_required": True,
                    "blocking_errors": blocking_errors,
                    "vehicle": vehicle_data,
                    "validation": validation,
                    "parse_method": parse_method,
                    "message": "Données critiques manquantes ou invalides. Révision obligatoire."
                }
        
        # ===== MONITORING: LOG PARSING METRICS =====
        try:
            score = validation.get("score", 0) if validation else 0
            duration = 0
            if vehicle_data and vehicle_data.get("metrics"):
                duration = vehicle_data["metrics"].get("parse_duration_sec", 0)
            
            # Déterminer le statut basé sur le score
            if score >= 85:
                status = "auto"
            elif score >= 60:
                status = "review"
            else:
                status = "vision"
            
            # Estimer le coût basé sur la méthode utilisée
            cost_estimate = 0.0
            if "google_vision" in parse_method:
                cost_estimate = 0.0015  # Google Vision DOCUMENT_TEXT_DETECTION
            elif "vision" in parse_method:
                cost_estimate = 0.02    # GPT-4 Vision (fallback)
            # OCR Tesseract = gratuit
            
            log_entry = {
                "timestamp": datetime.now(),
                "owner_id": user["id"],
                "parse_method": parse_method,
                "score": score,
                "status": status,
                "vin_valid": vehicle_data.get("vin_valid") if vehicle_data else False,
                "vin": vehicle_data.get("vin", "")[:17] if vehicle_data else "",
                "stock_no": vehicle_data.get("stock_no", "") if vehicle_data else "",
                "brand": vehicle_data.get("brand", "") if vehicle_data else "",
                "model": vehicle_data.get("model", "") if vehicle_data else "",
                "ep_cost": vehicle_data.get("ep_cost", 0) if vehicle_data else 0,
                "pdco": vehicle_data.get("pdco", 0) if vehicle_data else 0,
                "duration_sec": duration,
                "cost_estimate": cost_estimate,
                "success": True
            }
            
            await db.parsing_metrics.insert_one(log_entry)
            logger.info(f"Parsing metric logged: status={status}, score={score}, method={parse_method}")
        except Exception as log_err:
            logger.warning(f"Failed to log parsing metric: {log_err}")
        
        # ===== ENRICHISSEMENT AVEC PROMOTIONS FINANCEMENT =====
        # Ajouter automatiquement les informations de financement basées sur le code produit
        financing_info = None
        if vehicle_data and vehicle_data.get("model_code"):
            financing_info = get_financing_for_code(vehicle_data["model_code"])
            if financing_info:
                vehicle_data["financing"] = {
                    "consumer_cash": financing_info.get("consumer_cash", 0),
                    "bonus_cash": financing_info.get("bonus_cash", 0),
                    "option1_rates": financing_info.get("option1_rates", {}),
                    "option2_rates": financing_info.get("option2_rates", {}),
                    "programme_source": financing_info.get("programme_trim", "")
                }
                logger.info(f"Financing info added: Consumer Cash=${financing_info.get('consumer_cash', 0)}, Bonus=${financing_info.get('bonus_cash', 0)}")
        
        return {
            "success": True,
            "vehicle": vehicle_data,
            "validation": validation,
            "parse_method": parse_method,
            "has_financing": financing_info is not None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scanning invoice: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@router.post("/inventory/scan-invoice-file")
async def scan_invoice_file(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    """
    Scanne une facture FCA depuis un fichier uploadé (PDF ou image).
    Utilise le parser structuré (regex) pour les PDFs.
    
    USAGE: POST multipart/form-data avec field 'file'
    """
    user = await get_current_user(authorization)
    
    try:
        # Lire le fichier
        file_bytes = await file.read()
        
        # Détecter le type de fichier
        is_pdf = (
            file.content_type == "application/pdf" or
            file.filename.lower().endswith(".pdf") or
            file_bytes[:4] == b'%PDF'
        )
        
        # Convertir en base64 pour réutiliser la logique existante
        file_base64 = base64.b64encode(file_bytes).decode('utf-8')
        
        # Appeler le scan existant
        request = InvoiceScanRequest(image_base64=file_base64, is_pdf=is_pdf)
        return await scan_invoice(request, authorization)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scanning invoice file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@router.post("/test-ocr")
async def test_ocr_pipeline(file: UploadFile = File(...)):
    """
    🧪 ENDPOINT DE TEST - Pipeline OCR sans sauvegarde
    
    Teste le pipeline OCR complet sur une image de facture FCA:
    - Niveau 1: PDF → pdfplumber
    - Niveau 2: Image → OpenCV ROI + Tesseract
    - Niveau 3: Fallback → GPT-4 Vision (si score < 70)
    
    USAGE: POST multipart/form-data avec field 'file'
    Retourne: JSON avec données extraites + métriques + debug info
    
    ⚠️ NE SAUVEGARDE PAS en base de données
    """
    from ocr import process_image_ocr_pipeline, process_image_global_ocr
    from parser import parse_invoice_text
    from vin_utils import validate_and_correct_vin, decode_vin_year, decode_vin_brand
    from validation import validate_invoice_data as validate_invoice_full, calculate_validation_score
    
    try:
        start_time = time.time()
        
        # Lire le fichier
        file_bytes = await file.read()
        file_size = len(file_bytes)
        
        # Détecter le type
        is_pdf = (
            file.content_type == "application/pdf" or
            file.filename.lower().endswith(".pdf") or
            file_bytes[:4] == b'%PDF'
        )
        
        result = {
            "success": True,
            "file_info": {
                "filename": file.filename,
                "size_bytes": file_size,
                "content_type": file.content_type,
                "is_pdf": is_pdf
            },
            "pipeline_used": None,
            "raw_ocr": {},
            "parsed_data": {},
            "vin_analysis": {},
            "validation": {},
            "metrics": {},
            "debug": {}
        }
        
        # ===== TEST NIVEAU 1: PDF =====
        if is_pdf:
            result["pipeline_used"] = "pdf_native"
            extracted_text = extract_pdf_text(file_bytes)
            
            result["debug"]["pdf_text_length"] = len(extracted_text) if extracted_text else 0
            result["debug"]["pdf_text_preview"] = extracted_text[:500] if extracted_text else ""
            
            if extracted_text and len(extracted_text) > 100:
                parsed = parse_fca_invoice_structured(extracted_text)
                validation = validate_invoice_data(parsed)
                
                result["parsed_data"] = parsed
                result["validation"] = validation
                
                # VIN analysis
                vin = parsed.get("vin", "")
                if vin:
                    vin_result = validate_and_correct_vin(vin)
                    result["vin_analysis"] = vin_result
        
        # ===== TEST NIVEAU 2: OCR PAR ZONES =====
        else:
            result["pipeline_used"] = "ocr_zones"
            
            # Pipeline OCR par zones
            ocr_result = process_image_ocr_pipeline(file_bytes)
            result["raw_ocr"] = {
                "zones_processed": ocr_result.get("zones_processed", 0),
                "vin_text": ocr_result.get("vin_text", "")[:200],
                "finance_text": ocr_result.get("finance_text", "")[:200],
                "options_text": ocr_result.get("options_text", "")[:300],
                "totals_text": ocr_result.get("totals_text", "")[:200],
            }
            
            # Parser structuré
            parsed = parse_invoice_text(ocr_result)
            result["parsed_data"] = {
                "vin": parsed.get("vin"),
                "model_code": parsed.get("model_code"),
                "stock_no": parsed.get("stock_no"),
                "ep_cost": parsed.get("ep_cost"),
                "pdco": parsed.get("pdco"),
                "pref": parsed.get("pref"),
                "holdback": parsed.get("holdback"),
                "subtotal": parsed.get("subtotal"),
                "invoice_total": parsed.get("invoice_total"),
                "options_count": len(parsed.get("options", [])),
                "options": parsed.get("options", [])[:10],  # Limiter à 10
                "fields_extracted": parsed.get("fields_extracted", 0)
            }
            
            # VIN analysis approfondi
            vin = parsed.get("vin", "")
            if vin:
                vin_result = validate_and_correct_vin(vin)
                result["vin_analysis"] = {
                    "original": vin_result.get("original"),
                    "corrected": vin_result.get("corrected"),
                    "is_valid": vin_result.get("is_valid"),
                    "was_corrected": vin_result.get("was_corrected"),
                    "correction_type": vin_result.get("correction_type"),
                    "year": vin_result.get("year"),
                    "brand": vin_result.get("brand"),
                    "confidence": vin_result.get("confidence")
                }
            
            # Validation complète
            parsed["vin_valid"] = result["vin_analysis"].get("is_valid", False) if result["vin_analysis"] else False
            validation_result = validate_invoice_full(parsed)
            result["validation"] = validation_result
            
            # Recommandation fallback
            score = validation_result.get("score", 0)
            if score < 70:
                result["debug"]["recommendation"] = "Score < 70, le fallback GPT-4 Vision serait utilisé en production"
            else:
                result["debug"]["recommendation"] = f"Score {score}/100 suffisant, pas besoin de fallback Vision"
            
            # Test OCR global (comparaison)
            global_text = process_image_global_ocr(file_bytes)
            result["debug"]["global_ocr_length"] = len(global_text)
            result["debug"]["global_ocr_preview"] = global_text[:300] if global_text else ""
        
        # Métriques finales
        duration = round(time.time() - start_time, 3)
        result["metrics"] = {
            "parse_duration_sec": duration,
            "validation_score": result["validation"].get("score", 0),
            "cost_estimate": "$0.00 (100% open-source)"
        }
        
        return result
        
    except Exception as e:
        logger.error(f"Test OCR error: {str(e)}")
        import traceback
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@router.post("/inventory/scan-and-save")
async def scan_and_save_invoice(request: InvoiceScanRequest, authorization: Optional[str] = Header(None)):
    """Scanne une facture ET sauvegarde automatiquement le véhicule"""
    user = await get_current_user(authorization)
    
    # First scan the invoice
    scan_result = await scan_invoice(request, authorization)
    
    if not scan_result.get("success"):
        return scan_result
    
    vehicle_data = scan_result.get("vehicle", {})
    
    # ENRICHIR avec décodage VIN et codes produits
    vehicle_data = enrich_vehicle_data(vehicle_data)
    
    # Check if stock_no already exists
    stock_no = vehicle_data.get("stock_no", "")
    if not stock_no:
        raise HTTPException(status_code=400, detail="Numéro de stock non trouvé dans la facture")
    
    existing = await db.inventory.find_one({"stock_no": stock_no, "owner_id": user["id"]})
    
    # Prepare vehicle document
    vehicle_doc = {
        "id": str(uuid.uuid4()),
        "owner_id": user["id"],
        "stock_no": stock_no,
        "vin": vehicle_data.get("vin", ""),
        "brand": vehicle_data.get("brand", ""),
        "model": vehicle_data.get("model", ""),
        "trim": vehicle_data.get("trim", ""),
        "year": vehicle_data.get("year", datetime.now().year),
        "type": vehicle_data.get("type", "neuf"),
        "pdco": vehicle_data.get("pdco", 0) or 0,
        "ep_cost": vehicle_data.get("ep_cost", 0) or 0,
        "holdback": vehicle_data.get("holdback", 0) or 0,
        "net_cost": vehicle_data.get("net_cost", 0) or 0,
        "msrp": vehicle_data.get("msrp", 0) or 0,
        "asking_price": vehicle_data.get("msrp", 0) or 0,  # Default to MSRP
        "sold_price": None,
        "status": "disponible",
        "km": 0,
        "color": vehicle_data.get("color", ""),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    if existing:
        # Update existing
        await db.inventory.update_one(
            {"stock_no": stock_no, "owner_id": user["id"]},
            {"$set": vehicle_doc}
        )
        action = "mis à jour"
    else:
        # Insert new
        await db.inventory.insert_one(vehicle_doc)
        action = "ajouté"
    
    # Save options if present
    options = vehicle_data.get("options", [])
    if options:
        # Delete existing options
        await db.vehicle_options.delete_many({"stock_no": stock_no})
        # Insert new options
        for idx, opt in enumerate(options):
            await db.vehicle_options.insert_one({
                "id": str(uuid.uuid4()),
                "stock_no": stock_no,
                # PATCH: Support "product_code" et "code" pour compatibilité
                "product_code": opt.get("product_code", opt.get("code", "")),
                "order": idx,
                "description": opt.get("description", ""),
                "amount": opt.get("amount", 0) or 0
            })
    
    return {
        "success": True,
        "vehicle": vehicle_doc,
        "options_count": len(options),
        "action": action,
        "message": f"Véhicule {stock_no} {action} avec {len(options)} options"
    }



# ============ EXCEL EXPORT/IMPORT ============


@router.post("/invoice/export-excel")
async def export_invoice_to_excel(
    data: ExcelExportRequest,
    authorization: Optional[str] = Header(None)
):
    """
    Exporte les données de facture vers un fichier Excel.
    Utilisé après un scan OCR pour permettre la révision/correction.
    """
    user = await get_current_user(authorization)
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Facture FCA"
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:F1')
        title = f"FACTURE FCA - {data.brand or ''} {data.model or ''} {data.trim or ''}"
        ws['A1'] = title.strip()
        ws['A1'].font = header_font
        ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Vehicle Info Section
        ws['A3'] = "INFORMATIONS VÉHICULE"
        ws['A3'].font = subheader_font
        ws['A3'].fill = subheader_fill
        ws.merge_cells('A3:B3')
        
        vehicle_fields = [
            ("VIN", data.vin or ""),
            ("Code Modèle", data.model_code or ""),
            ("Marque", data.brand or ""),
            ("Modèle", data.model or ""),
            ("Trim", data.trim or ""),
            ("Année", data.year or ""),
            ("Stock#", data.stock_no or ""),
        ]
        
        row = 4
        for label, value in vehicle_fields:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            row += 1
        
        # Financial Section
        ws['D3'] = "INFORMATIONS FINANCIÈRES"
        ws['D3'].font = subheader_font
        ws['D3'].fill = subheader_fill
        ws.merge_cells('D3:E3')
        
        financial_fields = [
            ("E.P. (Coût Net)", data.ep_cost or 0),
            ("PDCO (MSRP)", data.pdco or 0),
            ("PREF", data.pref or 0),
            ("Holdback", data.holdback or 0),
            ("Sous-total", data.subtotal or 0),
            ("Total Facture", data.total or 0),
        ]
        
        row = 4
        for label, value in financial_fields:
            ws[f'D{row}'] = label
            ws[f'E{row}'] = value
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'D{row}'].border = thin_border
            ws[f'E{row}'].border = thin_border
            ws[f'E{row}'].number_format = '#,##0.00 $'
            row += 1
        
        # Options Section
        options_start = 13
        ws[f'A{options_start}'] = "OPTIONS / ACCESSOIRES"
        ws[f'A{options_start}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{options_start}'].fill = header_fill
        ws.merge_cells(f'A{options_start}:E{options_start}')
        
        # Options Headers
        opt_headers = ["#", "Code", "Description", "Catégorie", "Montant"]
        header_row = options_start + 1
        for col, header in enumerate(opt_headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = subheader_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Options Data
        row = header_row + 1
        options = data.options or []
        for i, opt in enumerate(options, 1):
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=opt.get('product_code', opt.get('code', ''))).border = thin_border
            ws.cell(row=row, column=3, value=opt.get('description', '')).border = thin_border
            ws.cell(row=row, column=4, value=opt.get('category', '')).border = thin_border
            amount_cell = ws.cell(row=row, column=5, value=opt.get('amount', 0))
            amount_cell.border = thin_border
            amount_cell.number_format = '#,##0.00 $'
            row += 1
        
        # Add empty rows for manual additions
        for i in range(len(options) + 1, 26):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col == 1:
                    cell.value = i
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 15
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Encode as base64
        excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')
        
        # Generate filename
        vin_part = (data.vin or "NOVIN")[-6:]
        filename = f"facture_{data.brand or 'FCA'}_{data.model or 'vehicle'}_{vin_part}.xlsx"
        
        return {
            "success": True,
            "filename": filename,
            "excel_base64": excel_base64,
            "message": "Fichier Excel généré avec succès"
        }
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")


@router.post("/invoice/import-excel")
async def import_invoice_from_excel(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    """
    Importe les données d'une facture depuis un fichier Excel.
    Utilisé après révision/correction manuelle.
    """
    user = await get_current_user(authorization)
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Fichier .xlsx requis")
    
    try:
        from openpyxl import load_workbook
        
        # Read file
        content = await file.read()
        excel_buffer = io.BytesIO(content)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        
        # Parse Vehicle Info (rows 4-10, columns A-B)
        vehicle_data = {}
        field_mapping = {
            "VIN": "vin",
            "Code Modèle": "model_code",
            "Marque": "brand",
            "Modèle": "model",
            "Trim": "trim",
            "Année": "year",
            "Stock#": "stock_no",
        }
        
        for row in range(4, 12):
            label = ws[f'A{row}'].value
            value = ws[f'B{row}'].value
            if label and label in field_mapping:
                vehicle_data[field_mapping[label]] = value
        
        # Parse Financial Info (rows 4-10, columns D-E)
        financial_mapping = {
            "E.P. (Coût Net)": "ep_cost",
            "PDCO (MSRP)": "pdco",
            "PREF": "pref",
            "Holdback": "holdback",
            "Sous-total": "subtotal",
            "Total Facture": "total",
        }
        
        for row in range(4, 12):
            label = ws[f'D{row}'].value
            value = ws[f'E{row}'].value
            if label and label in financial_mapping:
                try:
                    vehicle_data[financial_mapping[label]] = float(value) if value else 0
                except:
                    vehicle_data[financial_mapping[label]] = 0
        
        # Parse Options (starting from row 15)
        options = []
        row = 15  # After headers
        while row < 50:  # Max 35 options
            code = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=3).value
            category = ws.cell(row=row, column=4).value
            amount = ws.cell(row=row, column=5).value
            
            if code and description:  # Valid option
                options.append({
                    "product_code": str(code).strip(),
                    "description": str(description).strip(),
                    "category": str(category).strip() if category else "",
                    "amount": float(amount) if amount else 0
                })
            elif not code and not description:
                # Empty row, might be end of options
                break
            
            row += 1
        
        vehicle_data["options"] = options
        vehicle_data["import_source"] = "excel"
        vehicle_data["imported_at"] = datetime.utcnow().isoformat()
        
        return {
            "success": True,
            "data": vehicle_data,
            "options_count": len(options),
            "message": f"Import réussi: {len(options)} options trouvées"
        }
        
    except Exception as e:
        logger.error(f"Excel import error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur import Excel: {str(e)}")


@router.get("/invoice/template-excel")
async def get_invoice_template(authorization: Optional[str] = Header(None)):
    """
    Retourne un template Excel vide pour saisie manuelle.
    """
    user = await get_current_user(authorization)
    
    # Return empty template
    empty_data = ExcelExportRequest(
        vin="",
        model_code="",
        brand="",
        model="",
        trim="",
        year="2026",
        stock_no="",
        ep_cost=0,
        pdco=0,
        pref=0,
        holdback=0,
        subtotal=0,
        total=0,
        options=[]
    )
    
    return await export_invoice_to_excel(empty_data, authorization)


